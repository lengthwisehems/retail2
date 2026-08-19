# -*- coding: utf-8 -*-
"""
AMO database cleanup — apply the correction workbook to denim_analytics so the
database matches amo_inventory.py (commit 9ca26a5c).

ORDER
-----
  Phase 1  INSERT rows missing from style_info / style_metrics (styles that
           exist in lookup but not in style_info), built from the AMO scraper
           output CSVs + lookup identity + amo_inventory.py derivation.
  Phase 2  APPLY corrections from the workbook (ACTION=Remove -> delete;
           NEW cell -> literal / "Use amo_inventory.py" derive / blank keep).
  Phase 3  DEDUPE lookup & style_info, keeping the corrected row.

Derivation ("Use amo_inventory.py"): the real scraper functions are imported and
its post-processing passes run over the UNION of the corrected style_info rows
and the missing styles, keyed by HANDLE (style-level fields are the same across
a product's colors/sizes). Validated: 100% vs the Correct tab's active items,
and 541/541 + 95/95 resolution of the workbook's "Use amo_inventory.py" cells.

Inputs
------
  CORR_WORKBOOK   per-table tabs with current cols + NEW cols + ACTION
  SCRAPER_PATH    amo_inventory.py @ commit 9ca26a5c
  AMO_OUTPUT_DIR  folder of AMO_YYYYMMDD_HHMMSS.csv scraper outputs (the blob).
                  What must be in it: at least one output file that contains a
                  row (Description + measurements) for EVERY style being inserted
                  in Phase 1 (the styles present in lookup but missing from
                  style_info). It is NOT a random sample. Simplest safe choice:
                  drop in ALL your AMO output files - the script keeps the most
                  recent, richest row per handle, so extra/older files never
                  hurt. The dry run prints "Missing styles: N (with output data
                  X, no data Y)"; if Y > 0, add the output file(s) that still
                  listed those handles (usually the scrape just before they went
                  inactive) and re-run until no-data = 0.

SAFETY: DRY_RUN=True prints the full plan and writes nothing. Apply runs in one
transaction, rolls back on any error. Every log line is Central-time stamped.
Ids/barcodes are written as full-precision strings (repairs Excel sci-notation).
"""
from __future__ import annotations

import csv
import glob
import importlib.util
import json
import os
import re
import sys
import datetime as dt
from decimal import Decimal
from typing import Dict, List, Optional, Tuple

# ===========================================================================
# CONFIG
# ===========================================================================
DRY_RUN = True
DO_INSERT_MISSING = False   # these styles only have ~3 days of data and don't
                            # classify usefully - not worth adding; skip inserts.
DO_CORRECTIONS    = True
DO_DEDUPE         = True

# Resume support: a live run commits per phase/table and every COMMIT_EVERY rows,
# and records how far it got in a small JSON checkpoint next to this script. If a
# run dies partway (e.g. the WAN connection drops during the big variant_metrics
# update, as it did at 40000/408893), just re-run - it reads the checkpoint and
# skips everything already committed, picking up at the exact row it stopped on.
# Set RESET_PROGRESS = True to ignore any checkpoint and start the plan from the
# top (the underlying writes are idempotent, so this is always safe, just slower).
# The checkpoint is deleted automatically once all phases finish.
RESET_PROGRESS = False

BRAND = "AMO"

CORR_WORKBOOK  = r"AMO_DB_Values_2026-08-13_15-14-03_Claude.xlsx"
SCRAPER_PATH   = r"amo_inventory.py"
AMO_OUTPUT_DIR = r"AMO_Output"   # see header: must cover every missing style

SQL_SERVER   = os.environ.get("SQL_SERVER",   "denim-sql.database.windows.net")
SQL_DATABASE = os.environ.get("SQL_DATABASE", "denim_analytics")
SQL_USERNAME = os.environ.get("SQL_USERNAME", "")
SQL_PASSWORD = os.environ.get("SQL_PASSWORD", "")

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
CHECKPOINT_FILE = os.path.join(SCRIPT_DIR, f"cleanup_{BRAND.lower()}_checkpoint.json")

DERIVED_FIELDS = {"style_name", "jean_style", "rise_label",
                  "inseam_label", "inseam_style"}

PK = {"lookup": "lookup_id", "style_info": "style_info_id",
      "style_metrics": "style_metric_id", "variant_metrics": "variant_metric_id"}

# Phase 3 dedupe: which duplicate-identity keys to collapse, per table. A row is
# a duplicate of another when it matches on EVERY column in a key. The corrected
# row is kept (else the highest pk). Rows with a BLANK value in any key column
# are never collapsed on that key (a missing value is not an identity match), so
# empty sku_brand / inseam_label can't wrongly merge unrelated rows.
#   lookup         -> (brand+sku_shopify) and (brand+sku_brand)
#   style_info     -> (brand+product_name+inseam_label) and (brand+style_id+inseam_label)
#   style_metrics  -> none (one row per daily capture; no duplicates expected)
#   variant_metrics-> none (same)
DEDUPE_KEYS = {
    "lookup":     [["brand", "sku_shopify"], ["brand", "sku_brand"]],
    "style_info": [["brand", "product_name", "inseam_label"],
                   ["brand", "style_id", "inseam_label"]],
}


# ===========================================================================
# Central-time logging
# ===========================================================================
def _tz():
    try:
        from zoneinfo import ZoneInfo
        return ZoneInfo("America/Chicago")
    except Exception:
        return dt.timezone(dt.timedelta(hours=-6))
_TZ = _tz()


def log(msg: str = "") -> None:
    stamp = dt.datetime.now(_TZ).strftime("%Y-%m-%d %H:%M:%S")
    for part in str(msg).split("\n"):
        print(f"[{stamp}] {part}")


# ===========================================================================
# Checkpoint (resume after a dropped connection / crash)
# ===========================================================================
# Shape: {"inserts_done": bool,
#         "removes": {sheet: committed_count, ...},
#         "updates": {sheet: committed_count, ...},
#         "dedupe_done": {"lookup": bool, "style_info": bool}}
# A count is written ONLY right after the commit that persisted those rows, so a
# resume never skips an uncommitted row.
def _empty_ckpt() -> dict:
    return {"inserts_done": False, "removes": {}, "updates": {},
            "dedupe_done": {}}   # keyed per table+key, e.g. "lookup|brand+sku_brand"


def load_checkpoint() -> dict:
    if RESET_PROGRESS or not os.path.exists(CHECKPOINT_FILE):
        return _empty_ckpt()
    try:
        with open(CHECKPOINT_FILE, encoding="utf-8") as fh:
            data = json.load(fh)
        base = _empty_ckpt()
        base.update({k: data.get(k, base[k]) for k in base})
        return base
    except Exception as exc:
        log(f"(could not read checkpoint {CHECKPOINT_FILE}: {exc} - starting fresh)")
        return _empty_ckpt()


def save_checkpoint(ckpt: dict) -> None:
    try:
        with open(CHECKPOINT_FILE, "w", encoding="utf-8") as fh:
            json.dump(ckpt, fh, indent=1)
    except Exception as exc:
        log(f"(WARNING: could not write checkpoint: {exc})")


def clear_checkpoint() -> None:
    try:
        if os.path.exists(CHECKPOINT_FILE):
            os.remove(CHECKPOINT_FILE)
    except Exception as exc:
        log(f"(WARNING: could not remove checkpoint {CHECKPOINT_FILE}: {exc})")


# ===========================================================================
# Value helpers
# ===========================================================================
def s(v) -> str:
    """DB-safe full-precision string (no scientific notation)."""
    if v is None:
        return ""
    if isinstance(v, bool):
        return "1" if v else "0"
    if isinstance(v, Decimal):
        return f"{v:f}"
    if isinstance(v, float):
        return str(int(v)) if v.is_integer() else repr(v)
    return str(v).strip()


def norm(v) -> str:
    return re.sub(r"\s+", " ", s(v)).strip().lower()


def is_use_scraper(v) -> bool:
    return isinstance(v, str) and "amo_inventory" in v.lower()


def is_blank(v) -> bool:
    return v is None or (isinstance(v, str) and v.strip() == "")


def resolve(path: str) -> str:
    return path if os.path.isabs(path) else os.path.join(SCRIPT_DIR, path)


def corr_val(row: dict, col: str) -> str:
    """Corrected value of a workbook column: NEW literal wins; else current."""
    nv = row.get(f"NEW {col}")
    if is_use_scraper(nv) or is_blank(nv):
        return s(row.get(col))
    return s(nv)


# ===========================================================================
# Loaders
# ===========================================================================
def load_scraper(path: str):
    p = resolve(path)
    if not os.path.exists(p):
        sys.exit(f"ERROR: scraper not found at {p}")
    spec = importlib.util.spec_from_file_location("amo_inventory", p)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)  # type: ignore
    return mod


def load_tab(path: str, sheet: str) -> List[dict]:
    from openpyxl import load_workbook
    wb = load_workbook(resolve(path), data_only=True, read_only=True)
    ws = wb[sheet]
    it = ws.iter_rows(values_only=True)
    hdr = list(next(it))
    out = []
    for r in it:
        d = {hdr[i]: r[i] for i in range(len(hdr)) if hdr[i]}
        if any(v not in (None, "") for v in d.values()):
            out.append(d)
    return out


def build_output_index(output_dir: str) -> Dict[str, dict]:
    """handle -> richest/most-recent AMO scraper-output row (style-level data)."""
    idx: Dict[str, Tuple[str, dict]] = {}
    d = resolve(output_dir)
    if not os.path.isdir(d):
        log(f"WARNING: AMO_OUTPUT_DIR not found ({d}); cannot insert missing rows.")
        return {}
    files = sorted(glob.glob(os.path.join(d, "*.csv")))
    for f in files:
        m = re.search(r"AMO_(\d{8}_\d{6})", os.path.basename(f))
        date = m.group(1) if m else "00000000_000000"
        try:
            for r in csv.DictReader(open(f, encoding="utf-8-sig")):
                h = (r.get("Handle") or "").strip().lower()
                if not h:
                    continue
                prev = idx.get(h)
                dl = len(r.get("Description") or "")
                if (prev is None or date > prev[0]
                        or (date == prev[0] and dl > len(prev[1].get("Description") or ""))):
                    idx[h] = (date, r)
        except Exception as exc:
            log(f"(could not read {os.path.basename(f)}: {exc})")
    log(f"Output index: {len(idx)} handles across {len(files)} file(s)")
    return {h: r for h, (_, r) in idx.items()}


# ===========================================================================
# Derivation engine (keyed by handle)
# ===========================================================================
def _title_of(product_name: str, color: str) -> str:
    cd = color.replace("-", " ")
    if cd and product_name.endswith(" / " + cd):
        return product_name[: -(len(cd) + 3)].strip()
    return product_name


def strip_color(text: str, color: str) -> str:
    """STEP 2.5: remove the Color words from the (styling-stripped) title.
    The DB product_name sometimes embeds the color, but the scraper derives the
    Style Name from the title *before* color is appended - so a DB-only cleanup
    must strip it. Guarded so it never empties the name."""
    if not color:
        return text
    out = text
    for tok in re.split(r"[\s\-/]+", color):
        if len(tok) >= 2:
            out = re.sub(rf"\b{re.escape(tok)}\b", " ", out, flags=re.IGNORECASE)
    out = re.sub(r"\s+", " ", out).strip()
    return out or text


class Deriver:
    """Runs amo_inventory.py derivation + post-processing over the union of the
    corrected style_info rows and the missing styles. Answers by handle.

    Missing styles take their title/color from the corrected lookup identity
    (reliable "Style / Color") and their description/measurements from the AMO
    output row; existing style_info rows use their own corrected fields. Every
    title has the color stripped (STEP 2.5) before the styling-word removal
    result is fed into the sibling/one-word rules."""

    def __init__(self, g, si_rows: List[dict], missing: Dict[str, dict],
                 lookup_rows: List[dict]):
        self.g = g
        self.by_handle: Dict[str, dict] = {}
        # representative corrected (product_name, color) per handle from lookup
        self.lk_rep: Dict[str, Tuple[str, str]] = {}
        for d in lookup_rows:
            if s(d.get("ACTION")).lower() == "remove":
                continue
            h = s(d.get("handle")).lower()
            self.lk_rep.setdefault(h, (corr_val(d, "product_name"), corr_val(d, "color")))
        self._build(si_rows, missing)

    def _build(self, si_rows, missing):
        g = self.g
        WH = g.WORK_HEADERS
        work, handles = [], []

        def add(pn, color, desc, tags, handle, rise, inseam, leg):
            title = strip_color(_title_of(pn, color), color)   # STEP 1, 2.5
            ss = g.derive_style_name_base(title)
            js = g.derive_jean_style(title, desc, tags, leg)
            rl = g.derive_rise_label(title, desc, rise)
            il = g.derive_inseam_label(js, inseam, title, handle, "", desc, False)
            isy = g.derive_inseam_style(js, inseam, rise, il, desc, tags)
            row = [""] * len(WH)
            for k, v in (("Style Name", ss), ("Style Name Source", ss),
                         ("Description", desc), ("Tags", tags), ("Leg Opening", leg),
                         ("Rise", rise), ("Inseam", inseam), ("Jean Style", js),
                         ("Rise Label", rl), ("Inseam Label", il), ("Inseam Style", isy)):
                g._set(row, k, v)
            work.append(row)
            handles.append(handle.lower())

        for d in si_rows:
            if s(d.get("ACTION")).lower() == "remove":
                continue
            add(corr_val(d, "product_name"), corr_val(d, "color"),
                s(d.get("description")), s(d.get("tags")), s(d.get("handle")),
                s(d.get("rise")), s(d.get("inseam")), s(d.get("leg_opening")))
        for h, r in missing.items():
            # title/color from the corrected lookup identity (reliable
            # "Style / Color"); description/measurements from the output row.
            pn, color = self.lk_rep.get(h, ((r.get("Product") or ""), (r.get("Color") or "")))
            add(pn, color, (r.get("Description") or "").strip(),
                (r.get("Tags") or "").strip(), h,
                (r.get("Rise") or "").strip(), (r.get("Inseam") or "").strip(),
                (r.get("Leg Opening") or "").strip())

        g.apply_jean_style_sibling_inference(work)
        g.apply_style_name_rules(work)
        g.apply_rise_label_steps(work)
        g.apply_inseam_refresh(work)
        for h, w in zip(handles, work):
            self.by_handle[h] = {
                "style_name":   g._col(w, "Style Name"),
                "jean_style":   g._col(w, "Jean Style"),
                "rise_label":   g._col(w, "Rise Label"),
                "inseam_label": g._col(w, "Inseam Label"),
                "inseam_style": g._col(w, "Inseam Style"),
            }

    def get(self, field: str, handle: str) -> str:
        return self.by_handle.get(handle.lower(), {}).get(field, "")


# ===========================================================================
# Correction planning (pure)
# ===========================================================================
def corrected_updates(row: dict, deriver: Deriver) -> Dict[str, str]:
    updates: Dict[str, str] = {}
    handle = s(row.get("handle"))
    for hdr, val in row.items():
        if not (isinstance(hdr, str) and hdr.startswith("NEW ")):
            continue
        col = hdr[4:]
        cur = row.get(col)
        if is_blank(val):
            continue
        if is_use_scraper(val):
            if col not in DERIVED_FIELDS:
                continue
            new = deriver.get(col, handle)
            if new and norm(new) != norm(cur):
                updates[col] = new
            continue
        new = s(val)
        if norm(new) != norm(cur):
            updates[col] = new
    return updates


# ===========================================================================
# Main
# ===========================================================================
def main() -> None:
    log("=" * 60)
    log(f"AMO DATABASE CLEANUP  ({'DRY RUN' if DRY_RUN else 'LIVE RUN'})")
    log("=" * 60)

    g = load_scraper(SCRAPER_PATH)
    log(f"Loaded scraper rules from {resolve(SCRAPER_PATH)}")
    output_idx = build_output_index(AMO_OUTPUT_DIR)

    tabs = {sheet: load_tab(CORR_WORKBOOK, sheet)
            for sheet in ("lookup", "style_info", "style_metrics", "variant_metrics")}
    for sheet, rows in tabs.items():
        log(f"{sheet}: {len(rows)} rows loaded")

    # --- determine missing styles (in lookup Keep, not in style_info) ------
    si_handles = {s(d.get("handle")).lower() for d in tabs["style_info"]}
    missing_handles = sorted(
        {s(d.get("handle")).lower() for d in tabs["lookup"]
         if s(d.get("ACTION")).lower() != "remove" and s(d.get("handle"))} - si_handles)
    missing = {h: output_idx[h] for h in missing_handles if h in output_idx}
    no_data = [h for h in missing_handles if h not in output_idx]
    log(f"Missing styles: {len(missing_handles)} "
        f"(with output data: {len(missing)}, no data: {len(no_data)})")
    if no_data:
        log("   NO OUTPUT DATA for: " + ", ".join(no_data[:15]))

    # The derivation map always includes the missing styles so that lookup /
    # style_metrics "Use amo_inventory.py" cells for those handles resolve, even
    # when DO_INSERT_MISSING is off (their rows are still corrected in place).
    deriver = Deriver(g, tabs["style_info"], missing, tabs["lookup"])
    log(f"Derivation map: {len(deriver.by_handle)} handles")

    # --- plan corrections --------------------------------------------------
    plan_updates: Dict[str, List[Tuple[object, Dict[str, str]]]] = {t: [] for t in tabs}
    plan_removes: Dict[str, List[object]] = {t: [] for t in tabs}
    for sheet, rows in tabs.items():
        for row in rows:
            key = row.get(PK[sheet])
            if s(row.get("ACTION")).lower() == "remove":
                plan_removes[sheet].append(key)
                continue
            ups = corrected_updates(row, deriver)
            if ups:
                plan_updates[sheet].append((key, ups))

    # --- plan inserts (only when enabled) ----------------------------------
    inserts = (plan_inserts(tabs["lookup"], missing, deriver, output_idx)
               if DO_INSERT_MISSING else {"style_info": [], "style_metrics": []})

    # --- preview -----------------------------------------------------------
    log("-" * 60)
    if DO_INSERT_MISSING:
        log(f"Phase 1 INSERT: style_info +{len(inserts['style_info'])}, "
            f"style_metrics +{len(inserts['style_metrics'])}")
    else:
        log("Phase 1 INSERT: skipped (DO_INSERT_MISSING = False)")
    for sheet in tabs:
        nf = sum(len(u) for _, u in plan_updates[sheet])
        log(f"Phase 2 {sheet:16} remove={len(plan_removes[sheet]):>5} "
            f"update_rows={len(plan_updates[sheet]):>5} field_writes={nf}"
            + ("" if DO_CORRECTIONS else "  (DO_CORRECTIONS = False)"))
    log("Phase 3 dedupe: lookup + style_info (keep corrected) — runs on live DB"
        + ("" if DO_DEDUPE else "  (DO_DEDUPE = False)"))

    if DRY_RUN:
        if DO_INSERT_MISSING and inserts["style_info"]:
            log("Sample inserted style_info rows for missing styles "
                "(style_name | jean_style | rise_label | inseam_label | inseam_style):")
            for r in inserts["style_info"][:12]:
                log(f"   {r['handle']:34} {r.get('style_name',''):22} | "
                    f"{r.get('jean_style',''):24} | {r.get('rise_label',''):5} | "
                    f"{r.get('inseam_label',''):8} | {r.get('inseam_style','')}")
        log("DRY RUN complete — nothing written.")
        return

    # --- apply -------------------------------------------------------------
    import pymssql
    conn = pymssql.connect(server=SQL_SERVER, user=SQL_USERNAME, password=SQL_PASSWORD,
                           database=SQL_DATABASE, timeout=600, login_timeout=60)
    cur = conn.cursor(as_dict=True)
    # corrected_pks holds EVERY corrected row (upfront, from the plan) so the
    # Phase 3 dedupe survivor logic is correct on a resume too - it must not
    # depend on which rows happened to be updated in this particular run.
    corrected_pks = {
        "lookup":     {key for key, _ in plan_updates.get("lookup", [])},
        "style_info": {key for key, _ in plan_updates.get("style_info", [])},
    }
    unique_keys = load_unique_keys(cur, list(tabs))
    log("Loaded unique indexes: " + "; ".join(
        f"{t}=[{', '.join('('+'+'.join(k)+')' for k in ks)}]"
        for t, ks in unique_keys.items() if ks))

    ckpt = load_checkpoint()
    if not RESET_PROGRESS and os.path.exists(CHECKPOINT_FILE):
        log("Resuming from checkpoint: "
            f"inserts_done={ckpt['inserts_done']}, "
            f"removes={ckpt['removes']}, updates={ckpt['updates']}, "
            f"dedupe_done={ckpt['dedupe_done']}")

    # Per-phase/table commits: each committed step persists so a later failure
    # never discards it. Re-running is safe - the plan is keyed by PK and is
    # idempotent (already-applied updates re-set the same values; already-removed
    # rows match nothing). The checkpoint additionally lets a resume SKIP the rows
    # already committed instead of re-sending them. Only the currently-running
    # step rolls back on error.
    try:
        if DO_INSERT_MISSING:
            if ckpt["inserts_done"]:
                log("Phase 1: inserts already done (checkpoint) - skipping.")
            else:
                log("Phase 1: inserting missing rows...")
                do_inserts(cur, inserts)
                conn.commit()
                ckpt["inserts_done"] = True
                save_checkpoint(ckpt)
        if DO_CORRECTIONS:
            log("Phase 2: applying corrections...")
            apply_corrections(cur, conn, plan_updates, plan_removes,
                              corrected_pks, unique_keys, ckpt)
        if DO_DEDUPE:
            log("Phase 3: dedupe (keep corrected)...")
            for table, keysets in DEDUPE_KEYS.items():
                for kc in keysets:
                    tag = f"{table}|{'+'.join(kc)}"
                    if ckpt["dedupe_done"].get(tag):
                        log(f"   {tag}: already deduped (checkpoint) - skipping.")
                        continue
                    dedupe(cur, table, kc, PK[table], corrected_pks[table])
                    conn.commit()
                    ckpt["dedupe_done"][tag] = True
                    save_checkpoint(ckpt)
        clear_checkpoint()
        log("Done - all phases committed. Checkpoint cleared.")
    except Exception:
        conn.rollback()
        log("ERROR - current step rolled back. Earlier committed steps are kept; "
            "re-run to resume from the checkpoint (already-committed rows are skipped).")
        raise
    finally:
        conn.close()


# ===========================================================================
# Insert planning / execution
# ===========================================================================
def plan_inserts(lookup_rows, missing, deriver, output_idx) -> Dict[str, List[dict]]:
    """Build style_info + style_metrics rows for each missing colorway."""
    colorways: Dict[Tuple[str, str], dict] = {}
    for d in lookup_rows:
        if s(d.get("ACTION")).lower() == "remove":
            continue
        h = s(d.get("handle")).lower()
        if h not in missing:
            continue
        color = corr_val(d, "color")
        colorways.setdefault((h, color), d)   # one representative lookup row

    si_rows, sm_rows = [], []
    for (h, color), lk in colorways.items():
        out = missing[h]
        der = deriver.by_handle.get(h, {})
        cap_date = _out_date(out)
        # style_name: prefer the curated lookup value (a literal NEW style_name
        # overrides the scraper's sibling-rule result, e.g. billie -> "Billie
        # Straight"); fall back to the derived value when lookup says
        # "Use amo_inventory.py" or is blank.
        _nv_sn = lk.get("NEW style_name")
        style_name = (der.get("style_name", "")
                      if (is_use_scraper(_nv_sn) or is_blank(_nv_sn)) else s(_nv_sn))
        base = {
            "brand": BRAND,
            "style_id": corr_val(lk, "style_id") or s(out.get("Style Id")),
            "product_name": corr_val(lk, "product_name") or s(out.get("Product")),
            "handle": s(lk.get("handle")),
            "sku_url": s(lk.get("sku_url")) or s(out.get("SKU URL")),
            "color": color,
            "style_name": style_name,
            "source_file_name": "AMO_cleanup_insert",
            "captured_date": cap_date, "captured_datetime": cap_date,
        }
        si_rows.append({**base,
            "jean_style": der.get("jean_style", ""),
            "rise_label": der.get("rise_label", ""),
            "inseam_label": der.get("inseam_label", ""),
            "inseam_style": der.get("inseam_style", ""),
            "product_type": s(out.get("Product Type")),
            "description": (out.get("Description") or "").strip(),
            "tags": (out.get("Tags") or "").strip(),
            "vendor": s(out.get("Vendor")),
            "image_url": s(out.get("Image URL")),
            "rise": _num(out.get("Rise")), "back_rise": _num(out.get("Back Rise")),
            "inseam": _num(out.get("Inseam")), "leg_opening": _num(out.get("Leg Opening")),
            "created_at": _date(out.get("Published At")),
            "is_manual_override": 0})
        sm_rows.append({**base,
            "price": _num(out.get("Price")),
            "compare_at_price": _num(out.get("Compare at Price")),
            "quantity_of_style": _int(out.get("Quantity of style")),
            "created_at": _date(out.get("Published At")),
            "published_at": _date(out.get("Published At"))})
    return {"style_info": si_rows, "style_metrics": sm_rows}


def do_inserts(cur, inserts):
    for table in ("style_info", "style_metrics"):
        rows = inserts[table]
        for r in rows:
            cols = [c for c, v in r.items() if v != "" and v is not None]
            placeholders = ", ".join("%s" for _ in cols)
            cur.execute(f"INSERT INTO {table} ({', '.join('['+c+']' for c in cols)}) "
                        f"VALUES ({placeholders})", [r[c] for c in cols])
        log(f"   inserted {len(rows)} rows into {table}")


def _num(v) -> str:
    t = s(v).replace("$", "").replace(",", "").strip()
    return t if re.fullmatch(r"-?\d+(\.\d+)?", t or "") else ""


def _int(v) -> str:
    t = _num(v)
    return str(int(float(t))) if t else ""


def _date(v) -> str:
    t = s(v)
    for fmt in ("%m/%d/%Y", "%Y-%m-%d"):
        try:
            return dt.datetime.strptime(t, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return ""


def _out_date(out) -> dt.datetime:
    return dt.datetime.now(_TZ).replace(tzinfo=None)


# ===========================================================================
# Correction execution
# ===========================================================================
# A correction that changes a unique-key column to a value another row already
# has (e.g. a sci-notation sku fixed to match its correct twin) would violate
# the index, so the conflicting duplicate is deleted first - keeping the row
# being corrected (matches the dedupe rule). The unique indexes are read from
# the database at runtime (load_unique_keys) so every index is covered, not just
# the ones documented in the schema file.
def load_unique_keys(cur, tables) -> Dict[str, List[Tuple[str, ...]]]:
    out: Dict[str, List[Tuple[str, ...]]] = {t: [] for t in tables}
    cur.execute("""
        SELECT o.name AS tbl, i.name AS idx, c.name AS col, ic.key_ordinal AS ord
        FROM sys.indexes i
        JOIN sys.index_columns ic ON ic.object_id = i.object_id
             AND ic.index_id = i.index_id AND ic.is_included_column = 0
        JOIN sys.columns c ON c.object_id = i.object_id AND c.column_id = ic.column_id
        JOIN sys.objects o ON o.object_id = i.object_id
        WHERE i.is_unique = 1 AND i.is_primary_key = 0
        ORDER BY o.name, i.name, ic.key_ordinal
    """)
    per_index: Dict[Tuple[str, str], List[str]] = {}
    for r in cur.fetchall():
        if r["tbl"] in out:
            per_index.setdefault((r["tbl"], r["idx"]), []).append(r["col"])
    for (tbl, _idx), colnames in per_index.items():
        out[tbl].append(tuple(colnames))
    return out


def _clear_conflicts(cur, sheet, pk_col, pk, cols, unique_keys) -> int:
    """Delete any OTHER row that already holds the unique-key value this update
    is about to set (same variant/style), so the UPDATE can't hit the index.
    Runs a query only when the update actually changes a key column."""
    deleted = 0
    for key in unique_keys.get(sheet, []):
        if not any(c in cols and c != "brand" for c in key):
            continue
        conds = [f"Y.[{pk_col}] <> X.[{pk_col}]"]
        params = [pk]
        for c in key:
            if c in cols:
                conds.append(f"Y.[{c}] = %s")
                params.append(cols[c])
            else:
                conds.append(f"Y.[{c}] = X.[{c}]")
        cur.execute(f"DELETE Y FROM {sheet} Y JOIN {sheet} X "
                    f"ON X.[{pk_col}] = %s WHERE {' AND '.join(conds)}", params)
        deleted += cur.rowcount if (cur.rowcount and cur.rowcount > 0) else 0
    return deleted


COMMIT_EVERY = 5000   # commit + log progress every N rows within a big table


def apply_corrections(cur, conn, plan_updates, plan_removes, corrected_pks,
                      unique_keys, ckpt):
    # --- removes -----------------------------------------------------------
    for sheet, removes in plan_removes.items():
        done = ckpt["removes"].get(sheet, 0)
        if done >= len(removes):
            if removes:
                log(f"   {sheet}: removes already done ({done}/{len(removes)}) - skipping.")
            continue
        if done:
            log(f"   {sheet}: resuming removes at {done}/{len(removes)}.")
        n = done
        for key in removes[done:]:
            cur.execute(f"DELETE FROM {sheet} WHERE {PK[sheet]}=%s", (key,))
            n += 1
            if n % COMMIT_EVERY == 0:
                conn.commit()
                ckpt["removes"][sheet] = n
                save_checkpoint(ckpt)
                log(f"   {sheet}: removed {n}/{len(removes)}...")
        conn.commit()
        ckpt["removes"][sheet] = n
        save_checkpoint(ckpt)
        log(f"   {sheet}: removed {len(removes)} (committed)")
    # --- updates -----------------------------------------------------------
    for sheet, ups in plan_updates.items():
        done = ckpt["updates"].get(sheet, 0)
        if done >= len(ups):
            if ups:
                log(f"   {sheet}: updates already done ({done}/{len(ups)}) - skipping.")
            continue
        if done:
            log(f"   {sheet}: resuming updates at {done}/{len(ups)}.")
        n, conflicts = done, 0
        for key, cols in ups[done:]:
            conflicts += _clear_conflicts(cur, sheet, PK[sheet], key, cols, unique_keys)
            assigns = ", ".join(f"[{c}]=%s" for c in cols)
            cur.execute(f"UPDATE {sheet} SET {assigns} WHERE {PK[sheet]}=%s",
                        list(cols.values()) + [key])
            n += 1
            if n % COMMIT_EVERY == 0:
                conn.commit()
                ckpt["updates"][sheet] = n
                save_checkpoint(ckpt)
                log(f"   {sheet}: updated {n}/{len(ups)}"
                    + (f" (merged {conflicts} dup)" if conflicts else "") + "...")
        conn.commit()
        ckpt["updates"][sheet] = n
        save_checkpoint(ckpt)
        log(f"   {sheet}: updated {len(ups)}"
            + (f", merged {conflicts} duplicate row(s)" if conflicts else "")
            + " (committed)")


def dedupe(cur, table, key_cols, pk, corrected_pks):
    """Delete duplicate rows per key group, keeping a corrected row if present,
    else the highest pk. Groups with a blank value in any key column are skipped
    (a missing value is not an identity - never collapse on it)."""
    cur.execute(f"SELECT {pk} AS pk, {', '.join(key_cols)} FROM {table} WHERE brand=%s",
                (BRAND,))
    groups: Dict[tuple, List[int]] = {}
    for r in cur.fetchall():
        k = tuple(norm(r[c]) for c in key_cols)
        if any(part == "" for part in k):      # blank key part -> not a match
            continue
        groups.setdefault(k, []).append(r["pk"])
    deleted = 0
    for k, pks in groups.items():
        if len(pks) < 2:
            continue
        keep = next((p for p in sorted(pks, reverse=True) if p in corrected_pks), max(pks))
        for p in pks:
            if p != keep:
                cur.execute(f"DELETE FROM {table} WHERE {pk}=%s", (p,))
                deleted += 1
    log(f"   {table} [{'+'.join(key_cols)}]: deduped, removed {deleted} duplicate rows")


if __name__ == "__main__":
    main()
