"""Explore Self Contrast's Globo preorder and Gravity variant data."""
from __future__ import annotations

import json
import logging
import re
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Sequence, Tuple

import requests
import urllib3
from bs4 import BeautifulSoup
from openpyxl import Workbook
from requests.adapters import HTTPAdapter, Retry

BASE_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = BASE_DIR / "Output"
LOG_PATH = BASE_DIR / "selfcontrast_probe_run.log"
FALLBACK_LOG_PATH = OUTPUT_DIR / "selfcontrast_probe_run.log"

COLLECTION_URLS: Sequence[str] = (
    "https://www.selfcontrast.com/collections/denim-2",
)
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
)
REQUEST_TIMEOUT = 40
STOREFRONT_TIMEOUT = 20

STOREFRONT_ENDPOINTS: Sequence[str] = (
    "https://self-contrast.myshopify.com/api/2025-10/graphql.json",
    "https://self-contrast.myshopify.com/api/2024-01/graphql.json",
    "https://self-contrast.myshopify.com/api/2025-01/graphql.json",
    "https://self-contrast.myshopify.com/api/2025-07/graphql.json",
    "https://self-contrast.myshopify.com/api/2025-04/graphql.json",
    "https://self-contrast.myshopify.com/api/unstable/graphql.json",
    "https://self-contrast.myshopify.com/api/2024-04/graphql.json",
    "https://self-contrast.myshopify.com/api/2023-01/graphql.json",
    "https://self-contrast.myshopify.com/api/2023-04/graphql.json",
)

STOREFRONT_COLLECTION_HANDLES: Sequence[str] = (
    "denim-2",
)

GRAPHQL_PAGE_SIZE = 100
TRANSIENT_STATUS = {429, 500, 502, 503, 504}

STOREFRONT_QUERY = """
query CollectionProducts($handle: String!, $cursor: String, $pageSize: Int!) {
  collection(handle: $handle) {
    id
    handle
    title
    products(first: $pageSize, after: $cursor) {
      edges {
        cursor
        node {
          id
          handle
          title
          productType
          tags
          vendor
          description
          descriptionHtml
          onlineStoreUrl
          publishedAt
          createdAt
          updatedAt
          availableForSale
          options { name values }
          featuredImage { url altText }
          images(first: 20) { edges { node { url altText } } }
          variants(first: 250) {
            edges {
              cursor
              node {
                id
                title
                sku
                barcode
                availableForSale
                price { amount currencyCode }
                compareAtPrice { amount currencyCode }
                selectedOptions { name value }
                image { url altText }
              }
            }
          }
        }
      }
      pageInfo { hasNextPage endCursor }
    }
  }
}
"""

OUTPUT_DIR.mkdir(exist_ok=True)


def configure_logging() -> logging.Logger:
    logger = logging.getLogger("selfcontrast_probe")
    if logger.handlers:
        return logger

    logger.setLevel(logging.INFO)

    handler: logging.Handler
    try:
        handler = logging.FileHandler(LOG_PATH, encoding="utf-8")
    except OSError:
        handler = logging.FileHandler(FALLBACK_LOG_PATH, encoding="utf-8")
        stream_handler = logging.StreamHandler()
        stream_handler.setLevel(logging.INFO)
        logger.addHandler(stream_handler)
        logger.warning(
            "Falling back to %s for logging", FALLBACK_LOG_PATH.as_posix()
        )
    else:
        stream_handler = logging.StreamHandler()
        stream_handler.setLevel(logging.INFO)
        logger.addHandler(stream_handler)

    handler.setLevel(logging.INFO)
    formatter = logging.Formatter(
        "%(asctime)s - %(levelname)s - %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
    handler.setFormatter(formatter)
    logger.addHandler(handler)

    return logger


def build_session() -> requests.Session:
    session = requests.Session()
    retries = Retry(
        total=5,
        backoff_factor=1.2,
        status_forcelist=(429, 500, 502, 503, 504),
        allowed_methods=("GET", "HEAD"),
    )
    adapter = HTTPAdapter(max_retries=retries)
    session.mount("https://", adapter)
    session.mount("http://", adapter)
    session.headers.update({"User-Agent": USER_AGENT})
    session.verify = False
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    return session


def fetch_collection_html(session: requests.Session, logger: logging.Logger) -> str:
    html_chunks: List[str] = []
    for url in COLLECTION_URLS:
        logger.info("Fetching collection HTML from %s", url)
        response = session.get(url, timeout=REQUEST_TIMEOUT, verify=False)
        response.raise_for_status()
        logger.info("Fetched %s bytes of HTML", len(response.text))
        html_chunks.append(response.text)
    return "\n".join(html_chunks)


def _consume_bracket_block(text: str, start: int, open_char: str = "[") -> Tuple[str, int]:
    close_char = "]" if open_char == "[" else "}"
    depth = 0
    for index in range(start, len(text)):
        char = text[index]
        if char == open_char:
            depth += 1
        elif char == close_char:
            depth -= 1
            if depth == 0:
                return text[start : index + 1], index + 1
    raise ValueError("Unbalanced brackets encountered while parsing script payload")


def extract_globo_script(html: str, logger: logging.Logger) -> str:
    soup = BeautifulSoup(html, "html.parser")
    for script in soup.find_all("script"):
        if script.string and "GloboPreorderParams" in script.string:
            logger.info("Located Globo script block")
            return script.string
    raise RuntimeError("Unable to locate window.GloboPreorderParams script in HTML")


def parse_globo_product_arrays(script_text: str, logger: logging.Logger) -> List[Dict[str, Any]]:
    products: List[Dict[str, Any]] = []
    cursor = 0
    needle = ".concat("
    while True:
        idx = script_text.find(needle, cursor)
        if idx == -1:
            break
        array_start = script_text.find("[", idx)
        if array_start == -1:
            logger.warning("Found concat call without array payload at index %s", idx)
            break
        array_text, cursor = _consume_bracket_block(script_text, array_start, "[")
        try:
            chunk = json.loads(array_text)
            products.extend(chunk)
            logger.info("Parsed %s Globo product entries", len(chunk))
        except json.JSONDecodeError as exc:
            logger.error("Failed to decode Globo product array: %s", exc)
    logger.info("Total Globo products parsed: %s", len(products))
    return products


def parse_preorder_settings(script_text: str, logger: logging.Logger) -> Dict[str, Any]:
    pattern = "window.GloboPreorderParams.preorderSettings = "
    cursor = 0
    settings: Dict[str, Any] = {}
    while True:
        idx = script_text.find(pattern, cursor)
        if idx == -1:
            break
        start = idx + len(pattern)
        if start >= len(script_text) or script_text[start] != "{":
            cursor = start
            continue
        block, cursor = _consume_bracket_block(script_text, start, "{")
        try:
            candidate = json.loads(block)
        except json.JSONDecodeError as exc:
            logger.error("Failed to parse preorderSettings block: %s", exc)
            continue
        if candidate:
            settings = candidate
    logger.info(
        "Parsed preorderSettings: %s product mappings", len(settings.get("products", {}))
    )
    return settings


def parse_preorder_profiles(script_text: str, logger: logging.Logger) -> Dict[str, Any]:
    profile_pattern = re.compile(
        r"Object\.assign\(window\.GloboPreorderParams\.preorderSettings\.profiles,\s*(\{.*?\})\)",
        re.DOTALL,
    )
    profiles: Dict[str, Any] = {}
    for match in profile_pattern.finditer(script_text):
        blob = match.group(1)
        try:
            data = json.loads(blob)
            profiles.update(data)
        except json.JSONDecodeError as exc:
            logger.error("Failed to parse preorder profile blob: %s", exc)
    logger.info("Parsed %s preorder profiles", len(profiles))
    return profiles


def normalize_value(value: Any) -> Any:
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value


def build_globo_variant_rows(
    products: Iterable[Dict[str, Any]],
    settings: Dict[str, Any],
    profiles: Dict[str, Any],
) -> List[Dict[str, Any]]:
    product_defaults: Dict[str, str] = {}
    variant_profiles: Dict[str, str] = {}
    for product_id, variant_map in settings.get("products", {}).items():
        for variant_id, profile_id in variant_map.items():
            if variant_id:
                variant_profiles[str(variant_id)] = str(profile_id)
            else:
                product_defaults[str(product_id)] = str(profile_id)

    rows: List[Dict[str, Any]] = []
    for product in products:
        base = {
            f"product_{key}": normalize_value(value)
            for key, value in product.items()
            if key != "variants"
        }
        product_id = str(product.get("id", ""))
        variants = product.get("variants", []) or []
        for variant in variants:
            variant_id = str(variant.get("id", ""))
            profile_id = variant_profiles.get(variant_id) or product_defaults.get(product_id)
            profile_data = profiles.get(profile_id) if profile_id else None

            row = base.copy()
            row.update({f"variant_{key}": normalize_value(value) for key, value in variant.items()})
            row["variant_id_str"] = variant_id
            row["profile_id"] = profile_id
            if profile_data:
                row["profile_name"] = profile_data.get("name")
                row["profile_message"] = profile_data.get("message")
                row["profile_raw"] = json.dumps(profile_data, ensure_ascii=False)
            rows.append(row)
    return rows


def parse_gravity_variants(html: str, logger: logging.Logger) -> List[Dict[str, Any]]:
    div_pattern = re.compile(
        r"(<div[^>]*class=\"[^\"]*gsProductVariants[^\"]*\"[^>]*>)(.*?)</div>",
        re.DOTALL | re.IGNORECASE,
    )
    rows: Dict[str, Dict[str, Any]] = {}
    matches = list(div_pattern.finditer(html))
    logger.info("Found %s Gravity variant containers", len(matches))
    for match in matches:
        start_tag, inner_html = match.groups()
        if "vquantity" not in inner_html:
            continue
        snippet = f"{start_tag}{inner_html}</div>"
        soup = BeautifulSoup(snippet, "html.parser")
        div = soup.div
        handle = div.get("gsproducthandler")
        for element in div.find_all(["p", "span"]):
            attrs = dict(element.attrs)
            variant_id = (
                attrs.get("variantpriceid")
                or attrs.get("varianttitleid")
                or attrs.get("variantid")
            )
            if not variant_id:
                continue
            row = rows.setdefault(variant_id, {"product_handle": handle, "variant_id": variant_id})
            if element.name == "p" and "vquantity" in attrs:
                prefix = "inventory_"
            elif element.name == "p":
                prefix = "details_"
            else:
                prefix = "image_"
            for key, value in attrs.items():
                row[f"{prefix}{key}"] = value
    logger.info("Parsed %s Gravity variant rows", len(rows))
    return list(rows.values())


class GraphQLRequestError(RuntimeError):
    """Raised when the Storefront API repeatedly fails."""


def iter_graphql_endpoints(preferred: str | None = None) -> List[str]:
    ordered = list(dict.fromkeys(STOREFRONT_ENDPOINTS))
    if preferred and preferred in ordered:
        ordered.remove(preferred)
        ordered.insert(0, preferred)
    return ordered


def execute_graphql(
    session: requests.Session,
    payload: Dict[str, Any],
    endpoints: Sequence[str],
    logger: logging.Logger,
) -> Dict[str, Any]:
    headers = {"Content-Type": "application/json"}
    last_error: Exception | None = None

    for attempt, endpoint in enumerate(endpoints):
        try:
            response = session.post(
                endpoint,
                json=payload,
                headers=headers,
                timeout=REQUEST_TIMEOUT,
                verify=False,
            )
        except requests.RequestException as exc:  # pragma: no cover - network dependent
            last_error = exc
            sleep_for = min(2 ** attempt, 30)
            logger.warning("POST %s failed (%s); sleeping %.1fs", endpoint, exc, sleep_for)
            time.sleep(sleep_for)
            continue

        if response.status_code in TRANSIENT_STATUS:
            last_error = RuntimeError(f"HTTP {response.status_code}")
            sleep_for = min(2 ** attempt, 30)
            logger.warning(
                "POST %s returned %s; sleeping %.1fs",
                endpoint,
                response.status_code,
                sleep_for,
            )
            time.sleep(sleep_for)
            continue

        try:
            data = response.json()
        except ValueError as exc:
            last_error = exc
            logger.warning("Invalid JSON from %s: %s", endpoint, exc)
            time.sleep(1.5)
            continue

        errors = data.get("errors") or []
        if errors:
            logger.warning(
                "GraphQL response from %s contained %s errors; proceeding with available data",
                endpoint,
                len(errors),
            )
        return data

    raise GraphQLRequestError(f"GraphQL request failed after retries: {last_error}")


def select_operational_endpoint(session: requests.Session, logger: logging.Logger) -> str | None:
    if not STOREFRONT_COLLECTION_HANDLES:
        logger.warning("No collection handles configured for Storefront extraction")
        return None

    test_handle = STOREFRONT_COLLECTION_HANDLES[0]
    payload = {
        "query": STOREFRONT_QUERY,
        "variables": {"handle": test_handle, "cursor": None, "pageSize": 1},
    }
    headers = {"Content-Type": "application/json"}

    for endpoint in STOREFRONT_ENDPOINTS:
        logger.info("Testing Storefront endpoint %s", endpoint)
        try:
            response = session.post(
                endpoint,
                json=payload,
                headers=headers,
                timeout=STOREFRONT_TIMEOUT,
                verify=False,
            )
        except requests.RequestException as exc:  # pragma: no cover - network dependent
            logger.warning("Probe request to %s failed: %s", endpoint, exc)
            continue

        if not response.ok:
            logger.info(
                "Endpoint %s returned status %s during probe", endpoint, response.status_code
            )
            continue

        try:
            data = response.json()
        except ValueError:
            logger.info("Endpoint %s returned non-JSON payload", endpoint)
            continue

        collection = (data.get("data", {}) or {}).get("collection")
        products = ((collection or {}).get("products") or {}).get("edges") or []
        if products:
            logger.info("Endpoint %s contains collection data; selecting for extraction", endpoint)
            return endpoint

    logger.warning("No operational Storefront endpoint produced collection data")
    return None


def normalise_field_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, Path):
        return str(value)
    return json.dumps(value, ensure_ascii=False, sort_keys=True)


def flatten_storefront_rows(session: requests.Session, logger: logging.Logger) -> List[Dict[str, Any]]:
    if not STOREFRONT_COLLECTION_HANDLES:
        logger.info("No Storefront collection handles configured; skipping extraction")
        return []

    preferred_endpoint = select_operational_endpoint(session, logger)
    endpoints = iter_graphql_endpoints(preferred_endpoint)
    if not endpoints:
        logger.warning("No Storefront endpoints configured; skipping extraction")
        return []

    rows: List[Dict[str, Any]] = []

    for handle in STOREFRONT_COLLECTION_HANDLES:
        logger.info("Fetching Storefront data for collection '%s'", handle)
        cursor: str | None = None

        while True:
            payload = {
                "query": STOREFRONT_QUERY,
                "variables": {"handle": handle, "cursor": cursor, "pageSize": GRAPHQL_PAGE_SIZE},
            }
            try:
                data = execute_graphql(session, payload, endpoints, logger)
            except GraphQLRequestError as exc:
                logger.error("Storefront request failed for handle '%s': %s", handle, exc)
                break

            collection = (data.get("data", {}) or {}).get("collection")
            if not collection:
                logger.warning("No collection returned for handle '%s'", handle)
                break

            collection_info = {
                "collection_id": normalise_field_value(collection.get("id")),
                "collection_handle": normalise_field_value(collection.get("handle")),
                "collection_title": normalise_field_value(collection.get("title")),
            }

            products_connection = collection.get("products") or {}
            product_edges: Iterable[Dict[str, Any]] = products_connection.get("edges", []) or []

            for edge in product_edges:
                product = edge.get("node") or {}
                base: Dict[str, Any] = dict(collection_info)
                base["products_edge_cursor"] = edge.get("cursor", "")

                for key, value in product.items():
                    if key == "variants":
                        continue
                    column_name = f"product_{key}"
                    base[column_name] = normalise_field_value(value)

                variants_connection = product.get("variants") or {}
                variant_edges: Iterable[Dict[str, Any]] = variants_connection.get("edges", []) or []

                if not variant_edges:
                    rows.append(dict(base))
                    continue

                for variant_edge in variant_edges:
                    variant = variant_edge.get("node") or {}
                    row = dict(base)
                    row["variant_edge_cursor"] = variant_edge.get("cursor", "")
                    for key, value in variant.items():
                        column_name = f"variant_{key}"
                        row[column_name] = normalise_field_value(value)
                    rows.append(row)

            page_info = products_connection.get("pageInfo") or {}
            if page_info.get("hasNextPage"):
                cursor = page_info.get("endCursor")
                logger.info(
                    "Collection '%s' has more pages; continuing with cursor %s",
                    handle,
                    cursor,
                )
                if not cursor:
                    logger.warning(
                        "Missing endCursor for collection '%s'; aborting pagination",
                        handle,
                    )
                    break
            else:
                break

    logger.info("Flattened %s rows from the Storefront API", len(rows))
    return rows


def probe_storefront_endpoints(
    session: requests.Session, logger: logging.Logger
) -> List[Dict[str, Any]]:
    results: List[Dict[str, Any]] = []
    payload = {"query": "query { shop { name } }"}
    headers = {"Content-Type": "application/json"}
    for endpoint in STOREFRONT_ENDPOINTS:
        logger.info("Probing Storefront endpoint %s", endpoint)
        entry: Dict[str, Any] = {"endpoint": endpoint}
        try:
            response = session.post(
                endpoint,
                json=payload,
                headers=headers,
                timeout=STOREFRONT_TIMEOUT,
                verify=False,
            )
        except requests.RequestException as exc:  # pragma: no cover - network dependent
            logger.warning("Storefront request to %s failed: %s", endpoint, exc)
            entry["status_code"] = None
            entry["ok"] = False
            entry["error"] = str(exc)
            results.append(entry)
            continue

        entry["status_code"] = response.status_code
        body_text = response.text.strip()

        try:
            data = response.json()
        except ValueError:
            logger.info(
                "Endpoint %s returned non-JSON payload (%s bytes)",
                endpoint,
                len(response.content),
            )
            entry["ok"] = response.ok
            entry["response_excerpt"] = body_text[:500]
        else:
            if response.ok and data.get("data") and not data.get("errors"):
                entry["ok"] = True
                entry["response_excerpt"] = json.dumps(
                    data.get("data"), ensure_ascii=False
                )
            else:
                entry["ok"] = False
                entry["response_excerpt"] = json.dumps(
                    data.get("errors") or data, ensure_ascii=False
                )[:500]

        results.append(entry)
    return results


def write_sheet(sheet, rows: List[Dict[str, Any]]):
    if not rows:
        sheet.append(["No data"])
        return
    columns = sorted({key for row in rows for key in row.keys()})
    sheet.append(columns)
    for row in rows:
        sheet.append([row.get(column) for column in columns])


def export_to_workbook(
    globo_rows: List[Dict[str, Any]],
    preorder_settings: Dict[str, Any],
    preorder_profiles: Dict[str, Any],
    gravity_rows: List[Dict[str, Any]],
    storefront_rows: List[Dict[str, Any]],
    storefront_probe_rows: List[Dict[str, Any]],
) -> Path:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "GloboVariants"
    write_sheet(worksheet, globo_rows)

    settings_sheet = workbook.create_sheet("PreorderSettings")
    settings_rows: List[Dict[str, Any]] = []
    for product_id, variant_map in preorder_settings.get("products", {}).items():
        for variant_id, profile_id in variant_map.items():
            settings_rows.append(
                {
                    "product_id": product_id,
                    "variant_id": variant_id,
                    "profile_id": profile_id,
                }
            )
    write_sheet(settings_sheet, settings_rows)

    profiles_sheet = workbook.create_sheet("Profiles")
    profile_rows = []
    for profile_id, payload in preorder_profiles.items():
        row = {"profile_id": profile_id}
        row.update({f"profile_{key}": normalize_value(value) for key, value in payload.items()})
        profile_rows.append(row)
    write_sheet(profiles_sheet, profile_rows)

    gravity_sheet = workbook.create_sheet("GravityVariants")
    write_sheet(gravity_sheet, gravity_rows)

    storefront_sheet = workbook.create_sheet("Storefront")
    write_sheet(storefront_sheet, storefront_rows)

    storefront_probe_sheet = workbook.create_sheet("StorefrontProbes")
    write_sheet(storefront_probe_sheet, storefront_probe_rows)

    timestamp = datetime.now(timezone.utc).astimezone().strftime("%Y-%m-%d_%H-%M-%S")
    output_path = OUTPUT_DIR / f"SELFCONTRAST_Globo_Gravity_{timestamp}.xlsx"
    workbook.save(output_path)
    return output_path


def main() -> None:
    logger = configure_logging()
    session = build_session()
    html = fetch_collection_html(session, logger)
    globo_script = extract_globo_script(html, logger)
    globo_products = parse_globo_product_arrays(globo_script, logger)
    preorder_settings = parse_preorder_settings(globo_script, logger)
    preorder_profiles = parse_preorder_profiles(globo_script, logger)
    globo_rows = build_globo_variant_rows(globo_products, preorder_settings, preorder_profiles)
    gravity_rows = parse_gravity_variants(html, logger)
    storefront_rows = flatten_storefront_rows(session, logger)
    storefront_probe_rows = probe_storefront_endpoints(session, logger)
    output_path = export_to_workbook(
        globo_rows,
        preorder_settings,
        preorder_profiles,
        gravity_rows,
        storefront_rows,
        storefront_probe_rows,
    )
    logger.info("Workbook written to %s", output_path.as_posix())


if __name__ == "__main__":
    main()
