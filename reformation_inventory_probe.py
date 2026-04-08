#!/usr/bin/env python3
"""Reformation (SFCC) inventory probe with master/sub PID handling.

Source order:
1) Product-ShowQuickAdd?pid=<MASTER7>
2) Product-ShowQuickAdd?pid=<MASTER7+COLOR3>
3) Product-ShowQuickView?pid=<MASTER7>
4) Search-ShowAjax refinements for facet labels
"""

from __future__ import annotations

import json
import logging
import re
import time
from collections import deque
from datetime import datetime
from fractions import Fraction
from pathlib import Path
from typing import Dict, Iterable, List, Set
from urllib.parse import parse_qs, unquote, urljoin, urlparse

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

BRAND = "Reformation"
BASE_URL = "https://www.thereformation.com"
LOCALE_PATH = "/on/demandware.store/Sites-reformation-us-Site/en_US"
SEARCH_AJAX_URL = f"{BASE_URL}{LOCALE_PATH}/Search-ShowAjax"
QUICK_ADD_URL = f"{BASE_URL}{LOCALE_PATH}/Product-ShowQuickAdd"
QUICK_VIEW_URL = f"{BASE_URL}{LOCALE_PATH}/Product-ShowQuickView"
MAX_MASTERS: int | None = None

BASE_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = BASE_DIR / "Output"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
)
TIMEOUT = 30
MAX_RETRIES = 5

HEADERS = [
    "Style Id", "Handle", "Published At", "Created At", "Product", "Style Name", "Product Type", "Tags", "Vendor",
    "Description", "Variant Title", "Color", "Color Code", "Style Id + Color", "Size", "Rise", "Inseam", "Leg Opening",
    "Price", "Compare at Price", "Profit", "Promo", "Currency", "Min Price", "Min Compare at Price", "Max Price",
    "Max Compare at Price", "Available for Sale", "Stock Status", "Waitlist Active", "In Stock Date", "Ship Date",
    "Quantity Available", "Quantity of Style", "Inventory Rank", "Add to Cart Rate", "Inventory Sort", "SKU - Shopify",
    "SKU - Brand", "Barcode", "Image URL", "SKU URL", "Jean Style", "Jean Style Description", "Product Line",
    "Inseam Label", "Inseam Style", "Rise Label", "Color - Simplified", "Country of Origin", "Stretch", "Class Group",
    "GenX Targeting", "Millennial Targeting", "GenZ Targeting", "Rating",
]


class ProbeError(RuntimeError):
    pass


def setup_logger() -> logging.Logger:
    logger = logging.getLogger("reformation_inventory_probe")
    logger.setLevel(logging.INFO)
    logger.handlers.clear()

    stamp = datetime.now().strftime("%Y-%m-%d")
    pref = OUTPUT_DIR / f"{BRAND.lower()}_inventory_probe_{stamp}.log"
    fallback = OUTPUT_DIR / f"{BRAND.lower()}_run.log"
    fmt = logging.Formatter("%(asctime)s | %(levelname)s | %(message)s")

    sh = logging.StreamHandler()
    sh.setFormatter(fmt)
    logger.addHandler(sh)

    try:
        fh = logging.FileHandler(pref, encoding="utf-8")
        fh.setFormatter(fmt)
        logger.addHandler(fh)
    except Exception:
        fh = logging.FileHandler(fallback, encoding="utf-8")
        fh.setFormatter(fmt)
        logger.addHandler(fh)

    return logger


def build_session() -> requests.Session:
    s = requests.Session()
    s.headers.update({
        "User-Agent": USER_AGENT,
        "Accept": "application/json,text/html,*/*",
        "X-Requested-With": "XMLHttpRequest",
    })
    return s


def get_with_retries(session: requests.Session, url: str, logger: logging.Logger, params: dict | None = None) -> requests.Response:
    delay = 1.0
    last_error: Exception | None = None
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            r = session.get(url, params=params, timeout=TIMEOUT)
            if r.status_code in {429, 500, 502, 503, 504}:
                logger.warning("Transient HTTP %s for %s (attempt %s/%s)", r.status_code, r.url, attempt, MAX_RETRIES)
                time.sleep(delay)
                delay *= 2
                continue
            r.raise_for_status()
            return r
        except requests.RequestException as exc:
            last_error = exc
            logger.warning("Request failed for %s (attempt %s/%s): %s", url, attempt, MAX_RETRIES, exc)
            time.sleep(delay)
            delay *= 2
    raise ProbeError(f"Failed request after {MAX_RETRIES}: {url} | {last_error}")


def to_text(v: object) -> str:
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    if isinstance(v, (int, float, bool)):
        return str(v)
    if isinstance(v, list):
        return " | ".join([to_text(x) for x in v if to_text(x)])
    if isinstance(v, dict):
        return json.dumps(v, ensure_ascii=False)
    return str(v)


def dedupe_join(parts: Iterable[str], sep: str = " | ") -> str:
    seen: Set[str] = set()
    out: List[str] = []
    for p in parts:
        t = re.sub(r"\s+", " ", to_text(p)).strip()
        if t and t.lower() not in seen:
            out.append(t)
            seen.add(t.lower())
    return sep.join(out)


def parse_date(value: object) -> str:
    t = to_text(value)
    if not t:
        return ""
    for fmt in ("%Y-%m-%d", "%Y-%m-%dT%H:%M:%S.%fZ", "%Y-%m-%dT%H:%M:%SZ", "%a %b %d %Y", "%b %d, %Y", "%b %d, %Y"):
        try:
            return datetime.strptime(t.strip(), fmt).strftime("%m/%d/%Y")
        except ValueError:
            continue
    t2 = t.strip().replace(",", "")
    for fmt in ("%b %d %Y", "%a %b %d %Y"):
        try:
            return datetime.strptime(t2, fmt).strftime("%m/%d/%Y")
        except ValueError:
            continue
    return ""


def parse_number_from_fit(text: str, label: str) -> str:
    if not text:
        return ""
    m = re.search(rf"{re.escape(label)}\s*:\s*([^,\.]+)", text, flags=re.I)
    if not m:
        return ""
    token = m.group(1).strip()
    n = re.search(r"(\d+\s+\d+/\d+|\d+/\d+|\d+(?:\.\d+)?)", token)
    if not n:
        return ""
    token = n.group(1)
    try:
        if " " in token and "/" in token:
            whole, frac = token.split()
            value = float(whole) + float(Fraction(frac))
        elif "/" in token:
            value = float(Fraction(token))
        else:
            value = float(token)
        return ("%.2f" % value).rstrip("0").rstrip(".")
    except Exception:
        return token


def get_attr(product: dict, attr_id: str) -> dict | None:
    for a in product.get("variationAttributes") or []:
        if (a.get("attributeId") or a.get("id")) == attr_id:
            return a
    return None


def extract_master_from_pid(pid: str) -> str:
    m = re.match(r"(\d{7})", pid or "")
    return m.group(1) if m else ""


def extract_color_from_pid(pid: str) -> str:
    m = re.match(r"\d{7}([A-Z]{3})", pid or "")
    return m.group(1) if m else ""


def extract_color_codes(product: dict) -> List[str]:
    codes: Set[str] = set()

    color_attr = get_attr(product, "color")
    if color_attr:
        for v in color_attr.get("values") or []:
            unavailable_colors = v.get("unavailableColors")
            # Keep only active colors from the dedicated color attribute.
            # Reformation marks active entries with unavailableColors: [].
            if unavailable_colors != []:
                continue
            c = to_text(v.get("id") or v.get("value"))
            if re.fullmatch(r"[A-Z]{3}", c):
                codes.add(c)

    # Fallback to sizeByColor traversal when color attribute data is missing
    # or does not expose active colors.
    sbc = get_attr(product, "sizeByColor")
    if sbc and not codes:
        for cp in sbc.get("values") or []:
            color = cp.get("color") or {}
            c = to_text(color.get("id") or color.get("value"))
            if re.fullmatch(r"[A-Z]{3}", c):
                codes.add(c)

    return sorted(codes)


def canonical_handle(raw_handle: str, style_id_color: str) -> str:
    handle = to_text(raw_handle)
    if not handle:
        return ""
    path = urlparse(handle).path or handle
    return re.sub(r"(\d{7}[A-Z]{3})\.html", f"{style_id_color}.html", path)


def parse_float(value: object) -> float | None:
    t = to_text(value)
    if not t:
        return None
    t = t.replace(",", "")
    m = re.search(r"-?\d+(?:\.\d+)?", t)
    if not m:
        return None
    try:
        return float(m.group(0))
    except ValueError:
        return None


def parse_currency(value: object) -> float | None:
    return parse_float(to_text(value).replace("$", ""))


def parse_percent(value: object) -> float | None:
    t = to_text(value)
    if not t:
        return None
    if "%" in t:
        n = parse_float(t)
        return (n / 100.0) if n is not None else None
    n = parse_float(t)
    if n is None:
        return None
    return n / 100.0 if n > 1 else n


def extract_pids_from_search_html(html: str) -> List[str]:
    soup = BeautifulSoup(html, "lxml")
    pids: Set[str] = set()
    for t in soup.select('[data-product-container][data-pid]'):
        pid = to_text(t.get("data-pid"))
        if pid:
            pids.add(pid)
    return sorted(pids)


def collect_initial_masters(session: requests.Session, logger: logging.Logger) -> List[str]:
    found: Set[str] = set()
    start, sz = 0, 16
    while True:
        resp = get_with_retries(session, SEARCH_AJAX_URL, logger, params={"cgid": "jeans", "pmpt": "qualifying", "start": start, "sz": sz})
        pids = extract_pids_from_search_html(resp.text)
        logger.info("Search-ShowAjax start=%s -> %s pids", start, len(pids))
        if not pids:
            break
        for pid in pids:
            m = extract_master_from_pid(pid)
            if m:
                found.add(m)
        start += sz
    masters = sorted(found)
    if MAX_MASTERS is not None:
        masters = masters[:MAX_MASTERS]
    logger.info("Initial master pids: %s", len(masters))
    return masters


def fetch_quick_add(session: requests.Session, logger: logging.Logger, pid: str) -> dict:
    return get_with_retries(session, QUICK_ADD_URL, logger, params={"pid": pid}).json()


def fetch_quick_view(session: requests.Session, logger: logging.Logger, pid: str) -> dict:
    return get_with_retries(session, QUICK_VIEW_URL, logger, params={"pid": pid}).json()


def collect_refinement_values(session: requests.Session, logger: logging.Logger) -> Dict[str, Set[str]]:
    resp = get_with_retries(session, SEARCH_AJAX_URL, logger, params={"cgid": "jeans", "pmpt": "qualifying"})
    soup = BeautifulSoup(resp.text, "lxml")
    values = {"rise": set(), "denim_wash": set(), "leg": set(), "fabric_type": set()}
    for a in soup.select('a[href*="prefn1="]'):
        href = to_text(a.get("href")).replace("&amp;", "&")
        qs = parse_qs(urlparse(href).query)
        f = to_text(qs.get("prefn1", [""])[0])
        v = to_text(qs.get("prefv1", [""])[0])
        if f in values and v:
            values[f].add(unquote(v))
    logger.info("Refinements: %s", {k: len(v) for k, v in values.items()})
    return values


def build_facet_reverse(session: requests.Session, logger: logging.Logger) -> Dict[str, Dict[str, List[str]]]:
    refinements = collect_refinement_values(session, logger)
    reverse: Dict[str, Dict[str, List[str]]] = {}
    for facet, vals in refinements.items():
        for val in sorted(vals):
            resp = get_with_retries(session, SEARCH_AJAX_URL, logger, params={"cgid": "jeans", "pmpt": "qualifying", "prefn1": facet, "prefv1": val})
            pids = extract_pids_from_search_html(resp.text)
            logger.info("Facet %s=%s -> %s pids", facet, val, len(pids))
            for pid in pids:
                m = extract_master_from_pid(pid)
                if not m:
                    continue
                reverse.setdefault(m, {}).setdefault(facet, []).append(val)
    for m in reverse:
        for f in reverse[m]:
            reverse[m][f] = sorted(set(reverse[m][f]))
    return reverse


def discover_master_graph(session: requests.Session, logger: logging.Logger, initial_masters: List[str]):
    master_quickadd: Dict[str, dict] = {}
    master_quickview: Dict[str, dict] = {}
    master_subpids: Dict[str, Set[str]] = {}

    q = deque(initial_masters)
    seen: Set[str] = set()

    while q:
        master = q.popleft()
        if master in seen:
            continue
        seen.add(master)

        try:
            qa = fetch_quick_add(session, logger, master)
            qv = fetch_quick_view(session, logger, master)
        except Exception as exc:
            logger.warning("Skipping master %s (%s)", master, exc)
            continue

        master_quickadd[master] = qa
        master_quickview[master] = qv
        master_subpids.setdefault(master, set())

        qa_product = qa.get("product") or {}
        qv_product = qv.get("product") or {}

        # Required: derive all color sub-pids from master product color variation values
        for color_code in sorted(set(extract_color_codes(qa_product) + extract_color_codes(qv_product))):
            master_subpids[master].add(master + color_code)

        sbc = get_attr(qa_product, "sizeByColor")
        if not sbc:
            continue

        for cp in sbc.get("values") or []:
            for sz in cp.get("sizes") or []:
                sp = sz.get("product") or {}
                lvl = sp.get("length_variation_links") or {}
                links = lvl.get("links") if isinstance(lvl, dict) else []
                for link in links or []:
                    if link.get("currentProduct") is False:
                        pid2 = to_text(link.get("productID"))
                        m2 = extract_master_from_pid(pid2)
                        if m2 and m2 not in seen:
                            q.append(m2)
                        msub = re.match(r"(\d{7}[A-Z]{3})", pid2)
                        if msub:
                            m_for_sub = extract_master_from_pid(msub.group(1))
                            master_subpids.setdefault(m_for_sub, set()).add(msub.group(1))

        logger.info("master=%s discovered colors=%s", master, len(master_subpids.get(master, set())))

    logger.info("Final master count after length-variation expansion: %s", len(master_quickadd))
    return master_quickadd, master_quickview, master_subpids


def quickview_color_maps(product: dict):
    metric_by_color_name: Dict[str, dict] = {}
    product_line_by_color_name: Dict[str, str] = {}
    color_attr = get_attr(product, "color")
    if not color_attr:
        return metric_by_color_name, product_line_by_color_name
    for v in color_attr.get("values") or []:
        name = to_text(v.get("displayValue"))
        if not name:
            continue
        key = name.lower()
        metric_by_color_name[key] = {
            "Inventory Rank": to_text(v.get("inventory_rank")),
            "Add to Cart Rate": to_text(v.get("atc_rate")),
            "Inventory Sort": to_text(v.get("inventory_sort")),
        }
        product_line_by_color_name[key] = to_text(v.get("fashion_core"))
    return metric_by_color_name, product_line_by_color_name


def find_sub_variant_product(sub_qa: dict, color_code: str, size_value: str) -> dict:
    p = sub_qa.get("product") or {}
    sbc = get_attr(p, "sizeByColor")
    if not sbc:
        return {}
    for cp in sbc.get("values") or []:
        c = cp.get("color") or {}
        if to_text(c.get("value")) != color_code:
            continue
        for sz in cp.get("sizes") or []:
            sv = to_text(sz.get("value") or sz.get("id") or sz.get("displayValue"))
            if sv == size_value:
                return sz.get("product") or {}
    return {}


def build_description(prod: dict) -> str:
    return dedupe_join([
        to_text(prod.get("features")),
        to_text(prod.get("material_description")),
        to_text(prod.get("marketing_description")),
        to_text(prod.get("fit_intent")),
        to_text(prod.get("fit_guidance")),
        to_text(prod.get("fit_tip")),
        to_text(prod.get("custom")),
        to_text(prod.get("fabric_info")),
        to_text(prod.get("fit_flexibility")),
    ])


def choose_image(color_obj: dict) -> str:
    images = color_obj.get("images") or {}
    for key in ("medium", "large"):
        arr = images.get(key)
        if isinstance(arr, list) and arr:
            return to_text(arr[0].get("absURL"))
    return ""


def promotion_text(promotions: object) -> str:
    if promotions is None:
        return ""
    if isinstance(promotions, str):
        return promotions.strip()
    if isinstance(promotions, list):
        parts = []
        for p in promotions:
            if isinstance(p, dict):
                parts.append(to_text(p.get("calloutMsg") or p.get("name") or p.get("id")))
            else:
                parts.append(to_text(p))
        return dedupe_join(parts)
    if isinstance(promotions, dict):
        return dedupe_join([to_text(promotions.get("calloutMsg")), to_text(promotions.get("name")), to_text(promotions.get("id"))])
    return to_text(promotions)


def extract_inseam_label(sub_variant_prod: dict) -> str:
    lvl = sub_variant_prod.get("length_variation_links") or {}
    if isinstance(lvl, dict):
        return to_text(lvl.get("length"))
    return ""


def extract_stock_status(availability: dict) -> str:
    msgs = availability.get("messages") or []
    return dedupe_join([to_text(m) for m in msgs])


def build_rows(master: str, master_qa: dict, master_qv: dict, sub_qa_map: Dict[str, dict], facet_reverse: Dict[str, Dict[str, List[str]]]) -> List[dict]:
    rows: List[dict] = []
    p_master = master_qa.get("product") or {}
    p_view = master_qv.get("product") or {}

    style_name = to_text(p_master.get("productName"))
    vendor = to_text(p_master.get("brand"))
    color_metrics, product_line_map = quickview_color_maps(p_view)
    qv_price = p_view.get("price") or {}

    min_price = to_text((((qv_price.get("min") or {}).get("sales") or {}).get("formatted")))
    min_compare = to_text((((qv_price.get("min") or {}).get("list") or {}).get("formatted")))
    max_price = to_text((((qv_price.get("max") or {}).get("sales") or {}).get("formatted")))
    max_compare = to_text((((qv_price.get("max") or {}).get("list") or {}).get("formatted")))

    facet = facet_reverse.get(master, {})
    jean_style = dedupe_join(facet.get("leg", []), sep=", ")
    rise_label = dedupe_join(facet.get("rise", []), sep=", ")
    color_simple = dedupe_join(facet.get("denim_wash", []), sep=", ")
    stretch = dedupe_join(facet.get("fabric_type", []), sep=", ")

    sbc = get_attr(p_master, "sizeByColor")
    if not sbc:
        return rows

    for cp in sbc.get("values") or []:
        color = cp.get("color") or {}
        color_name = to_text(color.get("displayValue"))
        color_code = to_text(color.get("value"))
        style_color = master + color_code if re.fullmatch(r"[A-Z]{3}", color_code) else to_text(color.get("productId"))
        raw_handle = to_text(color.get("url"))
        handle = canonical_handle(raw_handle, style_color)
        sku_url = urljoin(BASE_URL, handle)
        image_url = choose_image(color)

        # requirement: drop rows where handle color token doesn't match Color Code
        mhandle = re.search(r'_color=([A-Z]{3})', raw_handle)
        if mhandle and mhandle.group(1) != color_code:
            continue

        subpid = master + color_code if re.fullmatch(r"[A-Z]{3}", color_code) else ""
        sub_qa = sub_qa_map.get(subpid, {})
        sub_root_prod = (sub_qa.get("product") or {})
        rating = to_text(sub_root_prod.get("rating"))
        profit = to_text(sub_root_prod.get("revenue"))
        promo = promotion_text(sub_root_prod.get("promotions"))

        cmetrics = color_metrics.get(color_name.lower(), {})
        product_line = to_text(product_line_map.get(color_name.lower(), ""))

        for sz in cp.get("sizes") or []:
            size_value = to_text(sz.get("value") or sz.get("id") or "")
            size_display = to_text(sz.get("displayValue"))
            base_variant = sz.get("product") or {}
            sub_variant = find_sub_variant_product(sub_qa, color_code, size_value) if sub_qa else {}
            variant = sub_variant or base_variant

            fit = to_text(variant.get("fit_features"))
            rise = parse_number_from_fit(fit, "rise")
            inseam = parse_number_from_fit(fit, "inseam")
            leg_opening = parse_number_from_fit(fit, "leg opening")

            price_sales = (base_variant.get("price") or {}).get("sales") or {}
            compare_list = (base_variant.get("price") or {}).get("list") or {}
            availability = base_variant.get("availability") or {}

            row = {
                "Style Id": parse_float(master) if parse_float(master) is not None else master,
                "Handle": handle,
                "Published At": "",
                "Created At": "",
                "Product": dedupe_join([style_name, color_name], sep=" "),
                "Style Name": style_name,
                "Product Type": to_text(base_variant.get("fit_guide_id")),
                "Tags": to_text(base_variant.get("product_tag")),
                "Vendor": vendor,
                "Description": build_description(base_variant),
                "Variant Title": dedupe_join([style_name, color_name, size_display], sep=" "),
                "Color": color_name,
                "Color Code": color_code,
                "Style Id + Color": style_color,
                "Size": size_display,
                "Rise": rise,
                "Inseam": inseam,
                "Leg Opening": leg_opening,
                "Price": to_text(price_sales.get("formatted")),
                "Compare at Price": to_text(compare_list.get("formatted")),
                "Profit": profit,
                "Promo": promo,
                "Currency": to_text(price_sales.get("currency")),
                "Min Price": min_price,
                "Min Compare at Price": min_compare,
                "Max Price": max_price,
                "Max Compare at Price": max_compare,
                "Available for Sale": to_text(base_variant.get("available")),
                "Stock Status": extract_stock_status(availability),
                "Waitlist Active": to_text(base_variant.get("wait_list")),
                "In Stock Date": parse_date(availability.get("inStockDate")),
                "Ship Date": parse_date(base_variant.get("estimated_ship_date")),
                "Quantity Available": to_text(availability.get("ats")),
                "Quantity of Style": "",
                "Inventory Rank": to_text(cmetrics.get("Inventory Rank", "")),
                "Add to Cart Rate": to_text(cmetrics.get("Add to Cart Rate", "")),
                "Inventory Sort": to_text(cmetrics.get("Inventory Sort", "")),
                "SKU - Shopify": "",
                "SKU - Brand": to_text(base_variant.get("id")),
                "Barcode": "",
                "Image URL": image_url,
                "SKU URL": sku_url,
                "Jean Style": jean_style,
                "Jean Style Description": to_text(base_variant.get("fit_intent")),
                "Product Line": product_line,
                "Inseam Label": extract_inseam_label(variant),
                "Inseam Style": "",
                "Rise Label": rise_label,
                "Color - Simplified": color_simple,
                "Country of Origin": to_text(base_variant.get("country_of_origin")),
                "Stretch": stretch,
                "Class Group": to_text(base_variant.get("class_group")),
                "GenX Targeting": to_text(base_variant.get("generation_GenX")),
                "Millennial Targeting": to_text(base_variant.get("generation_Millennials")),
                "GenZ Targeting": to_text(base_variant.get("generation_GenZ")),
                "Rating": rating,
            }
            rows.append(row)

    return rows


def fill_quantity_of_style(rows: List[dict]) -> None:
    totals: Dict[str, float] = {}
    for r in rows:
        key = r.get("Product", "")
        try:
            qty = float(to_text(r.get("Quantity Available")) or 0)
        except Exception:
            qty = 0.0
        totals[key] = totals.get(key, 0.0) + qty
    for r in rows:
        t = totals.get(r.get("Product", ""), 0.0)
        r["Quantity of Style"] = str(int(t)) if float(t).is_integer() else str(round(t, 2))


def write_excel(rows: List[dict], logger: logging.Logger) -> Path:
    stamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    out = OUTPUT_DIR / f"{BRAND}_Inventory_Probe_{stamp}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"
    ws.append(HEADERS)

    number_cols = {"Style Id", "Rise", "Inseam", "Leg Opening", "Quantity Available", "Quantity of Style", "Inventory Rank", "Inventory Sort", "Rating"}
    dollar_cols = {"Price", "Compare at Price", "Profit", "Min Price", "Min Compare at Price", "Max Price", "Max Compare at Price"}
    percent_cols = {"GenX Targeting", "Millennial Targeting", "GenZ Targeting"}

    for row in rows:
        out_row = []
        for h in HEADERS:
            v = row.get(h, "")
            if h in number_cols:
                n = parse_float(v)
                out_row.append(n if n is not None else v)
            elif h in dollar_cols:
                n = parse_currency(v)
                out_row.append(n if n is not None else v)
            elif h in percent_cols:
                n = parse_percent(v)
                out_row.append(n if n is not None else v)
            else:
                out_row.append(v)
        ws.append(out_row)

    header_idx = {h: i + 1 for i, h in enumerate(HEADERS)}
    for h in dollar_cols:
        col = header_idx[h]
        for r in range(2, ws.max_row + 1):
            ws.cell(r, col).number_format = '$#,##0.00'
    for h in percent_cols:
        col = header_idx[h]
        for r in range(2, ws.max_row + 1):
            ws.cell(r, col).number_format = '0.00%'

    wb.save(out)
    logger.info("Excel written: %s", out.resolve())
    return out


def main() -> None:
    logger = setup_logger()
    session = build_session()

    initial_masters = collect_initial_masters(session, logger)
    facet_reverse = build_facet_reverse(session, logger)

    master_quickadd, master_quickview, master_subpids = discover_master_graph(session, logger, initial_masters)

    # fetch sub pid quickadd payloads once
    sub_qa_cache: Dict[str, dict] = {}
    for master, subpids in master_subpids.items():
        for subpid in sorted(subpids):
            if subpid in sub_qa_cache:
                continue
            try:
                sub_qa_cache[subpid] = fetch_quick_add(session, logger, subpid)
            except Exception as exc:
                logger.warning("subpid fetch failed %s: %s", subpid, exc)

    all_rows: List[dict] = []
    for idx, master in enumerate(sorted(master_quickadd.keys()), start=1):
        rows = build_rows(master, master_quickadd[master], master_quickview.get(master, {}), sub_qa_cache, facet_reverse)
        all_rows.extend(rows)
        logger.info("%s/%s master=%s rows=%s", idx, len(master_quickadd), master, len(rows))

    # de-dupe exact SKU-brand rows from accidental overlaps
    dedup: Dict[str, dict] = {}
    for row in all_rows:
        key = f"{row.get('SKU - Brand','')}|{row.get('Handle','')}"
        dedup[key] = row
    final_rows = list(dedup.values())

    fill_quantity_of_style(final_rows)
    write_excel(final_rows, logger)


if __name__ == "__main__":
    main()
