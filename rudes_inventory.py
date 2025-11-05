# rudes_inventory.py
# -- Rudes Denim full-catalog scraper (CSV + daily Excel) --
# Quantities come from PDP <script> ReStock-config (JavaScript, not JSON).
# Author: you & ChatGPT

from __future__ import annotations

import csv
import html
import json
import logging
import re
import time
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple
from urllib.parse import urljoin, urlparse, urlunparse

import requests

BASE_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = BASE_DIR / "Output"
LOG_PATH = BASE_DIR / "rudes_run.log"
FALLBACK_LOG_PATH = OUTPUT_DIR / "rudes_run.log"

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def configure_logging() -> logging.Logger:
    logger = logging.getLogger("rudes")
    if logger.handlers:
        return logger

    logger.setLevel(logging.INFO)
    formatter = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
    selected_path: Path | None = None
    fallback_used = False

    for path in (LOG_PATH, FALLBACK_LOG_PATH):
        try:
            path.parent.mkdir(parents=True, exist_ok=True)
            handler = logging.FileHandler(path, encoding="utf-8")
            handler.setFormatter(formatter)
            logger.addHandler(handler)
            selected_path = path
            if path != LOG_PATH:
                fallback_used = True
            break
        except (OSError, PermissionError) as exc:
            print(
                f"WARNING: Unable to open log file {path}: {exc}. Continuing without this destination.",
                flush=True,
            )

    stream_handler = logging.StreamHandler()
    stream_handler.setFormatter(formatter)
    logger.addHandler(stream_handler)

    if selected_path is None:
        logger.warning("File logging disabled; continuing with console logging only.")
    elif fallback_used:
        logger.warning("Primary log path %s unavailable. Using fallback log at %s.", LOG_PATH, selected_path)

    return logger


LOGGER = configure_logging()

# --------- CONFIG ----------
PRIMARY_HOSTS = [
    "https://rudesdenim.com",
    "https://rudes-jeans.myshopify.com",
]
BASE = PRIMARY_HOSTS[0]
PRODUCTS_PATH = "/products.json"
PDP_PATH_TEMPLATE = "/products/{handle}"
EXCEL_PATH = OUTPUT_DIR / "RUDES_daily.xlsx"

# Turn OCR on only when you’re ready. It needs external Tesseract installed & PATH set.
OCR_MEASUREMENTS = False            # <- keep False for now (body_html rules first)
TESSERACT_PATH   = r"C:\Program Files\Tesseract-OCR\tesseract.exe"  # if you later turn OCR on

RETRIES = 2

# Allowed sizes to disambiguate size vs color
ALLOWED_SIZES = {
    "23","24","25","26","27","28","29","30","31","32","33","34",
    "x-small","small","medium","large","x-large","xx-large","xs","s","m","l","xl","2xl",
    "X-Small","Small","Medium","Large","X-Large","XX-Large","XS","S","M","L","XL","2XL"
}

# --------- UTILS ----------

def log(message: str, level: int = logging.INFO) -> None:
    LOGGER.log(level, message)


def _candidate_urls(url: str) -> Iterable[str]:
    parsed = urlparse(url)
    if parsed.scheme and parsed.netloc:
        for host in PRIMARY_HOSTS:
            host_parsed = urlparse(host)
            yield urlunparse(
                (
                    host_parsed.scheme,
                    host_parsed.netloc,
                    parsed.path,
                    parsed.params,
                    parsed.query,
                    parsed.fragment,
                )
            )
    else:
        for host in PRIMARY_HOSTS:
            yield urljoin(host, url)


def get(url: str, **kw) -> requests.Response:
    last_exc: Exception | None = None
    for candidate in _candidate_urls(url):
        for attempt in range(1, RETRIES + 2):
            try:
                response = requests.get(candidate, timeout=30, **kw)
                response.raise_for_status()
                if candidate != url:
                    log(f"Fetched {url} via fallback host {candidate}")
                return response
            except requests.RequestException as exc:  # pragma: no cover - network edge cases
                last_exc = exc
                log(
                    f"Request failed for {candidate} (attempt {attempt}/{RETRIES + 1}): {exc}",
                    level=logging.WARNING,
                )
                time.sleep(0.4)
        log(f"Switching host for {url}", level=logging.INFO)
    if last_exc:
        raise last_exc
    raise RuntimeError(f"Unable to fetch {url}")

def money_cents_to_str(v):
    # Rudes prices look like cents; keep “$185.00” formatting if cents given;
    # if already dollars, just str() it.
    if v is None: return ""
    try:
        iv = int(v)
        return f"${iv/100:.2f}"
    except Exception:
        return str(v)

def date_only_mmddyy(published_at: Optional[str]) -> str:
    if not published_at: return ""
    # examples: 2025-09-16T08:50:39-07:00
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", published_at)
    if not m: return ""
    y, mm, dd = m.group(1), m.group(2), m.group(3)
    return f"{mm}/{dd}/{y[-2:]}"

def clean_html_to_text(h: str) -> str:
    if not h: return ""
    # very light cleanup; avoid full HTML parser to keep deps light
    txt = html.unescape(re.sub(r"<br\s*/?>", " ", h, flags=re.I))
    txt = re.sub(r"<[^>]+>", " ", txt)
    txt = re.sub(r"\s+", " ", txt).strip()
    return txt

def parse_number_like(s: str) -> Optional[str]:
    """Turn 12 1/2, 12½, 12 3/4, 33 1/2\", 21\" into a plain number string. Keep as-is if unknown."""
    if not s: return ""
    s = s.replace("”","\"").replace("″","\"").replace("’","'").replace("“","\"")
    s = s.replace("½"," 1/2").replace("¼"," 1/4").replace("¾"," 3/4")
    s = s.replace('"', '').strip()
    # extract leading number / number fraction
    m = re.search(r"(-?\d+(?:\.\d+)?)(?:\s+(\d+)\/(\d+))?", s)
    if not m: 
        return s.strip()
    base = float(m.group(1))
    if m.group(2) and m.group(3):
        base += float(m.group(2)) / float(m.group(3))
    # trim .0
    if abs(base - round(base)) < 1e-6:
        return str(int(round(base)))
    return f"{base:.2f}".rstrip("0").rstrip(".")

# --------- PRODUCTS ----------
def fetch_all_products() -> List[dict]:
    allp = []
    page = 1
    while True:
        url = f"{PRODUCTS_PATH}?limit=250&page={page}"
        r = get(url)
        data = r.json()
        prods = data.get("products") or []
        log(f"[products.json] page {page} -> {len(prods)} products")
        if not prods:
            break
        allp.extend(prods)
        if len(prods) < 250:
            break
        page += 1
        time.sleep(0.2)
    log(f"[products.json] TOTAL products: {len(allp)}")
    return allp

# --------- PDP QUANTITIES FROM _ReStockConfig ----------
VARIANT_BLOCK_RE = re.compile(
    r"variants\s*:\s*\[(?P<body>.*?)\]\s*,?\s*\}", re.S)

# Object snippets like:
# { id: 507..., ... quantity: 2, }
VARIANT_OBJ_RE = re.compile(
    r"\{[^{}]*?\bid\s*:\s*(?P<id>\d+)[^{}]*?\bquantity\s*:\s*(?P<qty>-?\d+)[^{}]*?\}", re.S)

def extract_restock_quantities_from_html(html_text: str) -> Dict[str,int]:
    """
    Parse the inline JS (not JSON). Return {variant_id_str: quantity_int}
    """
    # narrow to the _ReStockConfig.product block
    # either "var _ReStockConfig = ..." or just "_ReStockConfig.product = { ... }"
    if "_ReStockConfig.product" not in html_text:
        return {}
    # Get the "variants: [ ... ]" slice first
    m = VARIANT_BLOCK_RE.search(html_text)
    if not m:
        return {}
    body = m.group("body")
    out = {}
    for mo in VARIANT_OBJ_RE.finditer(body):
        vid = mo.group("id")
        qty = int(mo.group("qty"))
        out[vid] = qty
    return out

# --------- MEASUREMENTS (body_html first; OCR optional) ----------
def extract_measures_from_body(body_html: str) -> Tuple[str,str,str]:
    """Try to read Front Rise / Inseam / Leg Opening from body_html text."""
    txt = clean_html_to_text(body_html or "")
    if not txt:
        return ("","","")
    # Try several labels
    def grab(label_list):
        for lab in label_list:
            m = re.search(rf"{lab}\s*[:\-]\s*([0-9][^,. ;]*)", txt, flags=re.I)
            if m:
                return parse_number_like(m.group(1))
        return ""
    rise   = grab(["Front Rise","Rise"])
    inseam = grab(["Inseam"])
    leg    = grab(["Leg Opening","Leg Openning","Opening"])
    return (rise, inseam, leg)

# (Optional) OCR pipeline stub – left in, but off unless OCR_MEASUREMENTS=True
def ocr_measurements_from_pdp_images(html_text: str, size_value: str) -> Tuple[str,str,str]:
    """
    Look for an image URL with triple underscores '___' and OCR it.
    Then select the cell where row name == Rise/Inseam/Leg Opening
    and column header == size_value.
    If not found, return ("","","").
    """
    if not OCR_MEASUREMENTS:
        return ("","","")
    try:
        import pytesseract
        from PIL import Image
        import io
        pytesseract.pytesseract.tesseract_cmd = TESSERACT_PATH
    except Exception as e:
        log(f"[OCR] disabled or libraries missing: {e}")
        return ("","","")

    # Find any image URL with '___' in it
    img_urls = re.findall(r"https?://[^\"']*___[^\"'> )]+", html_text)
    if not img_urls:
        return ("","","")
    # Pick the first; download and OCR
    try:
        resp = get(img_urls[0])
        from PIL import Image
        im = Image.open(io.BytesIO(resp.content))
        text = pytesseract.image_to_string(im)
        # Extremely heuristic parsing:
        # find header row with sizes, then rows with Rise/Inseam/Leg Opening, then pick value under size_value
        # (For now, just return blanks; expand later if needed)
        return ("","","")
    except Exception as e:
        log(f"[OCR] failed: {e}")
        return ("","","")

# --------- SIZE / COLOR NORMALIZATION ----------
def normalize_size_color(variant: dict) -> Tuple[str,str]:
    """
    Return (size, color) choosing from option1/option2/title so that size is in ALLOWED_SIZES.
    """
    t  = (variant.get("title") or "")
    o1 = (variant.get("option1") or "").strip()
    o2 = (variant.get("option2") or "").strip()

    cand = []
    if o1: cand.append(o1)
    if o2: cand.append(o2)
    # also split title like "23 / Lagoon Blue"
    if "/" in t:
        parts = [x.strip() for x in t.split("/") if x.strip()]
        cand.extend(parts)

    size_found = ""
    for c in cand:
        if c.lower() in {x.lower() for x in ALLOWED_SIZES}:
            size_found = c
            break

    if not size_found:
        # best-effort: pick a 2-digit number 23..34 as size
        m = re.search(r"\b(2[3-9]|3[0-4])\b", " ".join(cand))
        if m: size_found = m.group(1)

    # choose color as "the other thing" when possible
    color = ""
    if size_found and size_found == o1 and o2:
        color = o2
    elif size_found and size_found == o2 and o1:
        color = o1
    else:
        # from title pieces
        if "/" in t:
            parts = [x.strip() for x in t.split("/") if x.strip()]
            for p in parts:
                if p != size_found:
                    color = p
                    break

    return size_found, color

# --------- MAIN ----------
def run():
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    date_col  = datetime.now().strftime("%Y-%m-%d")
    time_col  = datetime.now().strftime("%H:%M:%S")

    csv_path = OUTPUT_DIR / f"RUDES_{timestamp}.csv"

    fields = [
        "Style Id","Handle","Published At","Product","Product Type","Vendor","Description",
        "Variant Title","Color","Size","Front Rise","Inseam","Leg Opening",
        "Price","Compare at Price","Available for Sale",
        "Quantity Available","Quantity of style",
        "SKU","Image URL","SKU URL","Date","Time"
    ]

    with csv_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()

        products = fetch_all_products()

        for idx, p in enumerate(products, start=1):
            handle   = p.get("handle","")
            prod_id  = p.get("id","")
            title    = p.get("title","") or ""
            vendor   = p.get("vendor","") or ""
            body     = p.get("body_html") or ""
            published= p.get("published_at") or ""
            product_type = title.split()[-1] if title.strip() else ""

            # PDP HTML (for quantities, and OCR if enabled)
            try:
                pdp_html = get(PDP_PATH_TEMPLATE.format(handle=handle)).text
            except Exception as e:
                log(f"[PDP] {handle} fetch error: {e}")
                pdp_html = ""

            qty_map = {}
            try:
                qty_map = extract_restock_quantities_from_html(pdp_html)
                if not qty_map:
                    log(f"[ReStock] no variant quantities parsed on {handle}")
            except Exception as e:
                log(f"[ReStock] extraction error on {handle}: {e}")

            # Body_html measures first
            rise, inseam, leg_open = extract_measures_from_body(body)

            # Variants + style-level total
            variants = p.get("variants") or []
            style_total = 0
            has_qty = False
            # First pass: total (using qty_map)
            for v in variants:
                vid = str(v.get("id",""))
                if vid in qty_map:
                    style_total += qty_map[vid]
                    has_qty = True

            # Images: pick first product image URL (not the OCR guide)
            images = p.get("images") or []
            first_img = ""
            if images:
                first_img = images[0].get("src") or ""

            for v in variants:
                vid   = str(v.get("id",""))
                vtitle= v.get("title") or ""
                price = v.get("price")
                comp  = v.get("compare_at_price")
                avail = v.get("available")

                size, color = normalize_size_color(v)
                # fallback to option fields if needed
                if not size:  size  = v.get("option1") or ""
                if not color: color = v.get("option2") or ""

                inv = qty_map.get(vid, "")

                # OCR (only if still empty AND you turn it on)
                if OCR_MEASUREMENTS and (not rise or not inseam or not leg_open):
                    r2,i2,l2 = ocr_measurements_from_pdp_images(pdp_html, size)
                    rise   = rise   or r2
                    inseam = inseam or i2
                    leg_open = leg_open or l2

                w.writerow({
                    "Style Id": prod_id,
                    "Handle": handle,
                    "Published At": date_only_mmddyy(published),
                    "Product": title,
                    "Product Type": product_type,
                    "Vendor": vendor,
                    "Description": clean_html_to_text(body),
                    "Variant Title": v.get("sku") or vtitle,  # you asked to use sku here
                    "Color": color,
                    "Size": size,
                    "Front Rise": rise,
                    "Inseam": inseam,
                    "Leg Opening": leg_open,
                    "Price": money_cents_to_str(price),
                    "Compare at Price": money_cents_to_str(comp),
                    "Available for Sale": avail,
                    "Quantity Available": inv,
                    "Quantity of style": style_total if has_qty else "",
                    "SKU": vid,
                    "Image URL": first_img,
                    "SKU URL": f"{BASE}/products/{handle}",
                    "Date": date_col,
                    "Time": time_col
                })

            log(f"[{idx}/{len(products)}] {handle} -> {len(variants)} variants (style_total={'N/A' if not has_qty else style_total})")

    log(f"CSV:   {csv_path}")

    # -------- Append to Excel (no formulas touched) --------
    try:
        from openpyxl import Workbook, load_workbook
        if EXCEL_PATH.exists():
            wb = load_workbook(EXCEL_PATH)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(fields)  # header

        with csv_path.open(newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            next(reader, None)
            for row in reader:
                ws.append(row)

        wb.save(EXCEL_PATH)
        log(f"Excel: {EXCEL_PATH}")
    except Exception as e:
        log(f"[Excel] append failed: {e}")

if __name__ == "__main__":
    run()
