# -*- coding: utf-8 -*-
"""
Export all denim_analytics rows for ONE brand to two files:

  1. [BRAND]_DB_Values_YYYY-MM-DD_HH-MM-SS.json  - compact, lossless, meant to
     be handed to Claude so it can load real data and test database-fix code.
  2. [BRAND]_DB_Values_YYYY-MM-DD_HH-MM-SS.xlsx  - one worksheet per table for
     you to review by eye.

Both files land in the SAME folder as this script. Timestamp is Central time.
Every log line is prefixed with a Central-time stamp like [2026-08-13 15:09:24];
the big tables (style_metrics, variant_metrics) also log progress every 10%.

CONFIGURE below: set BRAND, then flip each table True (pull it) / False (skip).

RUN (Windows Command Prompt - the black "cmd" window):
    pip install pymssql openpyxl
    set SQL_USERNAME=carrieboromisa
    set SQL_PASSWORD=your-password
    python "C:\\path\\to\\export_db_values.py"
(or fill HARDCODED_USERNAME / HARDCODED_PASSWORD below to skip the set lines)
"""
from __future__ import annotations

import json
import os
import sys
import datetime as dt
from decimal import Decimal
from pathlib import Path

# ===========================================================================
# CONFIGURE
# ===========================================================================
BRAND = "TRIARCHY"

# True  = include this table in the export
# False = skip it
TABLES = {
    "dbo.lookup":            True,
    "dbo.style_info":        True,
    "dbo.style_metrics":     True,
    "dbo.variant_metrics":   True,
    "dbo.image_url_history": False,   # flip True if you want image history too
}

# Tables that get 10%-increment progress logging (the big, slow ones).
PROGRESS_TABLES = {"style_metrics", "variant_metrics"}

WRITE_JSON = True    # the Claude-readable file
WRITE_XLSX = True    # the Excel review file
MAX_ROWS_PER_TABLE = None   # e.g. 5000 to cap; None = no cap (pull everything)

# DO_DEDUPE master switch:
#   True  -> collapse the metrics tables per DEDUPE_RULES below (one row per
#            unique record, keeping the OLDEST captured_datetime).
#   False -> keep ALL values; every row is exported, no deduping at all.
DO_DEDUPE = True

# The metrics tables keep a row per daily capture, which makes the exports too
# large to share. Collapse each to one row per unique record: rows sharing the
# same key are deduped down to the one with the OLDEST captured_datetime. The
# key itself is NOT added to the output, and no other column changes.
# (Only used when DO_DEDUPE = True; set DEDUPE_RULES = {} to keep a table's rows
# even while other tables dedupe.)
#   variant_metrics -> V_UNIQUE_KEY
#   style_metrics   -> S_UNIQUE_KEY
DEDUPE_RULES = {
    "variant_metrics": (["brand", "sku_shopify", "sku_brand", "barcode",
                         "variant_title", "size"], "captured_datetime"),
    "style_metrics":   (["brand", "style_id", "product_name", "handle", "color",
                         "style_name", "inseam", "inseam_label"], "captured_datetime"),
}

# Database connection --------------------------------------------------------
SQL_SERVER   = os.environ.get("SQL_SERVER",   "denim-sql.database.windows.net")
SQL_DATABASE = os.environ.get("SQL_DATABASE", "denim_analytics")
HARDCODED_USERNAME = ""
HARDCODED_PASSWORD = ""
SQL_USERNAME = HARDCODED_USERNAME or os.environ.get("SQL_USERNAME", "")
SQL_PASSWORD = HARDCODED_PASSWORD or os.environ.get("SQL_PASSWORD", "")

SCRIPT_DIR = Path(os.path.dirname(os.path.abspath(__file__)))
FETCH_CHUNK = 2000   # rows per fetchmany batch (for progress reporting)

# pymssql sends Python str params as NVARCHAR, but [brand] is VARCHAR. The
# comparison varchar_col = nvarchar_param converts the COLUMN (non-SARGable),
# which throws away the index and scans the ENTIRE multi-brand table just to find
# one brand - then the dedupe adds a ROW_NUMBER sort on top, so variant_metrics
# times out. Casting the parameter to varchar keeps the brand filter an index
# seek. (This is why it worked before the dedupe: a plain streaming scan finished
# in time, but scan + big sort does not.)
VC = "CAST(%s AS varchar(255))"


# ===========================================================================
# Central-time logging (every line is timestamped)
# ===========================================================================
def _resolve_tz():
    try:
        from zoneinfo import ZoneInfo
        return ZoneInfo("America/Chicago"), False
    except Exception:
        return dt.timezone(dt.timedelta(hours=-6)), True


_TZ, _TZ_FALLBACK = _resolve_tz()


def now_central() -> dt.datetime:
    return dt.datetime.now(_TZ)


def log(msg: str = "") -> None:
    print(f"[{now_central().strftime('%Y-%m-%d %H:%M:%S')}] {msg}")


# ===========================================================================
# Value serialization
# ===========================================================================
def json_safe(v):
    if v is None:
        return None
    if isinstance(v, Decimal):
        return str(v)
    if isinstance(v, (dt.datetime, dt.date)):
        return v.isoformat()
    if isinstance(v, (bytes, bytearray)):
        return v.decode("latin-1", errors="replace")
    return v


def xlsx_safe(v):
    if isinstance(v, Decimal):
        return float(v)
    if isinstance(v, (bytes, bytearray)):
        return v.decode("latin-1", errors="replace")
    if isinstance(v, str) and len(v) > 32767:   # Excel hard cell limit
        return v[:32764] + "..."
    return v   # None, int, float, str, datetime/date are fine for openpyxl


def table_short(name: str) -> str:
    """'dbo.style_info' -> 'style_info'."""
    return name.split(".")[-1].strip().strip("[]")


class Decile:
    """Emit a log line each time progress crosses the next 10% boundary."""
    def __init__(self, total: int, label: str, enabled: bool):
        self.total = total
        self.label = label
        self.enabled = enabled and total > 0
        self.next = 10

    def update(self, done: int) -> None:
        if not self.enabled:
            return
        while self.next <= 100 and done >= self.total * self.next / 100:
            log(f"   {self.label}: {self.next}% ({min(done, self.total)}/{self.total})")
            self.next += 10


# ===========================================================================
# Fetch one table (chunked, with optional 10% progress)
# ===========================================================================
def fetch_table(cur, tbl: str, progress: bool):
    top = f"TOP {int(MAX_ROWS_PER_TABLE)} " if MAX_ROWS_PER_TABLE else ""

    # Deduped tables are collapsed IN SQL (ROW_NUMBER per key, oldest
    # captured_datetime) so only the survivors cross the wire - the full history
    # never leaves the server, which avoids the huge-fetch timeout.
    rule = DEDUPE_RULES.get(tbl.lower()) if DO_DEDUPE else None
    if rule:
        return _fetch_deduped(cur, tbl, rule, top)

    total = None
    if progress:
        cur.execute(f"SELECT COUNT(*) FROM [dbo].[{tbl}] WHERE brand = {VC}", (BRAND,))
        total = cur.fetchone()[0]
        if MAX_ROWS_PER_TABLE:
            total = min(total, int(MAX_ROWS_PER_TABLE))
        log(f"{tbl}: fetching {total} rows...")
    else:
        log(f"{tbl}: fetching...")

    cur.execute(f"SELECT {top}* FROM [dbo].[{tbl}] WHERE brand = {VC} ORDER BY 1", (BRAND,))
    cols = [d[0] for d in cur.description]

    if not progress:
        rows = cur.fetchall()
    else:
        rows = []
        dec = Decile(total or 0, f"{tbl} fetch", True)
        while True:
            chunk = cur.fetchmany(FETCH_CHUNK)
            if not chunk:
                break
            rows.extend(chunk)
            dec.update(len(rows))
    return cols, rows


def _fetch_deduped(cur, tbl, rule, top):
    """Server-side dedupe: keep one row per key with the oldest date column."""
    key_cols, date_col = rule
    # real column list, in order (so the output matches a plain SELECT *)
    cur.execute("SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS "
                "WHERE TABLE_NAME = %s AND TABLE_SCHEMA = 'dbo' "
                "ORDER BY ORDINAL_POSITION", (tbl,))
    allcols = [r[0] for r in cur.fetchall()]
    pk = allcols[0]
    collist = ", ".join(f"[{c}]" for c in allcols)
    partition = ", ".join(f"[{c}]" for c in key_cols)
    log(f"{tbl}: fetching (deduped in SQL on {len(key_cols)}-col key, "
        f"oldest {date_col})...")
    cur.execute(
        f"WITH ranked AS ("
        f"  SELECT {collist}, ROW_NUMBER() OVER ("
        f"    PARTITION BY {partition} "
        f"    ORDER BY [{date_col}] ASC, [{pk}] ASC) AS _rn "
        f"  FROM [dbo].[{tbl}] WHERE [brand] = {VC}) "
        f"SELECT {top}{collist} FROM ranked WHERE _rn = 1", (BRAND,))
    cols = [d[0] for d in cur.description]
    rows = cur.fetchall()
    return cols, rows


# ===========================================================================
# Main
# ===========================================================================
def main() -> None:
    selected = [t for t, on in TABLES.items() if on]
    if not selected:
        sys.exit("Nothing to do - every table in TABLES is False.")
    if not SQL_USERNAME or not SQL_PASSWORD:
        sys.exit("ERROR: set SQL_USERNAME and SQL_PASSWORD (env vars) or fill "
                 "HARDCODED_USERNAME/PASSWORD near the top of this file.")

    stamp = now_central().strftime("%Y-%m-%d_%H-%M-%S")
    base = f"{BRAND}_DB_Values_{stamp}"
    log("=" * 58)
    log(f"EXPORT {BRAND} DB VALUES")
    log("Tables: " + ", ".join(table_short(t) for t in selected))
    if _TZ_FALLBACK:
        log("(tzdata not found - timestamps use a fixed UTC-6; "
            "pip install tzdata for exact Central time)")
    log("=" * 58)

    import pymssql
    log("Connecting to database...")
    conn = pymssql.connect(server=SQL_SERVER, user=SQL_USERNAME,
                           password=SQL_PASSWORD, database=SQL_DATABASE,
                           timeout=600, login_timeout=60)
    cur = conn.cursor()

    export: dict = {}
    for t in selected:
        tbl = table_short(t)
        progress = tbl.lower() in {p.lower() for p in PROGRESS_TABLES}
        try:
            cols, rows = fetch_table(cur, tbl, progress)
        except Exception as exc:
            log(f"{tbl}: SKIPPED - query failed: {exc}")
            continue
        export[tbl] = {"columns": cols, "row_count": len(rows), "rows": rows}
        log(f"{tbl}: done - {len(rows)} rows, {len(cols)} columns")

    conn.close()

    if not export:
        sys.exit("No data pulled - nothing written.")

    # ---- 1) Claude-readable JSON ------------------------------------------
    if WRITE_JSON:
        log("Writing Claude JSON...")
        payload = {
            "brand": BRAND,
            "generated_at_central": stamp,
            "database": SQL_DATABASE,
            "tables": {
                tbl: {
                    "columns": d["columns"],
                    "row_count": d["row_count"],
                    "rows": [[json_safe(v) for v in row] for row in d["rows"]],
                }
                for tbl, d in export.items()
            },
        }
        json_path = SCRIPT_DIR / f"{base}.json"
        with json_path.open("w", encoding="utf-8") as fh:
            json.dump(payload, fh, ensure_ascii=False, indent=1)
        log(f"Claude JSON written: {json_path}")

    # ---- 2) Excel review workbook -----------------------------------------
    if WRITE_XLSX:
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
        except ImportError:
            log("Excel skipped - run 'pip install openpyxl' to enable it "
                "(the JSON was still written).")
        else:
            log("Writing Excel workbook...")
            wb = Workbook()
            wb.remove(wb.active)
            for tbl, d in export.items():
                ws = wb.create_sheet(title=tbl[:31])
                ws.append(d["columns"])
                progress = tbl.lower() in {p.lower() for p in PROGRESS_TABLES}
                dec = Decile(d["row_count"], f"{tbl} xlsx", progress)
                for i, row in enumerate(d["rows"], start=1):
                    ws.append([xlsx_safe(v) for v in row])
                    dec.update(i)
                ws.freeze_panes = "A2"
                for i, cname in enumerate(d["columns"], start=1):
                    ws.column_dimensions[get_column_letter(i)].width = \
                        min(max(len(str(cname)) + 2, 12), 60)
                log(f"{tbl}: sheet written ({d['row_count']} rows)")
            xlsx_path = SCRIPT_DIR / f"{base}.xlsx"
            log("Saving Excel file (this can take a moment for large tables)...")
            wb.save(xlsx_path)
            log(f"Excel file written: {xlsx_path}")

    log("Done.")


if __name__ == "__main__":
    main()
