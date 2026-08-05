"""
Backfills blank style_id, style_name, style_name_grouping, sku_shopify, and
barcode values across dbo.lookup, dbo.style_info, dbo.style_metrics,
dbo.variant_metrics in denim_analytics, using an authoritative fill-value
list keyed by dbo.lookup.lookup_id (e.g. a "Fill in ... missing values.xlsx"
export).

This is a thin runner around fill_missing_style_fields.sql: it populates
#fill_values from the input file using parameterized inserts, then executes
the rest of that script unmodified so there is a single source of truth for
the backfill logic.

Credentials are read from environment variables only -- never hardcode them
here or pass them on the command line, since this repo is version controlled.

Expected input columns (.xlsx or .csv), matching the "Fill in ... missing
values" sheet layout:
    lookup_id
    If style_id is blank, fill with
    If style_name is blank, fill with
    If style_name_grouping is blank, fill with
    If sku_shopify is blank, fill with
    If barcode is blank, fill with
(any other columns are ignored)

Usage:
    export SQL_SERVER=denim-sql.database.windows.net
    export SQL_DATABASE=denim_analytics
    export SQL_USERNAME=...
    export SQL_PASSWORD=...
    pip install pyodbc openpyxl
    python sql/fill_missing_style_fields.py --input "Fill_in_Good_American_missing_values.xlsx"          # dry run
    python sql/fill_missing_style_fields.py --input "Fill_in_Good_American_missing_values.xlsx" --apply  # commits
"""

import argparse
import csv
import os
import sys
from pathlib import Path

SQL_FILE = Path(__file__).resolve().parent / "fill_missing_style_fields.sql"
STEP0_MARKER = "-- === END STEP 0 ==="

COLUMN_MAP = {
    "lookup_id": "lookup_id",
    "If style_id is blank, fill with": "new_style_id",
    "If style_name is blank, fill with": "new_style_name",
    "If style_name_grouping is blank, fill with": "new_style_name_grouping",
    "If sku_shopify is blank, fill with": "new_sku_shopify",
    "If barcode is blank, fill with": "new_barcode",
}
FILL_COLUMNS = ["new_style_id", "new_style_name", "new_style_name_grouping", "new_sku_shopify", "new_barcode"]


def _clean(value):
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def load_fill_values_csv(path: Path) -> list[dict]:
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        missing = [c for c in COLUMN_MAP if c not in (reader.fieldnames or [])]
        if missing:
            raise SystemExit(f"{path} is missing expected column(s): {missing}")
        rows = []
        for row in reader:
            rows.append({COLUMN_MAP[k]: _clean(row[k]) for k in COLUMN_MAP})
        return rows


def load_fill_values_xlsx(path: Path, sheet: str | None) -> list[dict]:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    missing = [c for c in COLUMN_MAP if c not in header]
    if missing:
        raise SystemExit(f"{path} sheet {ws.title!r} is missing expected column(s): {missing}")
    col_index = {name: header.index(name) for name in COLUMN_MAP}
    rows = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        if raw[col_index["lookup_id"]] is None:
            continue
        rows.append({dest: _clean(raw[col_index[src]]) for src, dest in COLUMN_MAP.items()})
    return rows


def load_fill_values(path: Path, sheet: str | None) -> list[dict]:
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        rows = load_fill_values_xlsx(path, sheet)
    elif path.suffix.lower() == ".csv":
        rows = load_fill_values_csv(path)
    else:
        raise SystemExit(f"Unsupported input file type: {path.suffix}")

    rows = [r for r in rows if r["lookup_id"] is not None]
    if not rows:
        raise SystemExit(f"No usable rows found in {path}")
    return rows


def load_script_body(apply_changes: bool) -> str:
    text = SQL_FILE.read_text(encoding="utf-8")
    if STEP0_MARKER not in text:
        raise SystemExit(f"Could not find marker {STEP0_MARKER!r} in {SQL_FILE}")
    body = text.split(STEP0_MARKER, 1)[1]
    if apply_changes:
        body = body.replace("DECLARE @dry_run BIT = 1;", "DECLARE @dry_run BIT = 0;", 1)
    return body


def connect():
    import pyodbc

    server = os.environ.get("SQL_SERVER")
    database = os.environ.get("SQL_DATABASE")
    username = os.environ.get("SQL_USERNAME")
    password = os.environ.get("SQL_PASSWORD")
    missing = [name for name, val in [
        ("SQL_SERVER", server), ("SQL_DATABASE", database),
        ("SQL_USERNAME", username), ("SQL_PASSWORD", password),
    ] if not val]
    if missing:
        raise SystemExit(f"Missing required environment variable(s): {', '.join(missing)}")

    driver = os.environ.get("SQL_ODBC_DRIVER", "ODBC Driver 18 for SQL Server")
    conn_str = (
        f"DRIVER={{{driver}}};SERVER={server};DATABASE={database};"
        f"UID={username};PWD={password};Encrypt=yes;TrustServerCertificate=no;"
        "Connection Timeout=30;"
    )
    return pyodbc.connect(conn_str, autocommit=True)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--input", required=True, type=Path, help="Fill-value spreadsheet (.xlsx) or CSV export of it")
    parser.add_argument("--sheet", default=None, help="Sheet name (defaults to the active sheet); .xlsx only")
    parser.add_argument("--apply", action="store_true", help="Commit changes (default is a dry run that rolls back)")
    args = parser.parse_args()

    fill_rows = load_fill_values(args.input, args.sheet)
    script_body = load_script_body(args.apply)

    print(f"Loaded {len(fill_rows)} fill-value row(s) from {args.input}")
    print("Mode: APPLY (will commit)" if args.apply else "Mode: DRY RUN (will roll back)")

    conn = connect()
    try:
        cursor = conn.cursor()
        cursor.execute(
            "IF OBJECT_ID('tempdb..#fill_values') IS NOT NULL DROP TABLE #fill_values;"
            "CREATE TABLE #fill_values ("
            "  lookup_id INT PRIMARY KEY,"
            "  new_style_id VARCHAR(100) COLLATE DATABASE_DEFAULT NULL,"
            "  new_style_name VARCHAR(255) COLLATE DATABASE_DEFAULT NULL,"
            "  new_style_name_grouping VARCHAR(100) COLLATE DATABASE_DEFAULT NULL,"
            "  new_sku_shopify VARCHAR(100) COLLATE DATABASE_DEFAULT NULL,"
            "  new_barcode VARCHAR(100) COLLATE DATABASE_DEFAULT NULL"
            ");"
        )
        cursor.executemany(
            "INSERT INTO #fill_values (lookup_id, new_style_id, new_style_name, new_style_name_grouping, new_sku_shopify, new_barcode) "
            "VALUES (?, ?, ?, ?, ?, ?)",
            [(int(r["lookup_id"]), *(r[c] for c in FILL_COLUMNS)) for r in fill_rows],
        )
        cursor.execute(script_body)

        # The script can emit more than one result set (the summary counts,
        # and -- if any sku_shopify values collided with an existing row --
        # a conflicts report). Print every result set the driver hands back.
        has_more = True
        while has_more:
            if cursor.description is not None:
                columns = [c[0] for c in cursor.description]
                rows = cursor.fetchall()
                print(f"--- {columns} ---")
                for row in rows:
                    print(dict(zip(columns, row)))
            has_more = cursor.nextset()

        for msg in conn.messages:
            print(msg[1])
    finally:
        conn.close()


if __name__ == "__main__":
    sys.exit(main())
