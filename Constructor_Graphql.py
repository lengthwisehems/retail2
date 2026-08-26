"""Constructor + Shopify Storefront GraphQL workbook probe.

Outputs one Excel file with two tabs:
- Constructor: one row per variant with product-level + variant-level Constructor fields
- GraphQL: product/variant rows with custom filter columns and ordered base columns
"""

from __future__ import annotations

import json
import logging
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Set, Tuple
from urllib.parse import urlparse

import requests
from openpyxl import Workbook
from requests.adapters import HTTPAdapter, Retry

# ---------------------------------------------------------------------------
# Inputs
# ---------------------------------------------------------------------------
BRAND = "AG Jeans"
COLLECTION_URL = [
    "https://www.agjeans.com/collections/womens-denim",
    "https://www.agjeans.com/collections/womens-sale",
]
MYSHOPIFY = "agjeans-store.myshopify.com"
GRAPHQL = "https://www.agjeans.com/api/unstable/graphql.json"
X_SHOPIFY_STOREFRONT_ACCESS_TOKEN = "ffae8e47a84566aa6fa059dfc56c7c56"

CATEGORY_FILTER = ["Jeans"]
PRODUCT_TYPE_FILTER = ["WOMENS BOTTOMS"]

# Constructor inputs
CONSTRUCTOR_API_KEY = "key_Ai9lmSZcQbh1bfYa"
CONSTRUCTOR_CLIENT_ID = "124bb124-e8d8-444b-9186-eaae4100af9f"
CONSTRUCTOR_SESSION = "20"
CONSTRUCTOR_BROWSE_ENDPOINT = "https://ac.cnstrc.com/browse/group_id"
CONSTRUCTOR_RESULTS_PER_PAGE = 26
CONSTRUCTOR_HIDDEN_FIELDS = ["prices.price_US", "compareAtPrices.compareprice_US"]
CONSTRUCTOR_EXTRA_PARAMS: Dict[str, str] = {}

# The "online"/warehouse fulfillment location. Everything else in the
# inventory array is a physical retail store. Edit here if the location name
# in Constructor changes.
ONLINE_LOCATION_NAME = "AG Jeans"

# Constructor facets whose per-item value is not returned in item data. We
# resolve them by querying the browse endpoint filtered to each facet value and
# tagging the matching products (one request per value). Add/remove names here.
CONSTRUCTOR_FACET_TAG_FIELDS = [
    "sizeOption",
    "gender",
    "country_code_of_origin",
    "cut",
    "influencer",
    "category",
    "variants_availability",
]

# ---------------------------------------------------------------------------
# GraphQL output ordering / skip rules
# ---------------------------------------------------------------------------
COLUMN_ORDER_BASE: Tuple[str, ...] = (
    "product.id",
    "product.handle",
    "product.published_at",
    "product.created_at",
    "product.title",
    "product.productType",
    "product.category.name",
    "product.tags_all",
    "product.vendor",
    "product.description",
    "product.descriptionHtml",
    "variant.title",
    "variant.option1",
    "variant.option2",
    "variant.option3",
    "variant.price",
    "variant.compare_at_price",
    "product.priceRange",
    "variant.available",
    "variant.quantityAvailable",
    "product.totalInventory",
    "variant.id",
    "variant.sku",
    "variant.barcode",
    "product.featuredImage",
    "product.onlineStoreUrl",
)

DEFAULT_FORBIDDEN_FIELDS: Dict[str, Set[str]] = {
    "ProductVariant": {
        "components",
        "groupedBy",
        "quantityPriceBreaks",
        "sellingPlanAllocations",
        "sellingPlanGroups",
        "storeAvailability",
    }
}


CONSTRUCTOR_COLUMN_ORDER: Tuple[str, ...] = (
    "product.id",
    "product.handle",
    "product.title",
    "product.title.v2",
    "product.title.v3",
    "product.title.v4",
    "product.productType",
    "product.category",
    "product.gender",
    "product.cut",
    "product.influencer",
    "product.country_code_of_origin",
    "product.variants_availability",
    "product.sizeOption",
    "product.sort",
    "product.labels",
    "product.matchedTerms",
    "product.tags_all",
    "product.description",
    "product.color",
    "product.color_name",
    "product.rise",
    "product.closure",
    "product.onlineStoreUrl",
    "variant.price",
    "variant.compare_at_price",
    "product.notifyBIS",
    "variant.quantityAvailable.Instore",
    "variant.quantityAvailable.Online",
    "variant.available",
    "variant.id",
    "variant.sku",
    "variant.size",
    "variant.bottoms_size",
    "source",
    "product.material",
    "product.fabric",
    "product.mill",
    "product.country",
    "collection.handles",
    "collection.url",
    "collection.handle",
    "product.highlight",
    "product.capsule",
    "product.image",
    "product.images.v2",
    "varient.raw",
)

EXTRA_FORBIDDEN_COLUMNS: Set[str] = {
    "product.collections.pageInfo.endCursor",
    "product.collections.pageInfo.hasNextPage",
    "product.encodedVariantAvailability",
    "product.encodedVariantExistence",
    "product.featuredImage.height",
    "product.featuredImage.thumbhash",
    "product.featuredImage.width",
    "product.images.pageInfo.endCursor",
    "product.images.pageInfo.hasNextPage",
    "product.isGiftCard",
    "product.media.pageInfo.endCursor",
    "product.media.pageInfo.hasNextPage",
    "products_edge_cursor",
    "variant.currentlyNotInStock",
    "variant.image.height",
    "variant.image.id",
    "variant.image.thumbhash",
    "variant.image.width",
    "variant.quantityRule.minimum",
    "variant_edge_cursor",
    "variants_endCursor",
    "variants_hasNextPage",
}

# ---------------------------------------------------------------------------
# Paths / logging
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = BASE_DIR / "Output"
OUTPUT_DIR.mkdir(exist_ok=True)

BRAND_SLUG = re.sub(r"[^a-z0-9]+", "_", BRAND.lower()).strip("_") or "brand"
LOG_PATH = OUTPUT_DIR / f"{BRAND_SLUG}_constructor_graphql.log"

REQUEST_TIMEOUT = 30


def configure_logger() -> logging.Logger:
    logger = logging.getLogger("constructor_graphql")
    logger.setLevel(logging.INFO)
    logger.handlers.clear()

    formatter = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")

    fh = logging.FileHandler(LOG_PATH, encoding="utf-8")
    fh.setFormatter(formatter)
    logger.addHandler(fh)

    sh = logging.StreamHandler()
    sh.setFormatter(formatter)
    logger.addHandler(sh)
    return logger


def build_session() -> requests.Session:
    session = requests.Session()
    retry = Retry(
        total=5,
        backoff_factor=0.5,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=["GET", "POST"],
    )
    adapter = HTTPAdapter(max_retries=retry)
    session.mount("http://", adapter)
    session.mount("https://", adapter)
    session.headers.update(
        {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/127.0.0.0 Safari/537.36"
            )
        }
    )
    return session


def collection_handle_from_url(url: str) -> str:
    path = urlparse(url).path.strip("/")
    parts = [p for p in path.split("/") if p]
    if len(parts) >= 2 and parts[0] == "collections":
        return parts[1]
    return parts[-1] if parts else ""


def list_to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, list):
        cleaned = []
        for item in value:
            if isinstance(item, (dict, list)):
                cleaned.append(json.dumps(item, ensure_ascii=False))
            else:
                cleaned.append(str(item))
        return ", ".join(cleaned)
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def maybe_number(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return str(value)
    if isinstance(value, str):
        m = re.search(r"-?\d+(?:\.\d+)?", value)
        return m.group(0) if m else ""
    if isinstance(value, dict):
        for key in ("price_US", "compareprice_US", "amount"):
            if key in value:
                return maybe_number(value.get(key))
        for v in value.values():
            got = maybe_number(v)
            if got:
                return got
        return ""
    if isinstance(value, list):
        for item in value:
            got = maybe_number(item)
            if got:
                return got
        return ""
    return ""


def sum_inventory(value: Any) -> str:
    if value is None:
        return ""
    total = 0

    def walk(node: Any) -> None:
        nonlocal total
        if isinstance(node, dict):
            for k, v in node.items():
                if k.lower() == "available" and isinstance(v, (int, float)):
                    total += int(v)
                else:
                    walk(v)
        elif isinstance(node, list):
            for i in node:
                walk(i)

    walk(value)
    return str(total)


def _inventory_entries(value: Any) -> List[Dict[str, Any]]:
    """Constructor inventory is a list of {available, locationName} dicts."""
    if isinstance(value, list):
        return [e for e in value if isinstance(e, dict)]
    if isinstance(value, dict):
        return [value]
    return []


def sum_inventory_by_location(value: Any, *, online: bool) -> str:
    """Sum 'available' across inventory entries, split by fulfillment location.

    online=True  -> only the ONLINE_LOCATION_NAME location (warehouse/online).
    online=False -> every OTHER (physical store) location combined.
    """
    entries = _inventory_entries(value)
    if not entries:
        return ""  # no inventory data at all
    total = 0
    target = ONLINE_LOCATION_NAME.strip().lower()
    for entry in entries:
        available = entry.get("available")
        if not isinstance(available, (int, float)):
            continue
        is_online = str(entry.get("locationName") or "").strip().lower() == target
        if is_online == online:
            total += int(available)
    return str(total)


def size_from_sku(sku: Any) -> str:
    """Variant size is the SKU suffix after the final '.' (e.g. 'MIU1H63IOS.22'
    -> '22'). Returns '' when the SKU has no size suffix."""
    text = str(sku or "").strip()
    if "." not in text:
        return ""
    return text.rsplit(".", 1)[-1].strip()


# ---------------------------------------------------------------------------
# Constructor collection -> merged variant rows
# ---------------------------------------------------------------------------
CONSTRUCTOR_REMOVE_HEADERS = {
    "constructor.data.prices",
    "constructor.data.inventory",
    "constructor.data.compareAtPrices",
    "constructor.data.productmedia_v1",
    "constructor.is_slotted",
    "constructor.parent_sku",
    "constructor.parent_value",
    "constructor.variation_index",
    "row_type",
    "source",
}


def fetch_constructor_results_for_collection(
    session: requests.Session,
    collection_url: str,
    logger: logging.Logger,
) -> List[Dict[str, Any]]:
    handle = collection_handle_from_url(collection_url)
    if not handle:
        logger.warning("Could not derive collection handle from URL: %s", collection_url)
        return []

    base_params: List[Tuple[str, str]] = [
        ("c", "cio-ui-plp-1.6.2"),
        ("key", CONSTRUCTOR_API_KEY),
        ("i", CONSTRUCTOR_CLIENT_ID),
        ("s", CONSTRUCTOR_SESSION),
        ("num_results_per_page", str(CONSTRUCTOR_RESULTS_PER_PAGE)),
    ]
    for hf in CONSTRUCTOR_HIDDEN_FIELDS:
        base_params.append(("fmt_options[hidden_fields]", hf))
    for k, v in CONSTRUCTOR_EXTRA_PARAMS.items():
        base_params.append((k, str(v)))

    page = 1
    total_pages: Optional[int] = None
    all_results: List[Dict[str, Any]] = []

    while True:
        url = f"{CONSTRUCTOR_BROWSE_ENDPOINT}/{handle}"
        params = base_params + [("page", str(page))]
        resp = session.get(url, params=params, timeout=REQUEST_TIMEOUT)
        resp.raise_for_status()
        payload = resp.json()
        response = payload.get("response") or {}
        results = response.get("results") or []

        if total_pages is None:
            total = int(response.get("total_num_results") or 0)
            per_page = int(response.get("num_results_per_page") or CONSTRUCTOR_RESULTS_PER_PAGE)
            total_pages = max((total + per_page - 1) // per_page, 1)
            logger.info("Constructor %s: total=%s pages=%s", handle, total, total_pages)

        if not results:
            break
        all_results.extend(results)

        if total_pages is not None and page >= total_pages:
            break
        page += 1

    return all_results


def merge_constructor_rows(results: List[Dict[str, Any]], collection_url: str) -> List[Dict[str, str]]:
    handle = collection_handle_from_url(collection_url)
    rows: List[Dict[str, str]] = []

    for result in results:
        data = result.get("data") if isinstance(result.get("data"), dict) else {}
        variations = result.get("variations") or []
        if not variations:
            variations = [{"value": None, "data": {}}]

        for idx, variation in enumerate(variations):
            variation = variation if isinstance(variation, dict) else {}
            v_data = variation.get("data") if isinstance(variation.get("data"), dict) else {}
            row: Dict[str, str] = {
                "collection.url": collection_url,
                "collection.handle": handle,
                "product.id": list_to_text(data.get("id")),
                "product.handle": list_to_text(data.get("handle")),
                "product.title": list_to_text(result.get("value")),
                "product.title.v2": list_to_text(data.get("subtitle")),
                "product.title.v3": list_to_text(data.get("product")),
                "product.title.v4": list_to_text(variation.get("value")),
                "product.productType": list_to_text(data.get("product_type")),
                "product.sort": list_to_text(data.get("featured")),
                "product.labels": list_to_text(result.get("labels")),
                "product.matchedTerms": list_to_text(result.get("matched_terms")),
                "product.tags_all": list_to_text(data.get("keywords")),
                "product.description": list_to_text(data.get("description")),
                "product.color": list_to_text(data.get("color_code")),
                "product.rise": list_to_text(data.get("rise")),
                "product.closure": list_to_text(data.get("closure")),
                "product.onlineStoreUrl": list_to_text(data.get("url")),
                "variant.price": maybe_number(v_data.get("prices")),
                "variant.compare_at_price": maybe_number(v_data.get("compareAtPrices")),
                "product.notifyBIS": list_to_text(data.get("show_klaviyo_bis")),
                "variant.quantityAvailable.Instore": sum_inventory_by_location(
                    v_data.get("inventory"), online=False
                ),
                "variant.quantityAvailable.Online": sum_inventory_by_location(
                    v_data.get("inventory"), online=True
                ),
                "variant.id": list_to_text(v_data.get("variation_id") or data.get("variation_id")),
                "variant.sku": list_to_text(v_data.get("sku") or data.get("sku")),
                "variant.size": size_from_sku(v_data.get("sku") or data.get("sku")),
                "product.material": list_to_text(data.get("material")),
                "product.fabric": list_to_text(data.get("fabric")),
                "product.mill": list_to_text(data.get("mill")),
                "product.country": list_to_text(data.get("coo")),
                "collection.handles": list_to_text(data.get("group_ids")),
                "product.highlight": list_to_text(data.get("highlight")),
                "product.capsule": list_to_text(data.get("capsule")),
                "product.image": list_to_text(data.get("image_url")),
                "product.images.v2": "" if data.get("image_url") else list_to_text(data.get("mediaImages")),
                "varient.raw": list_to_text(result.get("variations")),
            }

            # carry unmapped fields as extras with [level].[label] naming
            mapped_levels = {
                "collection": {"url", "handle", "handles"},
                "product": {
                    "id",
                    "handle",
                    "title",
                    "title.v2",
                    "title.v3",
                    "title.v4",
                    "productType",
                    "sort",
                    "labels",
                    "matchedTerms",
                    "tags_all",
                    "description",
                    "color",
                    "rise",
                    "closure",
                    "onlineStoreUrl",
                    "notifyBIS",
                    "material",
                    "fabric",
                    "mill",
                    "country",
                    "highlight",
                    "capsule",
                    "image",
                    "images.v2",
                },
                "variant": {
                    "price",
                    "compare_at_price",
                    "quantityAvailable.Instore",
                    "quantityAvailable.Online",
                    "size",
                    "id",
                    "sku",
                },
            }

            for key, value in data.items():
                source_header = f"constructor.data.{key}"
                if source_header in CONSTRUCTOR_REMOVE_HEADERS:
                    continue
                label = key
                new_col = f"product.{label}"
                if label not in mapped_levels["product"] and new_col not in row:
                    row[new_col] = list_to_text(value)

            for key, value in v_data.items():
                label = key
                new_col = f"variant.{label}"
                if label not in mapped_levels["variant"] and new_col not in row:
                    row[new_col] = list_to_text(value)

            rows.append(row)

    # Post-pass column rules
    def all_same(col: str, baseline_col: str) -> bool:
        non_blank = [r.get(col, "") for r in rows if r.get(col, "") != ""]
        baseline = [r.get(baseline_col, "") for r in rows if r.get(col, "") != ""]
        if not non_blank:
            return True
        return all(a == b for a, b in zip(non_blank, baseline))

    drop_cols: Set[str] = set()
    if all_same("product.title.v3", "product.title"):
        drop_cols.add("product.title.v3")
    if all_same("product.title.v4", "product.title"):
        drop_cols.add("product.title.v4")

    for c in ("product.labels", "product.matchedTerms"):
        if all((r.get(c, "") == "" or r.get(c, "") == "{}" or r.get(c, "") == "[]") for r in rows):
            drop_cols.add(c)

    if drop_cols:
        for r in rows:
            for c in drop_cols:
                r.pop(c, None)

    return rows


def fetch_constructor_rows(session: requests.Session, logger: logging.Logger) -> List[Dict[str, str]]:
    all_rows: List[Dict[str, str]] = []
    for collection_url in COLLECTION_URL:
        try:
            results = fetch_constructor_results_for_collection(session, collection_url, logger)
            all_rows.extend(merge_constructor_rows(results, collection_url))
        except requests.RequestException as exc:
            logger.warning("Constructor fetch failed for %s -> %s", collection_url, exc)
    logger.info("Constructor rows collected: %s", len(all_rows))
    return all_rows


def _to_int(value: Any) -> Optional[int]:
    try:
        return int(float(str(value)))
    except (TypeError, ValueError):
        return None


def backfill_oos_constructor_rows(
    constructor_rows: List[Dict[str, str]],
    graphql_rows: List[Dict[str, str]],
    logger: logging.Logger,
) -> List[Dict[str, str]]:
    """Constructor only indexes variants that are in stock at some location, so
    sold-out variants never come back from its API. Re-add them from the
    GraphQL rows already fetched this run: any variant that is out of stock
    (quantityAvailable <= 0) or not available for sale and is missing from the
    Constructor results. A variant absent from Constructor has no stock at any
    location, so both inventory columns are 0.

    Product-level Constructor enrichment (color, rise, fabric, ...) is reused by
    handle when the product itself is still in Constructor; otherwise only the
    fields GraphQL provides are populated. Every row is tagged via `source`.
    """
    existing_skus = {r.get("variant.sku", "") for r in constructor_rows if r.get("variant.sku")}

    # Product-level Constructor template per handle (drop variant-level fields).
    templates: Dict[str, Dict[str, str]] = {}
    for r in constructor_rows:
        handle = r.get("product.handle", "")
        if handle and handle not in templates:
            templates[handle] = {k: v for k, v in r.items() if not k.startswith("variant.")}

    # Tag native rows and give them an availability flag.
    for r in constructor_rows:
        r.setdefault("source", "constructor")
        if "variant.available" not in r:
            instore = _to_int(r.get("variant.quantityAvailable.Instore")) or 0
            online = _to_int(r.get("variant.quantityAvailable.Online")) or 0
            r["variant.available"] = "True" if (instore + online) > 0 else "False"

    added: List[Dict[str, str]] = []
    seen_backfill: Set[str] = set()
    for g in graphql_rows:
        sku = g.get("variant.sku", "")
        if not sku or sku in existing_skus or sku in seen_backfill:
            continue
        qty = _to_int(g.get("variant.quantityAvailable"))
        available_flag = str(g.get("variant.available", "")).strip().lower()
        is_oos = (qty is not None and qty <= 0) or available_flag in {"false", "no", "0"}
        if not is_oos:
            continue

        handle = g.get("product.handle", "")
        row = dict(templates.get(handle, {}))
        row.setdefault("product.handle", handle)
        row.setdefault("product.id", g.get("product.id", ""))
        row.setdefault("product.title", g.get("product.title", ""))
        row.setdefault("product.onlineStoreUrl", g.get("product.onlineStoreUrl", ""))
        row["variant.id"] = g.get("variant.id", "")
        row["variant.sku"] = sku
        row["variant.size"] = size_from_sku(sku) or g.get("variant.option1", "")
        row["variant.price"] = g.get("variant.price", "")
        row["variant.compare_at_price"] = g.get("variant.compare_at_price", "")
        row["variant.quantityAvailable.Instore"] = "0"
        row["variant.quantityAvailable.Online"] = "0"
        row["variant.available"] = "False"
        row["source"] = "graphql_oos"
        added.append(row)
        seen_backfill.add(sku)

    constructor_rows.extend(added)
    logger.info("Constructor OOS backfill from GraphQL: +%s rows", len(added))
    return constructor_rows


def _constructor_facet_options(
    session: requests.Session, handle: str
) -> Dict[str, List[str]]:
    """Return {facet_name: [option values]} including hidden facets."""
    params = [
        ("c", "cio-ui-plp-1.6.2"),
        ("key", CONSTRUCTOR_API_KEY),
        ("i", CONSTRUCTOR_CLIENT_ID),
        ("s", CONSTRUCTOR_SESSION),
        ("num_results_per_page", "1"),
        ("fmt_options[show_hidden_facets]", "true"),
    ]
    resp = session.get(
        f"{CONSTRUCTOR_BROWSE_ENDPOINT}/{handle}", params=params, timeout=REQUEST_TIMEOUT
    )
    resp.raise_for_status()
    facets = (resp.json().get("response") or {}).get("facets") or []
    out: Dict[str, List[str]] = {}
    for facet in facets:
        name = facet.get("name")
        values = [o.get("value") for o in (facet.get("options") or []) if o.get("value") is not None]
        if name:
            out[name] = values
    return out


def fetch_constructor_facet_map(
    session: requests.Session,
    collection_url: str,
    facet_names: Sequence[str],
    logger: logging.Logger,
) -> Dict[str, Dict[str, str]]:
    """Resolve normally-unexposed facet values per product by filtering the
    browse endpoint to each facet value and recording the matching product ids.

    Returns {product_id: {facet_name: "value1, value2"}}.
    """
    handle = collection_handle_from_url(collection_url)
    if not handle:
        return {}
    try:
        facet_options = _constructor_facet_options(session, handle)
    except requests.RequestException as exc:
        logger.warning("Facet options fetch failed for %s -> %s", handle, exc)
        return {}

    result: Dict[str, Dict[str, Set[str]]] = {}
    base = [
        ("c", "cio-ui-plp-1.6.2"),
        ("key", CONSTRUCTOR_API_KEY),
        ("i", CONSTRUCTOR_CLIENT_ID),
        ("s", CONSTRUCTOR_SESSION),
        ("num_results_per_page", "100"),
        ("fmt_options[show_hidden_facets]", "true"),
    ]
    for facet in facet_names:
        values = facet_options.get(facet) or []
        for value in values:
            page = 1
            while True:
                params = base + [("page", str(page)), (f"filters[{facet}]", value)]
                try:
                    resp = session.get(
                        f"{CONSTRUCTOR_BROWSE_ENDPOINT}/{handle}",
                        params=params,
                        timeout=REQUEST_TIMEOUT,
                    )
                    resp.raise_for_status()
                except requests.RequestException as exc:
                    logger.warning("Facet tag %s=%s failed -> %s", facet, value, exc)
                    break
                response = resp.json().get("response") or {}
                results = response.get("results") or []
                if not results:
                    break
                for item in results:
                    pid = str((item.get("data") or {}).get("id") or "")
                    if pid:
                        result.setdefault(pid, {}).setdefault(facet, set()).add(str(value))
                total = int(response.get("total_num_results") or 0)
                if page * 100 >= total:
                    break
                page += 1
        logger.info("Facet '%s': %s values tagged", facet, len(values))

    # Flatten sets to comma-joined strings.
    flat: Dict[str, Dict[str, str]] = {}
    for pid, facets in result.items():
        flat[pid] = {name: ", ".join(sorted(vals)) for name, vals in facets.items()}
    return flat


def enrich_constructor_rows(
    constructor_rows: List[Dict[str, str]],
    graphql_rows: List[Dict[str, str]],
    facet_map: Dict[str, Dict[str, str]],
) -> None:
    """Add the requested facet-derived columns to each Constructor row:
    color_name (from the GraphQL Color option, by handle), bottoms_size (the
    variant's own size), and the filter-tagged facets (by product id)."""
    # handle -> color name, from the GraphQL "Color" option (variant.option2).
    handle_color: Dict[str, str] = {}
    for g in graphql_rows:
        handle = g.get("product.handle", "")
        color = g.get("variant.option2", "")
        if handle and color and handle not in handle_color:
            handle_color[handle] = color

    for row in constructor_rows:
        handle = row.get("product.handle", "")
        pid = str(row.get("product.id", ""))
        row["product.color_name"] = handle_color.get(handle, "")
        row["variant.bottoms_size"] = row.get("variant.size", "")
        for facet, value in facet_map.get(pid, {}).items():
            row[f"product.{facet}"] = value
        # Ensure every requested facet column exists even when unmatched.
        for facet in CONSTRUCTOR_FACET_TAG_FIELDS:
            row.setdefault(f"product.{facet}", "")


# ---------------------------------------------------------------------------
# GraphQL (with introspection pass first)
# ---------------------------------------------------------------------------
INTROSPECTION_TYPE_FIELDS_QUERY = """
query($t: String!) {
  __type(name: $t) {
    name
    fields {
      name
      type {
        kind
        name
        ofType {
          kind
          name
          ofType {
            kind
            name
          }
        }
      }
      args {
        name
        type {
          kind
          name
          ofType { kind name }
        }
      }
    }
  }
}
"""

COLLECTION_PRODUCTS_QUERY = """
query CollectionProducts($handle: String!, $cursor: String, $pageSize: Int!) {
  collection(handle: $handle) {
    id
    handle
    title
    products(first: $pageSize, after: $cursor) {
      pageInfo {
        hasNextPage
        endCursor
      }
      edges {
        node {
          id
          handle
          title
          publishedAt
          createdAt
          productType
          tags
          vendor
          onlineStoreUrl
          description
          descriptionHtml
          totalInventory
          category {
            name
          }
          featuredImage {
            url
            altText
          }
          priceRange {
            minVariantPrice { amount currencyCode }
            maxVariantPrice { amount currencyCode }
          }
          options {
            name
            values
          }
          variants(first: 100) {
            pageInfo {
              hasNextPage
              endCursor
            }
            edges {
              cursor
              node {
                id
                title
                sku
                barcode
                availableForSale
                quantityAvailable
                price {
                  amount
                  currencyCode
                }
                compareAtPrice {
                  amount
                  currencyCode
                }
                selectedOptions {
                  name
                  value
                }
              }
            }
          }
        }
      }
    }
  }
}
"""

FILTER_PROBE_QUERIES = [
    """
query FilterProbeA($handle: String!) {
  collection(handle: $handle) {
    products(first: 1) {
      filters {
        id
        label
        type
        values {
          id
          label
          count
        }
      }
    }
  }
}
""",
    """
query FilterProbeB($handle: String!) {
  collection(handle: $handle) {
    products(first: 1) {
      productFilters {
        id
        label
        type
        values {
          id
          label
          count
        }
      }
    }
  }
}
""",
]


def perform_graphql_request(
    session: requests.Session,
    endpoint: str,
    query: str,
    variables: Dict[str, Any],
) -> Dict[str, Any]:
    headers = {
        "Content-Type": "application/json",
        "X-Shopify-Storefront-Access-Token": X_SHOPIFY_STOREFRONT_ACCESS_TOKEN,
    }
    resp = session.post(
        endpoint,
        json={"query": query, "variables": variables},
        headers=headers,
        timeout=REQUEST_TIMEOUT,
    )
    resp.raise_for_status()
    payload = resp.json()
    if payload.get("errors"):
        raise requests.RequestException(f"GraphQL errors: {payload['errors']}")
    return payload


def graphql_introspection_probe(session: requests.Session, endpoint: str, logger: logging.Logger) -> None:
    for type_name in ("Product", "ProductVariant"):
        try:
            data = perform_graphql_request(
                session,
                endpoint,
                INTROSPECTION_TYPE_FIELDS_QUERY,
                {"t": type_name},
            )
        except requests.RequestException as exc:
            logger.warning("Introspection failed for %s -> %s", type_name, exc)
            continue
        fields = (((data.get("data") or {}).get("__type") or {}).get("fields") or [])
        logger.info("%s field count: %s", type_name, len(fields))
        for f in fields:
            type_obj = f.get("type") or {}
            type_name_guess = type_obj.get("name") or ((type_obj.get("ofType") or {}).get("name")) or type_obj.get("kind")
            logger.info("%s.%s -> %s", type_name, f.get("name"), type_name_guess)


def discover_collection_filters(
    session: requests.Session,
    endpoint: str,
    handle: str,
    logger: logging.Logger,
) -> Dict[str, List[str]]:
    for query in FILTER_PROBE_QUERIES:
        try:
            data = perform_graphql_request(session, endpoint, query, {"handle": handle})
        except requests.RequestException:
            continue
        collection = ((data.get("data") or {}).get("collection") or {})
        products = collection.get("products") or {}
        filters_block = products.get("filters") or products.get("productFilters")
        if not filters_block:
            continue
        parsed: Dict[str, List[str]] = {}
        for fil in filters_block:
            if not isinstance(fil, dict):
                continue
            label = str(fil.get("label") or fil.get("id") or "").strip()
            if not label:
                continue
            values = []
            for item in fil.get("values") or []:
                if isinstance(item, dict):
                    candidate = item.get("label") or item.get("id")
                    if candidate:
                        values.append(str(candidate))
            if values:
                parsed[label] = values
        if parsed:
            logger.info("GraphQL filter groups for %s: %s", handle, len(parsed))
            return parsed
    return {}


def product_matches_filters(product: Dict[str, Any]) -> bool:
    if PRODUCT_TYPE_FILTER:
        ptype = str(product.get("productType") or "").strip().lower()
        allowed = {x.strip().lower() for x in PRODUCT_TYPE_FILTER if x.strip()}
        if ptype not in allowed:
            return False

    if CATEGORY_FILTER:
        tags = [str(t) for t in (product.get("tags") or [])]
        corpus = " ".join(
            [
                str(product.get("title") or ""),
                str(product.get("productType") or ""),
                " ".join(tags),
            ]
        ).lower()
        if not any(cat.strip().lower() in corpus for cat in CATEGORY_FILTER if cat.strip()):
            return False

    return True


def normalize_filter_name(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(name).lower()).strip("_") or "unnamed"


def flatten_graphql_row(
    collection_url: str,
    collection_handle: str,
    collection_title: str,
    filters: Dict[str, List[str]],
    product: Dict[str, Any],
    variant: Optional[Dict[str, Any]],
) -> Dict[str, str]:
    row: Dict[str, str] = {
        "collection.url": collection_url,
        "collection.handle": collection_handle,
        "collection.title": collection_title,
        "product.id": list_to_text(product.get("id")),
        "product.handle": list_to_text(product.get("handle")),
        "product.published_at": list_to_text(product.get("publishedAt")),
        "product.created_at": list_to_text(product.get("createdAt")),
        "product.title": list_to_text(product.get("title")),
        "product.productType": list_to_text(product.get("productType")),
        "product.category.name": list_to_text((product.get("category") or {}).get("name")),
        "product.tags_all": list_to_text(product.get("tags")),
        "product.vendor": list_to_text(product.get("vendor")),
        "product.description": list_to_text(product.get("description")),
        "product.descriptionHtml": list_to_text(product.get("descriptionHtml")),
        "product.priceRange": list_to_text(product.get("priceRange")),
        "product.totalInventory": list_to_text(product.get("totalInventory")),
        "product.featuredImage": list_to_text(product.get("featuredImage")),
        "product.onlineStoreUrl": list_to_text(product.get("onlineStoreUrl")),
    }

    corpus = " ".join(
        [
            row.get("product.title", ""),
            row.get("product.productType", ""),
            row.get("product.tags_all", ""),
        ]
    ).lower()
    for name, vals in filters.items():
        matches = [v for v in vals if str(v).lower() in corpus]
        if matches:
            row[f"filter.{normalize_filter_name(name)}"] = ", ".join(sorted(set(matches)))

    if variant:
        row["variant.id"] = list_to_text(variant.get("id"))
        row["variant.title"] = list_to_text(variant.get("title"))
        row["variant.sku"] = list_to_text(variant.get("sku"))
        row["variant.barcode"] = list_to_text(variant.get("barcode"))
        row["variant.available"] = list_to_text(variant.get("availableForSale"))
        row["variant.quantityAvailable"] = list_to_text(variant.get("quantityAvailable"))
        row["variant.price"] = maybe_number(variant.get("price"))
        row["variant.compare_at_price"] = maybe_number(variant.get("compareAtPrice"))
        selected = variant.get("selectedOptions") or []
        for opt in selected:
            if not isinstance(opt, dict):
                continue
            oname = str(opt.get("name") or "").strip().lower()
            oval = list_to_text(opt.get("value"))
            if oname in {"size", "option1"}:
                row["variant.option1"] = oval
            elif oname in {"color", "option2"}:
                row["variant.option2"] = oval
            else:
                if "variant.option3" not in row:
                    row["variant.option3"] = oval

    return row


def build_column_order(rows: List[Dict[str, str]]) -> List[str]:
    all_cols = {k for r in rows for k in r.keys()}
    base = [c for c in COLUMN_ORDER_BASE if c in all_cols]
    extras = []
    for col in sorted(all_cols):
        if col in base:
            continue
        if col in EXTRA_FORBIDDEN_COLUMNS:
            continue
        extras.append(col)
    return base + extras


def build_constructor_column_order(rows: List[Dict[str, str]]) -> List[str]:
    all_cols = {k for r in rows for k in r.keys()}
    base = [c for c in CONSTRUCTOR_COLUMN_ORDER if c in all_cols]
    extras = [c for c in sorted(all_cols) if c not in base]
    return base + extras


def fetch_graphql_rows(session: requests.Session, logger: logging.Logger) -> List[Dict[str, str]]:
    endpoint = GRAPHQL
    graphql_introspection_probe(session, endpoint, logger)

    rows: List[Dict[str, str]] = []

    for collection_url in COLLECTION_URL:
        handle = collection_handle_from_url(collection_url)
        if not handle:
            continue
        filters = discover_collection_filters(session, endpoint, handle, logger)
        cursor: Optional[str] = None

        while True:
            payload = perform_graphql_request(
                session,
                endpoint,
                COLLECTION_PRODUCTS_QUERY,
                {"handle": handle, "cursor": cursor, "pageSize": 100},
            )
            collection = ((payload.get("data") or {}).get("collection") or {})
            products_conn = collection.get("products") or {}
            edges = products_conn.get("edges") or []

            for edge in edges:
                product = (edge or {}).get("node") or {}
                if not product_matches_filters(product):
                    continue

                variants = ((product.get("variants") or {}).get("edges") or [])
                if not variants:
                    rows.append(
                        flatten_graphql_row(
                            collection_url,
                            handle,
                            list_to_text(collection.get("title")),
                            filters,
                            product,
                            None,
                        )
                    )
                else:
                    for v_edge in variants:
                        variant = (v_edge or {}).get("node") or {}
                        rows.append(
                            flatten_graphql_row(
                                collection_url,
                                handle,
                                list_to_text(collection.get("title")),
                                filters,
                                product,
                                variant,
                            )
                        )

            page_info = products_conn.get("pageInfo") or {}
            if page_info.get("hasNextPage"):
                cursor = page_info.get("endCursor")
            else:
                break

    logger.info("GraphQL rows collected: %s", len(rows))
    return rows


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------
def write_sheet(ws, rows: List[Dict[str, str]], ordered_columns: Optional[List[str]] = None) -> None:
    if not rows:
        ws.append(["note"])
        ws.append(["No rows collected"])
        return

    columns = ordered_columns or sorted({k for row in rows for k in row.keys()})
    ws.append(columns)
    for row in rows:
        ws.append([row.get(col, "") for col in columns])


def write_workbook(
    constructor_rows: List[Dict[str, str]],
    graphql_rows: List[Dict[str, str]],
    logger: logging.Logger,
) -> Path:
    ts = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    output = OUTPUT_DIR / f"{BRAND_SLUG}_Constructor_Graphql_{ts}.xlsx"

    wb = Workbook()
    wb.remove(wb.active)

    ws_constructor = wb.create_sheet("Constructor")
    write_sheet(
        ws_constructor,
        constructor_rows,
        ordered_columns=build_constructor_column_order(constructor_rows),
    )

    ws_graphql = wb.create_sheet("GraphQL")
    write_sheet(ws_graphql, graphql_rows, ordered_columns=build_column_order(graphql_rows))

    wb.save(output)
    logger.info("Workbook written: %s", output.resolve())
    return output


def main() -> None:
    logger = configure_logger()
    session = build_session()

    constructor_rows = fetch_constructor_rows(session, logger)
    graphql_rows = fetch_graphql_rows(session, logger)

    # Constructor omits sold-out variants; re-add them from GraphQL.
    constructor_rows = backfill_oos_constructor_rows(constructor_rows, graphql_rows, logger)

    # Resolve normally-unexposed facet fields (color_name, gender, category, ...)
    # and add them as columns.
    facet_map: Dict[str, Dict[str, str]] = {}
    for collection_url in COLLECTION_URL:
        facet_map.update(
            fetch_constructor_facet_map(session, collection_url, CONSTRUCTOR_FACET_TAG_FIELDS, logger)
        )
    enrich_constructor_rows(constructor_rows, graphql_rows, facet_map)

    output = write_workbook(constructor_rows, graphql_rows, logger)

    if not constructor_rows:
        logger.warning("Constructor tab is blank.")
    if not graphql_rows:
        logger.warning("GraphQL tab is blank.")
    logger.info("Done. Output: %s", output)


if __name__ == "__main__":
    main()