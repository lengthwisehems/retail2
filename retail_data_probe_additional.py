"""Exploratory app/product probe with multi-sheet Excel output."""
from __future__ import annotations

import json
import logging
import os
import re
import time
import uuid
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple
from urllib.parse import parse_qsl, urlparse, urlsplit

import requests
import urllib3
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from requests.adapters import HTTPAdapter, Retry

import retail_data_probe as base_probe

urllib3.util.connection.HAS_IPV6 = False

# ---------------------------------------------------------------------------
# Brand/site configuration (minimal hard-coding)
# ---------------------------------------------------------------------------
BRAND = "EB Denim"
BASE_URL = "https://www.ebdenim.com"
COLLECTION_PATH = "/collections/pants"
MYSHOPIFY = "https://ebdenim-com.myshopify.com"
GRAPHQL = ""
X_SHOPIFY_STOREFRONT_ACCESS_TOKEN = ""

# App discovery URLs (kept together for easy tweaking)
GLOBO_SCRIPT_URLS = [
    "https://www.ebdenim.com/cdn/shop/t/46/assets/v5.globo.filter.lib.js",
    "https://www.ebdenim.com/cdn/shop/t/46/assets/v5.globo.search.css",
]
GLOBO_ENDPOINT = "https://filter-x3.globo.io/api/apiFilter"
GLOBO_FALLBACK_PARAMS = {}
REBUY_DISCOVERY_URLS: List[str] = []
REBUNK_WIDGET_BASE = "https://rebuyengine.com/api/v1"
RESTOCK_SDK = "https://cdn.hengam.io/restock-alerts-sdk.js"
AVADA_SCRIPT_URLS = [
    "https://seo.apps.avada.io/scripttag/avada-seo-installed.js?shop=ebdenim-com.myshopify.com",
]
BUNDLER_STATUS = "https://bundler.nice-team.net/app/shop/status/ebdenim-com.myshopify.com.js"
POSTSCRIPT_ENDPOINTS = [
    "https://sdk-api-proxy.postscript.io/sdk/config?shop_id=575770",
    "https://sdk-api-proxy.postscript.io/v2/public/popups/575770/desktop",
    "https://sdk-api-proxy.postscript.io/v2/public/block_popups/575770",
    "https://sdk-api-proxy.postscript.io/public/klaviyo_form_status/575770",
]

# ---------------------------------------------------------------------------
# Derived paths and constants
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = BASE_DIR / "Output"
OUTPUT_DIR.mkdir(exist_ok=True)

BRAND_SLUG = BRAND.lower().replace(" ", "_") or "brand"
LOG_PATH = BASE_DIR / f"{BRAND_SLUG}_probe_run.log"
FALLBACK_LOG_PATH = OUTPUT_DIR / f"{BRAND_SLUG}_probe_run.log"
REBUY_DOMAIN = urlsplit(BASE_URL).netloc.replace("www.", "")

REQUEST_TIMEOUT = 30
TRANSIENT_STATUS = {429, 500, 502, 503, 504}
TOKEN_REGEX = re.compile(r"\b[0-9a-f]{32}\b", re.IGNORECASE)
DEFAULT_GRAPHQL_VERSIONS = [
    "api/2025-10/graphql.json",
    "api/2025-07/graphql.json",
    "api/unstable/graphql.json",
]
MAX_SCRIPT_FETCHES = 25

COLUMN_ORDER_BASE: Tuple[str, ...] = (
    "product.id",
    "product.handle",
    "product.published_at",
    "product.created_at",
    "product.title",
    "product.productType",
    "product.tags_all",
    "product.vendor",
    "product.description",
    "product.descriptionHtml",
    "variant.title",
    "variant.option1",
    "variant.option2",
    "variant.option3",
    "variant.price",
    "variant.compare_at_price",
    "variant.available",
    "variant.quantityAvailable",
    "product.totalInventory",
    "variant.id",
    "variant.sku",
    "variant.barcode",
    "product.images[0].src",
    "product.onlineStoreUrl",
)

DEFAULT_FORBIDDEN_FIELDS = {
    "ProductVariant": {
        "components",
        "groupedBy",
        "quantityPriceBreaks",
        "sellingPlanAllocations",
        "sellingPlanGroups",
        "storeAvailability",
    }
}

DROP_FEATURED_IMAGE_FIELDS = {
    "id",
    "product_id",
    "position",
    "created_at",
    "updated_at",
    "alt",
    "width",
    "height",
}


# Restock parsing helpers
VARIANT_BLOCK_RE = re.compile(r"variants\s*:\s*\[(?P<body>.*?)\]\s*,?\s*\}", re.S)
VARIANT_OBJ_RE = re.compile(r"\{[^{}]*?\bid\s*:\s*(?P<id>\d+)[^{}]*?\bquantity\s*:\s*(?P<qty>-?\d+)[^{}]*?\}", re.S)

AVADA_INVQTY_RE = re.compile(r"AVADA_INVQTY\s*\{(?P<body>.*?)\}", re.S)
AVADA_PAIR_RE = re.compile(r"(\d+)\s*[:=]\s*(\d+)")


def sync_base_probe_constants() -> None:
    base_probe.BRAND = BRAND
    base_probe.COLLECTION_URL = f"{BASE_URL}{COLLECTION_PATH}"
    base_probe.MYSHOPIFY = MYSHOPIFY
    base_probe.GRAPHQL = GRAPHQL
    base_probe.X_SHOPIFY_STOREFRONT_ACCESS_TOKEN = X_SHOPIFY_STOREFRONT_ACCESS_TOKEN
    base_probe.GRAPHQL_FILTER_TAG = ""
    base_probe.STOREFRONT_COLLECTION_HANDLES = ["pants"]
    base_probe.DEFAULT_GRAPHQL_VERSIONS = list(DEFAULT_GRAPHQL_VERSIONS)
    base_probe.REQUEST_TIMEOUT = REQUEST_TIMEOUT
    base_probe.TOKEN_REGEX = TOKEN_REGEX

# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------
@dataclass
class FetchResult:
    url: str
    status_code: int
    text: str


@dataclass
class ProductHandle:
    product_id: int
    handle: str
    title: str
    url: str

# ---------------------------------------------------------------------------
# Logging / session helpers
# ---------------------------------------------------------------------------

def configure_logger() -> logging.Logger:
    logger = logging.getLogger(BRAND_SLUG)
    if logger.handlers:
        return logger

    logger.setLevel(logging.INFO)
    formatter = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")

    handler = None
    try:
        file_handler = logging.FileHandler(LOG_PATH, mode="a", encoding="utf-8")
        file_handler.setFormatter(formatter)
        logger.addHandler(file_handler)
        handler = file_handler
    except OSError:
        try:
            file_handler = logging.FileHandler(FALLBACK_LOG_PATH, mode="a", encoding="utf-8")
            file_handler.setFormatter(formatter)
            logger.addHandler(file_handler)
            handler = file_handler
        except OSError:
            handler = None

    stream_handler = logging.StreamHandler()
    stream_handler.setFormatter(formatter)
    logger.addHandler(stream_handler)

    if handler:
        logger.info("Logging to %s", getattr(handler, "baseFilename", handler))
    else:
        logger.warning("File logging unavailable; console only")
    return logger


def build_session(logger: logging.Logger) -> requests.Session:
    session = requests.Session()
    session.verify = False
    session.proxies.update(requests.utils.get_environ_proxies(BASE_URL))
    retries = Retry(
        total=5,
        backoff_factor=1.5,
        status_forcelist=TRANSIENT_STATUS,
        allowed_methods=["GET", "POST"],
    )
    adapter = HTTPAdapter(max_retries=retries)
    session.mount("http://", adapter)
    session.mount("https://", adapter)
    session.headers.update(
        {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
            "(KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
        }
    )
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    return session


def request(session: requests.Session, url: str, logger: logging.Logger, method: str = "GET", **kwargs: Any) -> FetchResult:
    resp = session.request(method, url, timeout=REQUEST_TIMEOUT, **kwargs)
    return FetchResult(url=url, status_code=resp.status_code, text=resp.text)


def request_json(
    session: requests.Session,
    url: str,
    logger: logging.Logger,
    *,
    params: Optional[Dict[str, Any]] = None,
    method: str = "GET",
    **kwargs: Any,
) -> Dict[str, Any]:
    resp = session.request(method, url, timeout=REQUEST_TIMEOUT, params=params, **kwargs)
    if resp.status_code != 200:
        logger.warning("Request failed %s -> %s", resp.url, resp.status_code)
        return {}
    try:
        return resp.json()
    except Exception as exc:  # noqa: BLE001
        logger.warning("JSON parse failed for %s: %s", resp.url, exc)
        return {}


# ---------------------------------------------------------------------------
# Utility helpers
# ---------------------------------------------------------------------------

def normalize_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, Path):
        return value.as_posix()
    return json.dumps(value, ensure_ascii=False, sort_keys=True)


def flatten_value(value: Any, prefix: str, *, skip_keys: Optional[Iterable[str]] = None) -> Dict[str, Any]:
    items: Dict[str, Any] = {}
    skip = set(skip_keys or [])
    if prefix and prefix in skip:
        return items
    if isinstance(value, dict):
        for key, inner in value.items():
            new_prefix = f"{prefix}.{key}" if prefix else key
            if new_prefix in skip:
                continue
            items.update(flatten_value(inner, new_prefix, skip_keys=skip))
    elif isinstance(value, list):
        for index, inner in enumerate(value):
            new_prefix = f"{prefix}[{index}]" if prefix else f"[{index}]"
            if new_prefix in skip:
                continue
            items.update(flatten_value(inner, new_prefix, skip_keys=skip))
    else:
        items[prefix] = value
    return items


def flatten_record(record: Dict[str, Any], *, skip_keys: Optional[Iterable[str]] = None) -> Dict[str, Any]:
    flat: Dict[str, Any] = {}
    for key, value in record.items():
        flat.update(flatten_value(value, key, skip_keys=skip_keys))
    return flat


def normalize_app_row_keys(row: Dict[str, Any]) -> None:
    if "handle" in row and "product.handle" not in row:
        row["product.handle"] = row.pop("handle")
    if "page_title" in row and "product.page_title" not in row:
        row["product.page_title"] = row.pop("page_title")
    if "quantity_available" in row and "variant.quantityAvailable" not in row:
        row["variant.quantityAvailable"] = row.pop("quantity_available")


def apply_name_value_columns(row: Dict[str, Any]) -> None:
    replacements: Dict[str, Any] = {}
    to_remove: List[str] = []
    for key, value in list(row.items()):
        if not key.endswith(".name"):
            continue
        prefix = key[:-5]
        name_value = str(value).strip()
        value_key = f"{prefix}.value"
        if not name_value or value_key not in row:
            continue
        cleaned = re.sub(r"[^0-9A-Za-z]+", "_", name_value).strip("_") or "value"
        new_key = f"{prefix}.{cleaned}"
        replacements[new_key] = row[value_key]
        to_remove.extend([key, value_key])
    for key in to_remove:
        row.pop(key, None)
    row.update(replacements)


def finalize_rows(rows: List[Dict[str, Any]], *, prune: bool = False) -> List[Dict[str, Any]]:
    cleaned: List[Dict[str, Any]] = []
    for row in rows:
        apply_name_value_columns(row)
        normalize_app_row_keys(row)
        cleaned.append(prune_probe_fields(row) if prune else row)
    return cleaned


def prune_probe_fields(row: Dict[str, Any]) -> Dict[str, Any]:
    cleaned = dict(row)
    drop_keys = set()
    for key in list(cleaned.keys()):
        if key.startswith("product.variants["):
            drop_keys.add(key)
        if key.startswith("variant.components") or key.startswith("variant.groupedBy"):
            drop_keys.add(key)
        if key.startswith("variant.quantityPriceBreaks") or key.startswith("variant.sellingPlan"):
            drop_keys.add(key)
        if key.startswith("variant.storeAvailability"):
            drop_keys.add(key)
        if key.startswith("product.featured_image") or key.startswith("variant.featured_image"):
            for suffix in DROP_FEATURED_IMAGE_FIELDS:
                if key.endswith(f".{suffix}") or key.endswith(f"[{suffix}]"):
                    drop_keys.add(key)
                    break
        if re.search(r"featured_image\.variant_ids\[\d+\]", key):
            drop_keys.add(key)
    for key in drop_keys:
        cleaned.pop(key, None)
    return cleaned


def normalize_base_fields(row: Dict[str, Any], product: Dict[str, Any]) -> Dict[str, Any]:
    tags = product.get("tags") or []
    if isinstance(tags, list):
        row.setdefault("product.tags_all", ",".join(str(t) for t in tags))
    published = product.get("published_at") or product.get("publishedAt")
    created = product.get("created_at") or product.get("createdAt")
    if published is not None:
        row.setdefault("product.published_at", published)
    if created is not None:
        row.setdefault("product.created_at", created)
    online = product.get("onlineStoreUrl") or product.get("online_store_url")
    if online:
        row.setdefault("product.onlineStoreUrl", online)
    return row


def build_column_order(
    rows: List[Dict[str, Any]],
    *,
    extra_priority: Optional[Sequence[str]] = None,
    use_base: bool = True,
) -> List[str]:
    priority: List[str] = list(COLUMN_ORDER_BASE) if use_base else []
    if extra_priority:
        for col in extra_priority:
            if col not in priority:
                priority.append(col)
    seen = set(priority)
    for row in rows:
        for key in row.keys():
            if key not in seen:
                priority.append(key)
                seen.add(key)
    return priority


def write_sheet(
    wb: Workbook,
    title: str,
    rows: List[Dict[str, Any]],
    *,
    column_order: Optional[List[str]] = None,
    use_base: bool = True,
    extra_priority: Optional[Sequence[str]] = None,
) -> None:
    ws = wb.create_sheet(title)
    if not rows:
        ws.append(["info"])
        ws.append([f"no data for {title}"])
        return
    ordered_keys = column_order or build_column_order(
        rows, extra_priority=extra_priority, use_base=use_base
    )
    ws.append(ordered_keys)
    for row in rows:
        ws.append([normalize_cell(row.get(k, "")) for k in ordered_keys])

    for idx, col in enumerate(ordered_keys, start=1):
        length = max(len(str(col)), *(len(str(r.get(col, ""))) for r in rows))
        ws.column_dimensions[get_column_letter(idx)].width = min(80, length + 2)


# ---------------------------------------------------------------------------
# Discovery helpers
# ---------------------------------------------------------------------------

def fetch_collection_products(session: requests.Session, logger: logging.Logger) -> List[Dict[str, Any]]:
    products: List[Dict[str, Any]] = []
    page = 1
    while True:
        params = {"limit": 250, "page": page}
        url = f"{BASE_URL}{COLLECTION_PATH}/products.json"
        logger.info("Fetching collection page %s", page)
        data = request_json(session, url, logger, params=params, verify=False)
        chunk = data.get("products", []) if isinstance(data, dict) else []
        if not chunk:
            break
        products.extend(chunk)
        logger.info("Fetched %s products from page %s", len(chunk), page)
        page += 1
    return products


def fetch_product_htmls(
    session: requests.Session, logger: logging.Logger, handles: Sequence[str]
) -> Dict[str, str]:
    html_map: Dict[str, str] = {}
    for handle in handles:
        url = f"{BASE_URL}/products/{handle}"
        logger.info("Fetching PDP HTML for handle %s", handle)
        resp = request(session, url, logger)
        if resp.status_code != 200:
            logger.warning("Failed PDP fetch %s -> %s", url, resp.status_code)
            continue
        html_map[handle] = resp.text
    return html_map


# ---------------------------------------------------------------------------
# App parsers
# ---------------------------------------------------------------------------

def parse_restock(html: str) -> Dict[str, Dict[str, Any]]:
    if "_ReStockConfig" not in html:
        return {}
    m = VARIANT_BLOCK_RE.search(html)
    if not m:
        return {}
    body = m.group("body")
    out: Dict[str, Dict[str, Any]] = {}
    for mo in VARIANT_OBJ_RE.finditer(body):
        vid = mo.group("id")
        qty = int(mo.group("qty"))
        out[vid] = {
            "quantity": qty,
            "quantity_available": max(qty, 0),
            "notify_me": abs(qty) if qty < 0 else 0,
        }
    return out


def parse_avada(html: str) -> Tuple[Dict[str, Dict[str, Any]], Dict[str, Any]]:
    inv: Dict[str, Dict[str, Any]] = {}
    for vid, qty in re.findall(r"AVADA_INVQTY\s*\[\s*(\d+)\s*\]\s*=\s*(-?\d+)", html):
        inv[vid] = {"quantity": int(qty), "quantity_available": max(int(qty), 0)}

    cdt_match = re.search(r"AVADA_CDT\s*=\s*(\{.*?\})\s*;", html, re.S)
    cdt_data: Dict[str, Any] = {}
    if cdt_match:
        try:
            cdt_data = json.loads(cdt_match.group(1))
        except json.JSONDecodeError:
            try:
                cdt_data = json.loads(cdt_match.group(1).replace("'", '"'))
            except json.JSONDecodeError:
                cdt_data = {"raw": cdt_match.group(1)}
    return inv, cdt_data


def build_rebuy_product_handles(products: List[Dict[str, Any]]) -> List[ProductHandle]:
    handles: List[ProductHandle] = []
    for prod in products:
        if not isinstance(prod, dict):
            continue
        pid = prod.get("id")
        handle = prod.get("handle")
        if pid is None or handle is None:
            continue
        try:
            pid_int = int(pid)
        except (TypeError, ValueError):
            continue
        title = prod.get("title", "")
        url = f"https://{REBUY_DOMAIN}/products/{handle}"
        handles.append(ProductHandle(pid_int, handle, title, url))
    return handles


def fetch_rebuy_input_products(
    session: requests.Session,
    product: ProductHandle,
    widget_id: str,
    api_key: str,
    logger: logging.Logger,
) -> List[Dict[str, Any]]:
    params = {
        "key": api_key,
        "shopify_product_ids": str(product.product_id),
        "limit": 50,
        "product_groups": "yes",
        "uuid": str(uuid.uuid4()),
        "url": product.url,
    }
    base_url = f"{REBUNK_WIDGET_BASE}/custom/id/{widget_id}"
    data = request_json(session, base_url, logger, params=params, verify=False)
    metadata = data.get("metadata", {}) if isinstance(data, dict) else {}
    input_products = metadata.get("input_products", []) if isinstance(metadata, dict) else []
    if not input_products and isinstance(data, dict):
        custom = data.get("custom", {}) if isinstance(data.get("custom", {}), dict) else {}
        input_products = custom.get("data", []) if isinstance(custom, dict) else []
    if not input_products:
        logger.warning("No input_products returned for %s (%s)", product.handle, product.product_id)
    return input_products if isinstance(input_products, list) else []



def probe_rebuy(
    session: requests.Session,
    logger: logging.Logger,
    html_sources: List[str],
    products: List[Dict[str, Any]],
) -> Tuple[List[Dict[str, Any]], List[str]]:
    rows: List[Dict[str, Any]] = []
    token_blobs: List[str] = []
    widget_ids: set = set()
    api_keys: set = set()

    discovery_regex = re.compile(
        r"https://rebuyengine.com/api/v1/custom/id/(\d+)\?[^\s\"']*key=([0-9a-f]{40})",
        re.I,
    )
    for blob in html_sources:
        for wid, key in discovery_regex.findall(blob):
            widget_ids.add(wid)
            api_keys.add(key)
        for wid in re.findall(r"custom/id/(\d+)", blob):
            widget_ids.add(wid)
        for key in re.findall(r"\b[0-9a-f]{40}\b", blob, re.I):
            api_keys.add(key)

    shop_domain = urlsplit(MYSHOPIFY or BASE_URL).netloc
    discovery_targets = [
        f"https://cdn.rebuyengine.com/onsite/js/rebuy.js?shop={shop_domain}",
        *REBUY_DISCOVERY_URLS,
    ]

    for url in discovery_targets:
        try:
            resp = session.get(url, timeout=REQUEST_TIMEOUT, verify=False)
            if resp.status_code != 200:
                continue
            text = resp.text
            token_blobs.append(text)
            for wid, key in discovery_regex.findall(text):
                widget_ids.add(wid)
                api_keys.add(key)
        except Exception as exc:  # noqa: BLE001
            logger.warning("Rebuy discovery failed for %s: %s", url, exc)

    if not widget_ids:
        logger.warning("No Rebuy widget ids discovered")
    if not api_keys:
        logger.warning("No Rebuy API keys discovered")

    product_handles = build_rebuy_product_handles(products)

    for product in product_handles:
        for widget in widget_ids:
            for api_key in api_keys:
                input_products = fetch_rebuy_input_products(session, product, widget, api_key, logger)
                for item in input_products:
                    if not isinstance(item, dict):
                        continue
                    base_row = flatten_record({"product": item}, skip_keys={"product.variants"})
                    base_row = normalize_base_fields(base_row, item)
                    variants = item.get("variants") or []
                    if not variants:
                        rows.append(base_row)
                        continue
                    for variant in variants:
                        if not isinstance(variant, dict):
                            continue
                        if variant.get("sku") == "ROUTEINS":
                            continue
                        variant_row = flatten_record({"variant": variant})
                        row = {**base_row, **variant_row}
                        rows.append(row)
    return rows, token_blobs


def probe_globo(
    session: requests.Session,
    logger: logging.Logger,
    collection_html: str,
    pdp_html_map: Dict[str, str],
) -> Tuple[List[Dict[str, Any]], List[str]]:
    rows: List[Dict[str, Any]] = []
    token_blobs: List[str] = []

    params: Dict[str, str] = dict(GLOBO_FALLBACK_PARAMS)

    def harvest_params(text: str) -> None:
        for url in re.findall(r"https://[^\s\"']*apiFilter[^\s\"']*", text):
            try:
                parsed = urlparse(url)
                for key, value in parse_qsl(parsed.query):
                    if key and value and key not in params:
                        params[key] = value
            except Exception:
                continue
        for key in [
            "callback",
            "filter_id",
            "shop",
            "collection",
            "sort_by",
            "country",
            "limit",
            "event",
            "cid",
            "did",
            "page_type",
            "ncp",
        ]:
            if key in params:
                continue
            m = re.search(fr"{key}\\?\"?[:=]\\?\"?([\w\-]+)", text)
            if m:
                params[key] = m.group(1)

    config_match = re.search(r"GloboFilterConfig\s*=\s*(\{.*?\})\s*;", collection_html, re.S)
    if config_match:
        config_raw = config_match.group(1)
        token_blobs.append(config_raw)
        for key in ["shop", "domain", "name"]:
            m = re.search(fr"{key}\s*:\s*\"([^\"]+)\"", config_raw)
            if m and "shop" not in params:
                params["shop"] = m.group(1) if "." in m.group(1) else f"{m.group(1)}.myshopify.com"
        page_id_match = re.search(r"page_id\s*=\s*(\d+)", collection_html)
        if page_id_match and "collection" not in params:
            params["collection"] = page_id_match.group(1)
        country_match = re.search(r"cur_country\s*:\s*\"([A-Z]{2})\"", config_raw)
        if country_match and "country" not in params:
            params["country"] = country_match.group(1)

    harvest_params(collection_html)
    for html in pdp_html_map.values():
        harvest_params(html)
    for url in GLOBO_SCRIPT_URLS:
        try:
            resp = request(session, url, logger)
            if resp.status_code != 200:
                continue
            token_blobs.append(resp.text)
            harvest_params(resp.text)
        except Exception as exc:  # noqa: BLE001
            logger.warning("Globo script fetch failed %s: %s", url, exc)

    params.setdefault("shop", urlsplit(MYSHOPIFY).netloc if MYSHOPIFY else "")
    params.setdefault("event", "init")
    params.setdefault("page_type", "collection")
    params.setdefault("limit", "200")

    required_keys = {"filter_id", "shop", "collection"}
    if not required_keys.issubset(params):
        logger.warning("Missing Globo parameters; returning script captures only")
        return rows, token_blobs

    try:
        resp = session.get(GLOBO_ENDPOINT, params=params, timeout=REQUEST_TIMEOUT)
        token_blobs.append(resp.text)
        if resp.status_code != 200:
            logger.warning("Globo API failed %s", resp.status_code)
            return rows, token_blobs
        text = resp.text
        json_text = re.sub(r"^[^(]*\(", "", text).rstrip(") ;\n")
        data = json.loads(json_text)
        product_list = data.get("product_list") or []
        source_collection = params.get("collection")
        for product in product_list:
            if not isinstance(product, dict):
                continue
            base_row: Dict[str, Any] = {
                "product.id": product.get("id") or product.get("product_id"),
                "product.handle": product.get("handle"),
                "product.title": product.get("title"),
                "product.productType": product.get("product_type") or product.get("productType"),
                "product.vendor": product.get("vendor"),
                "product.description": product.get("description"),
                "product.tags_all": ",".join(product.get("tags") or []),
                "product.images[0].src": product.get("featured_image"),
                "product.onlineStoreUrl": product.get("url") or product.get("link"),
                "product.totalInventory": product.get("ss_inventory_count"),
                "source_collection": source_collection,
            }
            variants = product.get("variants") or []
            if not variants and product.get("ss_variants"):
                variants = product.get("ss_variants")
            if not variants:
                row = {**base_row, **flatten_record({"globo": product})}
                rows.append(row)
                continue
            for variant in variants:
                if not isinstance(variant, dict):
                    continue
                variant_row = {
                    "variant.id": variant.get("id") or variant.get("variant_id"),
                    "variant.title": variant.get("title"),
                    "variant.quantityAvailable": variant.get("quantity")
                    or variant.get("inventoryQuantity")
                    or variant.get("quantityAvailable"),
                    "variant.available": variant.get("available"),
                    "variant.sku": variant.get("sku"),
                    "variant.barcode": variant.get("barcode"),
                }
                extra = flatten_record({"globo": variant})
                row = {**base_row, **variant_row, **extra}
                rows.append(row)
    except Exception as exc:  # noqa: BLE001
        logger.warning("Globo API parse failed: %s", exc)

    return rows, token_blobs


def probe_bundler(
    session: requests.Session, logger: logging.Logger
) -> Tuple[List[Dict[str, Any]], List[str]]:
    rows: List[Dict[str, Any]] = []
    token_blobs: List[str] = []
    try:
        resp = request(session, BUNDLER_STATUS, logger)
        if resp.status_code == 200:
            text = resp.text.strip()
            token_blobs.append(text)
            if text.startswith("var status"):
                cleaned = re.sub(r"^var\s+status\s*=\s*", "", text).rstrip(";")
            else:
                cleaned = text
            data = json.loads(cleaned)
            rows.append(flatten_record({"bundler": {"url": BUNDLER_STATUS, "data": data}}))
    except Exception as exc:  # noqa: BLE001
        logger.warning("Bundler parse failed: %s", exc)
    return rows, token_blobs


def probe_postscript(
    session: requests.Session, logger: logging.Logger, handles: List[str]
) -> Tuple[List[Dict[str, Any]], List[str]]:
    rows: List[Dict[str, Any]] = []
    token_blobs: List[str] = []
    for url in POSTSCRIPT_ENDPOINTS:
        try:
            resp = request(session, url, logger)
            if resp.status_code != 200:
                continue
            text = resp.text.strip()
            token_blobs.append(text)
            try:
                data = json.loads(text)
            except json.JSONDecodeError:
                continue
            if isinstance(data, dict) and data.get("popups"):
                for popup in data.get("popups", []):
                    if not isinstance(popup, dict):
                        continue
                    products = popup.get("products") or []
                    for prod in products:
                        if not isinstance(prod, dict):
                            continue
                        row = flatten_record({"popup": popup, "product": prod})
                        rows.append(row)
        except Exception as exc:  # noqa: BLE001
            logger.warning("Postscript parse failed for %s: %s", url, exc)

    for handle in handles:
        pop_url = (
            f"https://sdk-api-proxy.postscript.io/v2/public/popups/575770/desktop"
            f"?origin={BASE_URL}/products/{handle}"
        )
        try:
            resp = request(session, pop_url, logger)
            if resp.status_code != 200:
                continue
            text = resp.text.strip()
            token_blobs.append(text)
            data = json.loads(text)
            for popup in data.get("popups", []):
                if not isinstance(popup, dict):
                    continue
                products = popup.get("products") or []
                for prod in products:
                    if not isinstance(prod, dict):
                        continue
                    row = flatten_record({"popup": popup, "product": prod})
                    row["product.handle"] = handle
                    rows.append(row)
        except Exception as exc:  # noqa: BLE001
            logger.warning("Postscript PDP probe failed for %s: %s", handle, exc)
    return rows, token_blobs


# ---------------------------------------------------------------------------
# Per-product enrichment
# ---------------------------------------------------------------------------

def enrich_from_html(
    session: requests.Session,
    logger: logging.Logger,
    handle: str,
    html: Optional[str] = None,
) -> Dict[str, List[Dict[str, Any]]]:
    if html is None:
        url = f"{BASE_URL}/products/{handle}"
        resp = request(session, url, logger)
        if resp.status_code != 200:
            return {}
        html = resp.text
    restock = parse_restock(html)
    avada_inv, avada_cdt = parse_avada(html)
    soup = BeautifulSoup(html, "html.parser")
    page_title = (soup.title.string or "").strip() if soup.title else ""
    rows: Dict[str, List[Dict[str, Any]]] = defaultdict(list)

    for vid, payload in restock.items():
        rows["Restock"].append(
            {
                "variant.id": vid,
                "product.handle": handle,
                "product.page_title": page_title,
                "variant.quantityAvailable": payload.get("quantity_available", ""),
                "variant.quantity": payload.get("quantity"),
                "restock.notify_me": payload.get("notify_me", ""),
                "source": "restock",
            }
        )
    for vid, payload in avada_inv.items():
        rows["Avada"].append(
            {
                "variant.id": vid,
                "product.handle": handle,
                "product.page_title": page_title,
                "variant.quantityAvailable": payload.get("quantity_available", ""),
                "variant.quantity": payload.get("quantity"),
                "source": "avada",
            }
        )
    if avada_cdt:
        flat_cdt = flatten_record({"avada": avada_cdt})
        apply_name_value_columns(flat_cdt)
        flat_cdt["product.handle"] = handle
        flat_cdt["product.page_title"] = page_title
        rows["Avada"].append(flat_cdt)
    return rows


# ---------------------------------------------------------------------------
# Workbook assembly
# ---------------------------------------------------------------------------

def build_workbook(
    graphql_rows: List[Dict[str, Any]],
    graphql_access_rows: List[Dict[str, Any]],
    catalog_rows: List[Dict[str, Any]],
    app_rows: Dict[str, List[Dict[str, Any]]],
    rebuy_rows: List[Dict[str, Any]],
    globo_rows: List[Dict[str, Any]],
    bundler_rows: List[Dict[str, Any]],
    postscript_rows: List[Dict[str, Any]],
) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    write_sheet(
        wb, "GraphQL", graphql_rows, column_order=build_column_order(graphql_rows)
    )
    write_sheet(wb, "Storefront_access", graphql_access_rows, use_base=False)
    write_sheet(
        wb,
        "CollectionJSON",
        catalog_rows,
        column_order=build_column_order(catalog_rows),
    )
    app_priority = [
        "product.handle",
        "product.page_title",
        "variant.id",
        "variant.title",
        "variant.quantityAvailable",
        "variant.quantity",
        "restock.notify_me",
        "source",
    ]
    write_sheet(
        wb, "Rebuyengine", rebuy_rows, use_base=True, extra_priority=app_priority
    )
    write_sheet(wb, "Globo", globo_rows, use_base=True, extra_priority=app_priority)
    write_sheet(
        wb,
        "Restock",
        app_rows.get("Restock", []),
        use_base=True,
        extra_priority=app_priority,
    )
    write_sheet(
        wb,
        "Avada",
        app_rows.get("Avada", []),
        use_base=True,
        extra_priority=app_priority,
    )
    write_sheet(wb, "Bundler", bundler_rows, use_base=False)
    write_sheet(wb, "Postscript", postscript_rows, use_base=False)
    return wb


# ---------------------------------------------------------------------------
# Main entry
# ---------------------------------------------------------------------------

def main() -> None:
    logger = configure_logger()
    sync_base_probe_constants()
    session = build_session(logger)

    collection_html = base_probe.fetch_collection_html(session, logger)
    catalog = fetch_collection_products(session, logger)
    handles = [p.get("handle") for p in catalog if isinstance(p, dict) and p.get("handle")]

    pdp_html_map = fetch_product_htmls(session, logger, handles)

    token_blobs: List[str] = []
    graphql_rows: List[Dict[str, Any]] = []
    graphql_access_rows: List[Dict[str, Any]] = []

    catalog_rows: List[Dict[str, Any]] = []
    for product in catalog:
        if not isinstance(product, dict):
            continue
        base = flatten_record({"product": product}, skip_keys={"product.variants"})
        base = normalize_base_fields(base, product)
        grouped_tags = base_probe.group_tags_for_columns(product.get("tags") or [])
        base.update({k: ",".join(v) for k, v in grouped_tags.items()})
        variants = product.get("variants") or []
        if not variants:
            apply_name_value_columns(base)
            catalog_rows.append(prune_probe_fields(base))
        else:
            for variant in variants:
                variant_flat = flatten_record({"variant": variant})
                row = {**base, **variant_flat}
                apply_name_value_columns(row)
                catalog_rows.append(prune_probe_fields(row))

    app_rows: Dict[str, List[Dict[str, Any]]] = defaultdict(list)

    for handle in handles:
        per_product = enrich_from_html(session, logger, handle, pdp_html_map.get(handle))
        for key, rows in per_product.items():
            app_rows[key].extend(rows)

    html_blobs_for_tokens = [collection_html, *pdp_html_map.values()]
    for blob in html_blobs_for_tokens:
        for token in TOKEN_REGEX.findall(blob):
            token_blobs.append(token)
    rebuy_rows_raw, rebuy_tokens = probe_rebuy(
        session, logger, html_blobs_for_tokens, catalog
    )
    globo_rows_raw, globo_tokens = probe_globo(
        session, logger, collection_html, pdp_html_map
    )
    bundler_rows_raw, bundler_tokens = probe_bundler(session, logger)
    postscript_rows_raw, postscript_tokens = probe_postscript(
        session, logger, handles
    )

    token_blobs.extend(rebuy_tokens)
    token_blobs.extend(globo_tokens)
    token_blobs.extend(bundler_tokens)
    token_blobs.extend(postscript_tokens)

    combined_html_tokens = "\n".join(
        [collection_html, *pdp_html_map.values(), *token_blobs]
    ).strip()
    storefront_rows, access_rows = base_probe.gather_storefront_data(
        session, combined_html_tokens, logger
    )
    graphql_rows = [prune_probe_fields(dict(row)) for row in storefront_rows]
    graphql_access_rows = access_rows

    rebuy_rows = finalize_rows(rebuy_rows_raw)
    globo_rows = finalize_rows(globo_rows_raw)
    bundler_rows = finalize_rows(bundler_rows_raw)
    postscript_rows = finalize_rows(postscript_rows_raw)
    app_rows = {k: finalize_rows(v) for k, v in app_rows.items()}

    wb = build_workbook(
        graphql_rows,
        graphql_access_rows,
        catalog_rows,
        app_rows,
        rebuy_rows,
        globo_rows,
        bundler_rows,
        postscript_rows,
    )

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    output_path = OUTPUT_DIR / f"{BRAND_SLUG}_apps_{timestamp}.xlsx"
    wb.save(output_path)
    logger.info("Workbook written: %s", output_path.resolve())


if __name__ == "__main__":
    main()
