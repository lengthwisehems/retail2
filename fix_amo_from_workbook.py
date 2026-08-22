# -*- coding: utf-8 -*-
"""
Apply the post-cleanup AMO edit workbook to denim_analytics.

After cleanup_amo_database.py finished, a review found leftover bad values that
came from the AMO_2026-08-18_23-53-21 scrape (old code). This applies the fixes
recorded in that review workbook, per sheet, keyed by primary-key id:

  * ACTION contains "Remove"  -> DELETE that row by its id.
  * ACTION contains "Edit"    -> UPDATE that row: for every "NEW <col>" cell that
                                 is filled in, set <col> to that value. Blank NEW
                                 cells are left as-is (keep the current value).
  * anything else / blank ACTION -> skip.

"NEW <col>" maps to the sheet's real column by name (e.g. "NEW Inseam" -> inseam,
"NEW variant_title" -> variant_title). lookup has no edits and is ignored.

Everything is keyed by the int primary key (style_info_id / style_metric_id /
variant_metric_id), which is the clustered PK, so each write is a fast seek - no
brand scans, no timeouts. DRY_RUN prints the full plan and writes nothing.
Commits in batches; every log line is Central-time stamped.

RUN (Windows cmd):
    pip install pymssql openpyxl
    set SQL_USERNAME=carrieboromisa
    set SQL_PASSWORD=your-password
    python "C:\\path\\to\\fix_amo_from_workbook.py"
"""
from __future__ import annotations

import os
import sys
import datetime as dt
from decimal import Decimal
from typing import Dict, List, Tuple

# ===========================================================================
# CONFIG
# ===========================================================================
DRY_RUN = True
BRAND = "AMO"

WORKBOOK = r"AMO_DB_Values_20260821_152509_share.xlsx"

# sheet -> primary-key column. lookup is intentionally absent (no edits).
SHEETS = {
    "style_info":      "style_info_id",
    "style_metrics":   "style_metric_id",
    "variant_metrics": "variant_metric_id",
}

SQL_SERVER   = os.environ.get("SQL_SERVER",   "denim-sql.database.windows.net")
SQL_DATABASE = os.environ.get("SQL_DATABASE", "denim_analytics")
SQL_USERNAME = os.environ.get("SQL_USERNAME", "carrieboromisa")
SQL_PASSWORD = os.environ.get("SQL_PASSWORD", "")

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
COMMIT_EVERY = 2000
VC = "CAST(%s AS varchar(255))"   # brand safety-check stays an index seek


# ===========================================================================
# Central-time logging
# ===========================================================================
def _tz():
    try:
        from zoneinfo import ZoneInfo
        return ZoneInfo("America/Chicago")
    except Exception:
        return dt.timezone(dt.timedelta(hours=-6))
_TZ = _tz()


def log(msg: str = "") -> None:
    stamp = dt.datetime.now(_TZ).strftime("%Y-%m-%d %H:%M:%S")
    for part in str(msg).split("\n"):
        print(f"[{stamp}] {part}")


# ===========================================================================
# Value helpers
# ===========================================================================
def s(v) -> str:
    """Full-precision string; no scientific notation, trimmed."""
    if v is None:
        return ""
    if isinstance(v, bool):
        return "1" if v else "0"
    if isinstance(v, Decimal):
        return f"{v:f}"
    if isinstance(v, float):
        return str(int(v)) if v.is_integer() else repr(v)
    return str(v).strip()


def is_blank(v) -> bool:
    return v is None or (isinstance(v, str) and v.strip() == "")


def resolve(path: str) -> str:
    return path if os.path.isabs(path) else os.path.join(SCRIPT_DIR, path)


# ===========================================================================
# Planning (pure - no DB) so it can be tested offline
# ===========================================================================
def _new_col_map(header: List) -> Dict[int, str]:
    """index of each 'NEW <col>' header -> real DB column name."""
    actual = {}
    for h in header:
        if isinstance(h, str) and h and not h.lower().startswith("new "):
            actual.setdefault(h.strip().lower(), h.strip())
    out: Dict[int, str] = {}
    for i, h in enumerate(header):
        if isinstance(h, str) and h.lower().startswith("new "):
            base = h[4:].strip()
            out[i] = actual.get(base.lower(), base.lower())
    return out


def build_plan(wb) -> Dict[str, dict]:
    """{sheet: {'pk': col, 'removes': [id...], 'updates': [(id, {col: val})...]}}"""
    plan: Dict[str, dict] = {}
    for sheet, pk in SHEETS.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        it = ws.iter_rows(values_only=True)
        header = list(next(it))
        idx = {h: i for i, h in enumerate(header) if h}
        if "ACTION" not in idx or pk not in idx:
            log(f"{sheet}: no ACTION/{pk} column - skipped")
            continue
        ai, pki = idx["ACTION"], idx[pk]
        newmap = _new_col_map(header)     # col-index -> db column
        removes: List = []
        updates: List[Tuple] = []
        skipped_noop = 0
        for r in it:
            if all(v in (None, "") for v in r):
                continue
            action = s(r[ai]).lower()
            key = r[pki]
            if key in (None, ""):
                continue
            if "remove" in action:
                removes.append(key)
            elif "edit" in action:
                ups = {}
                for ci, dbcol in newmap.items():
                    nv = r[ci] if ci < len(r) else None
                    if is_blank(nv):
                        continue
                    cur = r[idx[dbcol]] if dbcol in idx else None
                    if s(nv) != s(cur):
                        ups[dbcol] = s(nv)
                if ups:
                    updates.append((key, ups))
                else:
                    skipped_noop += 1
        plan[sheet] = {"pk": pk, "removes": removes, "updates": updates,
                       "noop": skipped_noop}
    return plan


# ===========================================================================
# Main
# ===========================================================================
def main() -> None:
    from openpyxl import load_workbook
    path = resolve(WORKBOOK)
    if not os.path.exists(path):
        sys.exit(f"ERROR: workbook not found at {path}")
    log("=" * 60)
    log(f"AMO WORKBOOK FIX  ({'DRY RUN' if DRY_RUN else 'LIVE RUN'})")
    log(f"Workbook: {path}")
    log("=" * 60)

    wb = load_workbook(path, data_only=True, read_only=True)
    plan = build_plan(wb)

    total_fields = 0
    for sheet, p in plan.items():
        nf = sum(len(u) for _, u in p["updates"])
        total_fields += nf
        log(f"{sheet:16} remove={len(p['removes']):>4}  "
            f"edit_rows={len(p['updates']):>5}  field_writes={nf:>5}"
            f"  (skipped {p['noop']} no-op edit rows)")

    # sample of each sheet's edits so the dry run is reviewable
    for sheet, p in plan.items():
        if p["updates"]:
            log(f"--- {sheet} sample edits ---")
            for key, ups in p["updates"][:6]:
                cols = ", ".join(f"{c}={v!r}" for c, v in ups.items())
                log(f"   {p['pk']}={key}: {cols}")
        if p["removes"]:
            log(f"--- {sheet} sample removes: "
                + ", ".join(str(k) for k in p["removes"][:8])
                + (" ..." if len(p["removes"]) > 8 else ""))

    if DRY_RUN:
        log("DRY RUN complete - nothing written.")
        return

    import pymssql
    conn = pymssql.connect(server=SQL_SERVER, user=SQL_USERNAME, password=SQL_PASSWORD,
                           database=SQL_DATABASE, timeout=600, login_timeout=60)
    cur = conn.cursor()
    try:
        for sheet, p in plan.items():
            pk = p["pk"]
            # removes first
            n = 0
            for key in p["removes"]:
                cur.execute(f"DELETE FROM {sheet} WHERE {pk}=%s AND brand={VC}",
                            (key, BRAND))
                n += 1
                if n % COMMIT_EVERY == 0:
                    conn.commit()
                    log(f"   {sheet}: removed {n}/{len(p['removes'])}...")
            if p["removes"]:
                conn.commit()
                log(f"   {sheet}: removed {len(p['removes'])} (committed)")
            # updates
            n = 0
            for key, ups in p["updates"]:
                assigns = ", ".join(f"[{c}]=%s" for c in ups)
                cur.execute(f"UPDATE {sheet} SET {assigns} WHERE {pk}=%s AND brand={VC}",
                            list(ups.values()) + [key, BRAND])
                n += 1
                if n % COMMIT_EVERY == 0:
                    conn.commit()
                    log(f"   {sheet}: updated {n}/{len(p['updates'])}...")
            if p["updates"]:
                conn.commit()
                log(f"   {sheet}: updated {len(p['updates'])} (committed)")
        log("Done - all edits committed.")
    except Exception:
        conn.rollback()
        log("ERROR - current step rolled back. Earlier committed steps are kept; "
            "re-run to finish (edits are idempotent, removes match nothing 2nd time).")
        raise
    finally:
        conn.close()


if __name__ == "__main__":
    main()
