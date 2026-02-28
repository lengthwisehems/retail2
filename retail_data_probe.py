"""Brand-agnostic probe that inspects Shopify collection feeds and Storefront APIs."""

from __future__ import annotations

import argparse
import html
import json
import logging
import re
import time
from datetime import datetime, timezone
from pathlib import Path
from collections import Counter, defaultdict
from typing import Any, Dict, Iterable, List, Optional, Sequence, Set, Tuple
from urllib.parse import parse_qsl, urlencode, urljoin, urlsplit, urlunsplit

import requests
import urllib3
from bs4 import BeautifulSoup
from openpyxl import Workbook
from requests.adapters import HTTPAdapter, Retry

# ---------------------------------------------------------------------------
# Brand-specific configuration
# ---------------------------------------------------------------------------
BRAND = "Citizens_of_Humanity"
COLLECTION_URL = [
    "https://citizensofhumanity.com/collections/womens-jeans",
]
MYSHOPIFY = "citizens-of-humanity.myshopify.com"
GRAPHQL = "https://citizens-of-humanity.myshopify.com/api/unstable/graphql.json"
X_SHOPIFY_STOREFRONT_ACCESS_TOKEN = ["a1b87221a13b15123b1c8b79a866f388"]
GRAPHQL_FILTER_TAG = ""
STOREFRONT_COLLECTION_HANDLES: List[str] = ["womens-jeans"]
SEARCHSPRING_SITE_ID = ""
SEARCHSPRING_URL = ""
SEARCHSPRING_EXTRA_PARAMS: Dict[str, Any] = {}
METAFIELD_IDENTIFIERS: List[Tuple[str, str]] = []
COLLECTION_TITLE_MAP: Dict[str, str] = {}
VIEW_JSON_ENRICHMENT_ENABLED = False
VIEW_JSON_FIELDS: List[str] = []
VIEW_JSON_PROBE_LIMIT = 5

# ---------------------------------------------------------------------------
# Derived paths and constants
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = BASE_DIR / "Output"
OUTPUT_DIR.mkdir(exist_ok=True)

BRAND_SLUG = BRAND.lower().replace(" ", "_") or "brand"
LOG_PATH = BASE_DIR / f"{BRAND_SLUG}_probe_run.log"
FALLBACK_LOG_PATH = OUTPUT_DIR / f"{BRAND_SLUG}_probe_run.log"

REQUEST_TIMEOUT = 30
TRANSIENT_STATUS = {429, 500, 502, 503, 504}
GRAPHQL_PAGE_SIZE = 100
MAX_SCRIPT_FETCHES = 25
TOKEN_REGEX = re.compile(r"\b[0-9a-f]{32}\b", re.IGNORECASE)

DEFAULT_GRAPHQL_VERSIONS = [
    "api/2025-10/graphql.json",
    "api/2024-01/graphql.json",
    "api/2025-01/graphql.json",
    "api/2025-07/graphql.json",
    "api/2025-04/graphql.json",
    "api/unstable/graphql.json",
    "api/2024-04/graphql.json",
    "api/2023-01/graphql.json",
    "api/2023-04/graphql.json",
]

COLUMN_ORDER_BASE: Tuple[str, ...] = (
    "product.id",
    "product.handle",
    "product.published_at",
    "product.created_at",
    "product.title",
    "product.productType",
    "product.tags_all",
    "product.vendor",
    "product.description",
    "product.descriptionHtml",
    "variant.title",
    "variant.option1",
    "variant.option2",
    "variant.option3",
    "variant.price",
    "variant.compare_at_price",
    "variant.available",
    "variant.quantityAvailable",
    "product.totalInventory",
    "variant.id",
    "variant.sku",
    "variant.barcode",
    "product.images[0].src",
    "product.onlineStoreUrl",
)

DEFAULT_FORBIDDEN_FIELDS: Dict[str, Set[str]] = {
    "ProductVariant": {
        "components",
        "groupedBy",
        "quantityPriceBreaks",
        "sellingPlanAllocations",
        "sellingPlanGroups",
        "storeAvailability",
    }
}

# Additional fields to skip in queries/outputs
EXTRA_FORBIDDEN_COLUMNS: Set[str] = {
    "product.collections.pageInfo.endCursor",
    "product.collections.pageInfo.hasNextPage",
    "product.encodedVariantAvailability",
    "product.encodedVariantExistence",
    "product.featuredImage.height",
    "product.featuredImage.thumbhash",
    "product.featuredImage.width",
    "product.images.pageInfo.endCursor",
    "product.images.pageInfo.hasNextPage",
    "product.isGiftCard",
    "product.media.pageInfo.endCursor",
    "product.media.pageInfo.hasNextPage",
    "products_edge_cursor",
    "variant.currentlyNotInStock",
    "variant.image.height",
    "variant.image.id",
    "variant.image.thumbhash",
    "variant.image.width",
    "variant.quantityRule.minimum",
    "variant_edge_cursor",
    "variants_endCursor",
    "variants_hasNextPage",
}


def parse_metafield_identifiers(raw: str) -> List[Tuple[str, str]]:
    identifiers: List[Tuple[str, str]] = []
    if not raw:
        return identifiers
    parts = [part.strip() for part in raw.split(",") if part.strip()]
    seen: Set[Tuple[str, str]] = set()
    for part in parts:
        if ":" not in part:
            continue
        namespace, key = part.split(":", 1)
        namespace = namespace.strip()
        key = key.strip()
        if not namespace or not key:
            continue
        tup = (namespace, key)
        if tup not in seen:
            identifiers.append(tup)
            seen.add(tup)
    return identifiers


def parse_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return value != 0
    return str(value).strip().lower() in {"1", "true", "yes", "y", "on"}


def parse_view_json_fields(raw: str) -> List[str]:
    if not raw:
        return []
    fields: List[str] = []
    seen: Set[str] = set()
    for part in raw.split(","):
        key = part.strip()
        if not key or key in seen:
            continue
        fields.append(key)
        seen.add(key)
    return fields


def normalize_tokens(value: Any) -> List[str]:
    """Return an ordered list of unique, non-empty tokens."""

    if not value:
        return []

    tokens: List[str] = []
    if isinstance(value, str):
        token = value.strip()
        if token:
            tokens.append(token)
    elif isinstance(value, (list, tuple, set)):
        for item in value:
            if not isinstance(item, str):
                continue
            token = item.strip()
            if token:
                tokens.append(token)

    seen: Set[str] = set()
    ordered: List[str] = []
    for token in tokens:
        if token in seen:
            continue
        seen.add(token)
        ordered.append(token)
    return ordered


def format_error_note(errors: Optional[List[Dict[str, Any]]]) -> str:
    """Summarize GraphQL errors for logging and the Storefront_access sheet.

    This keeps the count while appending the first error's path/message so
    entries like "errors:1" have immediate context when a probe returns HTTP 200
    but Shopify still reports GraphQL errors.
    """

    if not errors:
        return "errors:0"

    first = errors[0] or {}
    path = first.get("path") or []
    message = first.get("message") or first.get("error") or ""
    path_str = ".".join(str(p) for p in path if p is not None)

    details: List[str] = []
    if path_str:
        details.append(f"path={path_str}")
    if message:
        details.append(f"msg={message}")

    suffix = f":{' | '.join(details)}" if details else ""
    return f"errors:{len(errors)}{suffix}"

FALLBACK_COLLECTION_QUERY = """
query CollectionFallback($handle: String!, $cursor: String, $pageSize: Int!) {
  collection(handle: $handle) {
    id
    handle
    title
    products(first: $pageSize, after: $cursor) {
      pageInfo {
        hasNextPage
        endCursor
      }
      edges {
        cursor
        node {
          id
          handle
          title
          productType
          tags
          vendor
          onlineStoreUrl
          createdAt
          updatedAt
          publishedAt
          variants(first: 100) {
            pageInfo {
              hasNextPage
              endCursor
            }
            edges {
              cursor
              node {
                id
                title
                sku
                availableForSale
                price {
                  amount
                  currencyCode
                }
              }
            }
          }
        }
      }
    }
  }
}
"""

def build_metafields_selection() -> str:
    if not METAFIELD_IDENTIFIERS:
        return ""
    identifiers_literal = ", ".join(
        f'{{namespace: "{ns}", key: "{key}"}}' for ns, key in METAFIELD_IDENTIFIERS
    )
    return (
        "metafields(identifiers: ["
        + identifiers_literal
        + "]) {\n  namespace\n  key\n  type\n  value\n}"
    )


def build_fallback_products_query() -> str:
    metafields_selection = build_metafields_selection()
    metafields_block = f"\n        {metafields_selection}" if metafields_selection else ""
    return f"""
query ProductsFallback($cursor: String, $pageSize: Int!, $query: String) {{
  products(first: $pageSize, after: $cursor, query: $query) {{
    pageInfo {{
      hasNextPage
      endCursor
    }}
    edges {{
      cursor
      node {{
        id
        handle
        title
        description
        productType
        tags
        vendor
        onlineStoreUrl
        createdAt
        updatedAt
        publishedAt
        collections(first: 50) {{
          edges {{
            node {{
              id
              handle
              title
            }}
          }}
        }}
        options {{
          name
          values
        }}{metafields_block}
        variants(first: 100) {{
          pageInfo {{
            hasNextPage
            endCursor
          }}
          edges {{
            cursor
            node {{
              id
              title
              sku
              availableForSale
              price {{
                amount
                currencyCode
              }}
            }}
          }}
        }}
      }}
    }}
  }}
}}
"""

SHOP_PROBE_QUERY = "query { shop { name primaryDomain { url } } }"

INTROSPECTION_QUERY = """
query ($typeName: String!) {
  __type(name: $typeName) {
    name
    fields {
      name
      args {
        name
        defaultValue
        type {
          kind
          name
          ofType {
            kind
            name
            ofType {
              kind
              name
              ofType {
                kind
                name
              }
            }
          }
        }
      }
      type {
        kind
        name
        ofType {
          kind
          name
          ofType {
            kind
            name
            ofType {
              kind
              name
            }
          }
        }
      }
    }
  }
}
"""

FILTER_PROBE_QUERIES = [
    """
query FiltersProbe($handle: String!) {
  collection(handle: $handle) {
    products(first: 1) {
      filters {
        id
        label
        type
        values {
          id
          label
          count
          input
        }
      }
    }
  }
}
    """,
    """
query FiltersProbe($handle: String!) {
  collection(handle: $handle) {
    products(first: 1) {
      productFilters {
        id
        label
        type
        values {
          id
          label
          count
          input
        }
      }
    }
  }
}
    """,
]

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


def _primary_collection_url() -> Optional[str]:
    if isinstance(COLLECTION_URL, (list, tuple)):
        return COLLECTION_URL[0] if COLLECTION_URL else None
    return COLLECTION_URL


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def configure_logging() -> logging.Logger:
    logger = logging.getLogger(f"retail_probe_{BRAND_SLUG}")
    logger.setLevel(logging.INFO)
    formatter = logging.Formatter("%(asctime)s %(levelname)s %(message)s")

    handler: logging.Handler
    try:
        handler = logging.FileHandler(LOG_PATH, mode="a", encoding="utf-8")
    except OSError as exc:
        fallback = FALLBACK_LOG_PATH
        handler = logging.FileHandler(fallback, mode="a", encoding="utf-8")
        logger.warning(
            "Primary log path %s unavailable (%s); using %s", LOG_PATH, exc, fallback
        )

    handler.setFormatter(formatter)
    logger.addHandler(handler)

    stream_handler = logging.StreamHandler()
    stream_handler.setFormatter(formatter)
    logger.addHandler(stream_handler)
    return logger


def build_session() -> requests.Session:
    session = requests.Session()
    retry = Retry(
        total=5,
        backoff_factor=0.5,
        status_forcelist=TRANSIENT_STATUS,
        allowed_methods=("GET", "POST"),
    )
    adapter = HTTPAdapter(max_retries=retry)
    session.mount("http://", adapter)
    session.mount("https://", adapter)
    session.headers.update(
        {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            )
        }
    )
    session.verify = False
    return session


def normalize_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, Path):
        return value.as_posix()
    return json.dumps(value, ensure_ascii=False, sort_keys=True)


def flatten_value(value: Any, prefix: str) -> Dict[str, Any]:
    items: Dict[str, Any] = {}
    if isinstance(value, dict):
        for key, inner in value.items():
            new_prefix = f"{prefix}.{key}" if prefix else key
            items.update(flatten_value(inner, new_prefix))
    elif isinstance(value, list):
        for index, inner in enumerate(value):
            new_prefix = f"{prefix}[{index}]" if prefix else f"[{index}]"
            items.update(flatten_value(inner, new_prefix))
    else:
        items[prefix] = value
    return items


def flatten_record(record: Dict[str, Any]) -> Dict[str, Any]:
    flat: Dict[str, Any] = {}
    for key, value in record.items():
        flat.update(flatten_value(value, key))
    return flat


def extract_graphql_variant_entries(
    variants_connection: Any,
) -> List[Dict[str, Any]]:
    if not isinstance(variants_connection, dict):
        return []

    entries: List[Dict[str, Any]] = []
    seen_ids: Set[Any] = set()

    edges = variants_connection.get("edges") or []
    for edge in edges:
        if not isinstance(edge, dict):
            continue
        node = edge.get("node")
        if not isinstance(node, dict):
            continue
        vid = node.get("id")
        if vid is not None:
            if vid in seen_ids:
                continue
            seen_ids.add(vid)
        entries.append({"cursor": edge.get("cursor", ""), "node": node})

    nodes = variants_connection.get("nodes") or []
    if isinstance(nodes, list):
        for node in nodes:
            if not isinstance(node, dict):
                continue
            vid = node.get("id")
            if vid is not None and vid in seen_ids:
                continue
            if vid is not None:
                seen_ids.add(vid)
            entries.append({"cursor": "", "node": node})

    return entries


def build_option_columns(options: Sequence[Dict[str, Any]]) -> Dict[str, str]:
    columns: Dict[str, str] = {}
    aggregate_values: List[str] = []
    for option in options or []:
        if not isinstance(option, dict):
            continue
        name = str(option.get("name") or "").strip()
        values = [str(v).strip() for v in option.get("values") or [] if str(v).strip()]
        if not values:
            continue
        joined = ", ".join(values)
        if name:
            columns[f"product.options.{name}"] = joined
        else:
            aggregate_values.append(joined)
    if aggregate_values and "product.options" not in columns:
        columns["product.options"] = ", ".join(aggregate_values)
    return columns


def sanitize_dynamic_header(value: str) -> str:
    cleaned = re.sub(r"[^0-9A-Za-z]+", "_", str(value).strip()).strip("_")
    return cleaned or "value"


def apply_name_value_columns(row: Dict[str, Any]) -> None:
    replacements: Dict[str, Any] = {}
    to_remove: List[str] = []
    for key, value in list(row.items()):
        if not key.endswith(".name"):
            continue
        prefix = key[:-5]
        name_value = str(value).strip()
        value_key = f"{prefix}.value"
        if not name_value or value_key not in row:
            continue
        new_key = f"{prefix}.{sanitize_dynamic_header(name_value)}"
        replacements[new_key] = row[value_key]
        to_remove.extend([key, value_key])
    for key in to_remove:
        row.pop(key, None)
    row.update(replacements)


def extract_first_image_src(product: Dict[str, Any]) -> Optional[str]:
    images = product.get("images")
    if isinstance(images, list):
        for item in images:
            if isinstance(item, dict):
                src = item.get("src") or item.get("url") or item.get("originalSrc")
                if src:
                    return src
    elif isinstance(images, dict):
        edges = images.get("edges") or []
        for edge in edges:
            if not isinstance(edge, dict):
                continue
            node = edge.get("node")
            if isinstance(node, dict):
                src = node.get("src") or node.get("url") or node.get("originalSrc")
                if src:
                    return src
    return None


def normalize_money_field(row: Dict[str, Any], base_key: str) -> None:
    amount_key = f"{base_key}.amount"
    if base_key not in row and amount_key in row:
        row[base_key] = row.pop(amount_key)
    elif amount_key in row and row.get(base_key) == row.get(amount_key):
        row.pop(amount_key, None)


def remove_matching_keys(
    row: Dict[str, Any], prefixes: Sequence[str], *, allowed: Optional[Sequence[str]] = None
) -> None:
    allowed_set = set(allowed or [])
    for key in list(row.keys()):
        lowered = key.lower()
        if "position" in lowered:
            row.pop(key, None)
            continue
        for prefix in prefixes:
            if key.startswith(prefix) and key not in allowed_set:
                row.pop(key, None)
                break


def extract_field_from_error_path(path: Sequence[Any]) -> Optional[str]:
    for segment in reversed(path or []):
        if isinstance(segment, str):
            return segment
    return None


def infer_error_target_type(path: Sequence[Any]) -> str:
    string_segments = [segment for segment in path if isinstance(segment, str)]
    return "ProductVariant" if "variants" in string_segments else "Product"


def populate_variant_options(row: Dict[str, Any], variant: Optional[Dict[str, Any]]) -> None:
    if variant is None:
        return
    selected = variant.get("selectedOptions") or []
    for index, option in enumerate(selected):
        if index >= 3 or not isinstance(option, dict):
            continue
        value = option.get("value")
        if value and not row.get(f"variant.option{index + 1}"):
            row[f"variant.option{index + 1}"] = value


def finalize_common_row(
    row: Dict[str, Any],
    product: Dict[str, Any],
    variant: Optional[Dict[str, Any]],
    *,
    source: str,
) -> None:
    tags = product.get("tags") or []
    if isinstance(tags, list) and tags:
        row["product.tags_all"] = ", ".join(str(tag) for tag in tags if str(tag))
    for key in list(row.keys()):
        if key.startswith("product.tags["):
            row.pop(key, None)

    option_columns = build_option_columns(product.get("options") or [])
    for key, value in option_columns.items():
        row[key] = value

    image_src = extract_first_image_src(product)
    if image_src:
        row["product.images[0].src"] = image_src

    remove_matching_keys(
        row,
        [
            "product.images[",
            "product.images.edges",
            "product.media.edges",
            "product.collections.edges",
            "product.options[",
            "variant.selectedOptions[",
            "variant.featured_image",
        ],
        allowed=["product.images[0].src", "variant.featured_image.src"],
    )

    normalize_money_field(row, "variant.price")
    normalize_money_field(row, "variant.compare_at_price")

    if "product.totalInventory" not in row:
        for candidate in ("product.total_inventory", "product.ss_available_qty"):
            if candidate in row:
                row["product.totalInventory"] = row.pop(candidate)
                break

    if "product.onlineStoreUrl" not in row and "product.url" in row:
        row["product.onlineStoreUrl"] = row.pop("product.url")

    if "product.id" not in row and "product.ss_id" in row:
        row["product.id"] = row.pop("product.ss_id")

    if "product.title" not in row and "product.name" in row:
        row["product.title"] = row.pop("product.name")

    if "variant.available" not in row:
        for candidate in (
            "variant.availableForSale",
            "variant.available_for_sale",
            "variant.available_for_sale?",
        ):
            if candidate in row:
                row["variant.available"] = row.pop(candidate)
                break

    if "variant.quantityAvailable" not in row:
        for candidate in (
            "variant.quantity_available",
            "variant.inventory_quantity",
        ):
            if candidate in row:
                row["variant.quantityAvailable"] = row.pop(candidate)
                break

    populate_variant_options(row, variant)

    apply_name_value_columns(row)

    if source == "storefront":
        if "product.publishedAt" in row and "product.published_at" not in row:
            row["product.published_at"] = row.pop("product.publishedAt")
        if "product.createdAt" in row and "product.created_at" not in row:
            row["product.created_at"] = row.pop("product.createdAt")
        if variant and "availableForSale" in variant and "variant.available" not in row:
            row["variant.available"] = variant.get("availableForSale")
    else:
        if "product.published_at" not in row and "product_published_at" in row:
            row["product.published_at"] = row.get("product_published_at")
        if "product_published_at" in row:
            row.pop("product_published_at", None)
        if "product.productType" not in row and "product.product_type" in row:
            row["product.productType"] = row.pop("product.product_type")
        if "product.body_html" in row:
            row.setdefault("product.descriptionHtml", row["product.body_html"])
            row.setdefault("product.description", row["product.body_html"])
            row.pop("product.body_html", None)

    if "variant.compare_at_price" not in row and variant is not None:
        compare_candidates = (
            variant.get("compareAtPrice"),
            variant.get("compare_at_price"),
        )
        for candidate in compare_candidates:
            if isinstance(candidate, dict):
                amount = candidate.get("amount")
                if amount is not None:
                    row["variant.compare_at_price"] = amount
                    break
            elif candidate not in (None, ""):
                row["variant.compare_at_price"] = candidate
                break

    if "variant.price" not in row and variant is not None:
        price_candidates = (
            variant.get("price"),
            variant.get("priceV2"),
        )
        for candidate in price_candidates:
            if isinstance(candidate, dict):
                amount = candidate.get("amount")
                if amount is not None:
                    row["variant.price"] = amount
                    break
            elif candidate not in (None, ""):
                row["variant.price"] = candidate
                break

    if variant is not None:
        for idx in range(1, 4):
            option_key = f"option{idx}"
            alt_key = f"variant.{option_key}"
            if alt_key not in row and option_key in variant:
                row[alt_key] = variant.get(option_key)

    for forbidden in list(EXTRA_FORBIDDEN_COLUMNS):
        row.pop(forbidden, None)


def finalize_json_row(row: Dict[str, Any], product: Dict[str, Any], variant: Optional[Dict[str, Any]]) -> None:
    finalize_common_row(row, product, variant, source="json")


def finalize_storefront_row(
    row: Dict[str, Any], product: Dict[str, Any], variant: Optional[Dict[str, Any]]
) -> None:
    finalize_common_row(row, product, variant, source="storefront")


def extract_collections(product: Dict[str, Any], collection_info: Dict[str, Any]) -> Tuple[List[str], List[str]]:
    handles: List[str] = []
    titles: List[str] = []
    collections = product.get("collections")
    if isinstance(collections, dict):
        edges = collections.get("edges") or []
        nodes = collections.get("nodes") or []
        if nodes:
            for node in nodes:
                if not isinstance(node, dict):
                    continue
                handle = node.get("handle")
                title = node.get("title")
                if handle:
                    handles.append(str(handle))
                if title:
                    titles.append(str(title))
        elif edges:
            for edge in edges:
                node = edge.get("node") if isinstance(edge, dict) else None
                if not isinstance(node, dict):
                    continue
                handle = node.get("handle")
                title = node.get("title")
                if handle:
                    handles.append(str(handle))
                if title:
                    titles.append(str(title))

    fallback_handle = collection_info.get("collection_handle")
    fallback_title = collection_info.get("collection_title")
    if fallback_handle and fallback_handle not in handles:
        handles.append(fallback_handle)
    if fallback_title and fallback_title not in titles:
        titles.append(fallback_title)
    if COLLECTION_TITLE_MAP and handles:
        for h in handles:
            if h in COLLECTION_TITLE_MAP and COLLECTION_TITLE_MAP[h] not in titles:
                titles.append(COLLECTION_TITLE_MAP[h])
    return handles, titles


def collect_metafields(product: Dict[str, Any]) -> List[Dict[str, Any]]:
    metafields: List[Dict[str, Any]] = []
    raw_metafields = product.get("metafields")
    if isinstance(raw_metafields, list):
        for mf in raw_metafields:
            if isinstance(mf, dict):
                metafields.append(
                    {
                        "namespace": mf.get("namespace"),
                        "key": mf.get("key"),
                        "type": mf.get("type"),
                        "value": mf.get("value"),
                    }
                )
    elif isinstance(raw_metafields, dict):
        # Support alias-based selections like mf_0: metafield(...)
        for value in raw_metafields.values():
            if isinstance(value, dict):
                metafields.append(
                    {
                        "namespace": value.get("namespace"),
                        "key": value.get("key"),
                        "type": value.get("type"),
                        "value": value.get("value"),
                    }
                )
    return metafields


def derive_filter_values(product: Dict[str, Any], metafields: List[Dict[str, Any]]) -> Dict[str, List[str]]:
    filters: Dict[str, Set[str]] = defaultdict(set)
    product_type = product.get("productType")
    vendor = product.get("vendor")
    tags = product.get("tags") or []
    options = product.get("options") or []

    if product_type:
        filters["productType"].add(str(product_type))
    if vendor:
        filters["vendor"].add(str(vendor))
    if isinstance(tags, list):
        for tag in tags:
            if tag:
                filters["tags"].add(str(tag))

    if isinstance(options, list):
        for opt in options:
            if not isinstance(opt, dict):
                continue
            name = opt.get("name") or opt.get("title")
            values = opt.get("values") or []
            if not name:
                continue
            for val in values:
                if val:
                    filters[str(name)].add(str(val))

    for mf in metafields:
        ns = mf.get("namespace")
        key = mf.get("key")
        value = mf.get("value")
        if ns and key and value not in (None, ""):
            filters[f"{ns}:{key}"].add(str(value))

    return {k: sorted(v) for k, v in filters.items() if v}


def normalize_filter_name(name: str) -> str:
    cleaned = re.sub(r"[^a-z0-9]+", "_", str(name).lower()).strip("_")
    return cleaned or "unnamed"


FILTER_COLUMN_SKIP = {"producttype", "vendor", "tags"}


def build_filter_corpus(product: Dict[str, Any]) -> Tuple[str, Set[str]]:
    parts: List[str] = []
    for field in ("handle", "title", "productType", "vendor"):
        val = product.get(field)
        if isinstance(val, str):
            parts.append(val)
    tags = product.get("tags") or []
    if isinstance(tags, list):
        parts.extend([t for t in tags if isinstance(t, str)])
    options = product.get("options") or []
    if isinstance(options, list):
        for opt in options:
            if not isinstance(opt, dict):
                continue
            values = opt.get("values") or []
            for val in values:
                if val:
                    parts.append(str(val))
    combined = " ".join(parts).lower()
    normalized_text = re.sub(r"[^a-z0-9]+", " ", combined)
    tokens = {tok for tok in normalized_text.split() if tok}
    return normalized_text, tokens


def select_filters_for_product(
    collection_filters: Dict[str, List[str]],
    product: Dict[str, Any],
    derived_filters: Dict[str, List[str]],
) -> Dict[str, List[str]]:
    final_filters: Dict[str, List[str]] = {}
    normalized_text, tokens = build_filter_corpus(product)

    for key, values in derived_filters.items():
        if not values:
            continue
        final_filters[key] = list(values)

    for key, candidates in (collection_filters or {}).items():
        if key in final_filters:
            continue
        matches: List[str] = []
        for candidate in candidates or []:
            cand_str = str(candidate)
            cand_norm = re.sub(r"[^a-z0-9]+", " ", cand_str.lower()).strip()
            if not cand_norm:
                continue
            cand_tokens = {tok for tok in cand_norm.split() if tok}
            if cand_norm in normalized_text or cand_tokens.issubset(tokens):
                matches.append(cand_str)
        if not matches and candidates and len(candidates) == 1:
            matches = [str(candidates[0])]
        if matches:
            final_filters[key] = matches

    return final_filters


def apply_filter_columns(row: Dict[str, Any], filter_values: Dict[str, List[str]]) -> None:
    for raw_key, values in (filter_values or {}).items():
        norm_key = normalize_filter_name(raw_key)
        if norm_key in FILTER_COLUMN_SKIP:
            continue
        if not values:
            continue
        column = f"filter.{norm_key}"
        row[column] = ", ".join(values)


def build_column_order(
    rows: List[Dict[str, Any]],
    *,
    extra_priority: Optional[Sequence[str]] = None,
) -> List[str]:
    all_columns = {key for row in rows for key in row.keys()}
    ordered = list(COLUMN_ORDER_BASE)
    priority: List[str] = []
    if extra_priority:
        for column in extra_priority:
            if column not in ordered and column in all_columns:
                priority.append(column)
    extras = [col for col in all_columns if col not in COLUMN_ORDER_BASE and col not in priority]
    extras.sort()
    return ordered + priority + extras


def unwrap_type(type_info: Optional[Dict[str, Any]]) -> Tuple[Optional[str], Optional[str], Tuple[str, ...]]:
    wrappers: List[str] = []
    current = type_info
    while current and current.get("kind") in {"NON_NULL", "LIST"}:
        wrappers.append(current["kind"])
        current = current.get("ofType")
    kind = current.get("kind") if current else None
    name = current.get("name") if current else None
    return kind, name, tuple(wrappers)


def field_has_required_args(field: Dict[str, Any]) -> bool:
    for arg in field.get("args", []):
        kind, _name, wrappers = unwrap_type(arg.get("type"))
        if "NON_NULL" in wrappers and arg.get("defaultValue") in (None, "null"):
            return True
        if kind == "NON_NULL" and arg.get("defaultValue") in (None, "null"):
            return True
    return False


def write_sheet(
    sheet,
    rows: List[Dict[str, Any]],
    *,
    column_order: Optional[Sequence[str]] = None,
):
    if not rows:
        sheet.append(["No data"])
        return
    if column_order is None:
        columns = sorted({key for row in rows for key in row.keys()})
    else:
        columns = list(column_order)
    sheet.append(columns)
    for row in rows:
        sheet.append([normalize_cell(row.get(column)) for column in columns])


def fetch_collection_html(session: requests.Session, logger: logging.Logger) -> List[Tuple[str, str]]:
    urls: List[str] = []
    if isinstance(COLLECTION_URL, (list, tuple)):
        urls = [url for url in COLLECTION_URL if url]
    elif isinstance(COLLECTION_URL, str) and COLLECTION_URL:
        urls = [COLLECTION_URL]

    if not urls:
        logger.info("No COLLECTION_URL configured; skipping HTML fetch")
        return []

    html_blobs: List[Tuple[str, str]] = []
    for url in urls:
        logger.info("Fetching collection HTML from %s", url)
        try:
            response = session.get(url, timeout=REQUEST_TIMEOUT, verify=False)
            response.raise_for_status()
            html_blobs.append((url, response.text))
        except requests.RequestException as exc:
            logger.warning("Failed to fetch collection HTML from %s: %s", url, exc)

    return html_blobs

def build_products_json_url() -> Optional[str]:
    url = _primary_collection_url()
    if not url:
        return None
    parts = urlsplit(url)
    path = parts.path.rstrip("/") + "/products.json"
    return urlunsplit((parts.scheme, parts.netloc, path, "", ""))


def build_products_json_urls() -> List[str]:
    urls: List[str] = []
    if isinstance(COLLECTION_URL, (list, tuple)):
        for item in COLLECTION_URL:
            if not item:
                continue
            parts = urlsplit(item)
            path = parts.path.rstrip("/") + "/products.json"
            urls.append(urlunsplit((parts.scheme, parts.netloc, path, "", "")))
    else:
        single = build_products_json_url()
        if single:
            urls.append(single)
    return urls


def fetch_collection_titles(session: requests.Session, logger: logging.Logger) -> Dict[str, str]:
    url = _primary_collection_url()
    if not url:
        return {}
    parts = urlsplit(url)
    base = f"{parts.scheme}://{parts.netloc}"
    titles: Dict[str, str] = {}
    page = 1
    while True:
        target = f"{base}/collections.json"
        params = {"page": page, "limit": 250}
        try:
            resp = session.get(target, params=params, timeout=REQUEST_TIMEOUT, verify=False)
        except requests.RequestException as exc:
            logger.debug("Failed to fetch collections.json: %s", exc)
            break
        if not resp.ok:
            break
        try:
            payload = resp.json()
        except ValueError:
            break
        collections = payload.get("collections") if isinstance(payload, dict) else None
        if not collections:
            break
        for coll in collections:
            if not isinstance(coll, dict):
                continue
            handle = coll.get("handle")
            title = coll.get("title")
            if handle and title:
                titles[str(handle)] = str(title)
        if len(collections) < 250:
            break
        page += 1
        time.sleep(0.25)
    if titles:
        logger.info("Discovered %s collections from collections.json", len(titles))
    return titles


def derive_tag_group_key(tag: str) -> str:
    normalized = str(tag or "").strip().lower()
    if not normalized:
        return "misc"
    prefix = normalized
    for separator in ("-", "_", " "):
        if separator in normalized:
            prefix = normalized.split(separator, 1)[0]
            break
    prefix = re.sub(r"[^a-z0-9]+", "_", prefix).strip("_")
    return prefix or "misc"


def group_tags_for_columns(tags: Sequence[str]) -> Dict[str, List[str]]:
    grouped: Dict[str, List[str]] = {}
    seen: Dict[str, set] = defaultdict(set)
    for raw_tag in tags:
        if not isinstance(raw_tag, str):
            continue
        tag = raw_tag.strip()
        if not tag:
            continue
        group_key = derive_tag_group_key(tag)
        column_name = f"tags_group_{group_key}"
        bucket = grouped.setdefault(column_name, [])
        if tag not in seen[column_name]:
            bucket.append(tag)
            seen[column_name].add(tag)
    return grouped


def collect_tag_values(record: Dict[str, Any]) -> List[str]:
    tags: List[str] = []
    for key, value in record.items():
        if "tag" not in key.lower():
            continue
        if isinstance(value, list):
            str_items = [str(item).strip() for item in value if isinstance(item, str) and str(item).strip()]
            tags.extend(str_items)
        elif isinstance(value, str):
            pieces = [part.strip() for part in value.split(",")]
            tags.extend([piece for piece in pieces if piece])
    return tags


def fetch_collection_json(
    session: requests.Session, logger: logging.Logger
) -> Tuple[List[Dict[str, Any]], List[str]]:
    products_json_urls = build_products_json_urls()
    if not products_json_urls:
        logger.info("No collection JSON URL computed; skipping JSON extraction")
        return [], []

    all_products: List[Dict[str, Any]] = []
    for products_json_url in products_json_urls:
        page = 1
        while True:
            params = {"limit": 250, "page": page}
            logger.info("Fetching collection JSON page %s from %s", page, products_json_url)
            try:
                response = session.get(
                    products_json_url, params=params, timeout=REQUEST_TIMEOUT, verify=False
                )
            except requests.RequestException as exc:
                logger.warning("Collection JSON request failed: %s", exc)
                break

            if not response.ok:
                logger.warning(
                    "Collection JSON request returned status %s", response.status_code
                )
                break

            try:
                data = response.json()
            except ValueError:
                logger.warning("Collection JSON response was not valid JSON")
                break

            products = data.get("products") or []
            if not products:
                logger.info("No products found on page %s; stopping pagination", page)
                break

            all_products.extend(products)
            if len(products) < 250:
                break
            page += 1
            time.sleep(0.5)

    logger.info("Collected %s products from collection JSON", len(all_products))
    rows: List[Dict[str, Any]] = []
    tag_group_counts: Counter[str] = Counter()
    for product in all_products:
        if not isinstance(product, dict):
            continue
        product_copy = dict(product)
        tags = list(product_copy.pop("tags", []) or [])
        tag_set = {tag for tag in tags if isinstance(tag, str)}
        for extra_tag in collect_tag_values(product):
            if extra_tag and extra_tag not in tag_set:
                tags.append(extra_tag)
                tag_set.add(extra_tag)
        variants = list(product_copy.get("variants", []) or [])
        product_copy.pop("variants", None)

        options = list(product.get("options") or [])
        option_columns = build_option_columns(options)

        derived_filters = derive_filter_values(product, [])
        filter_values = select_filters_for_product({}, product, derived_filters)

        images = product_copy.get("images") or []
        first_image_src = None
        if isinstance(images, list) and images:
            first_image = images[0]
            if isinstance(first_image, dict):
                first_image_src = (
                    first_image.get("src")
                    or first_image.get("url")
                    or first_image.get("originalSrc")
                )
        if first_image_src:
            product_copy["images"] = [{"src": first_image_src}]
        elif "images" in product_copy:
            product_copy["images"] = []

        flat_product = flatten_record({"product": product_copy})
        base_row = dict(flat_product)
        if tags:
            base_row["product.tags_all"] = ", ".join(tags)
        apply_filter_columns(base_row, filter_values)
        for key, value in option_columns.items():
            base_row[key] = value

        tag_groups = group_tags_for_columns(tags)

        def attach_tag_groups(target_row: Dict[str, Any]) -> None:
            for column_name, tag_values in tag_groups.items():
                joined = ", ".join(tag_values)
                target_row[column_name] = joined
                if joined:
                    tag_group_counts[column_name] += 1

        if not variants:
            row = dict(base_row)
            attach_tag_groups(row)
            finalize_json_row(row, product, None)
            rows.append(row)
            continue

        for variant in variants:
            if not isinstance(variant, dict):
                continue
            variant_copy = dict(variant)
            featured = variant_copy.get("featured_image")
            if isinstance(featured, dict):
                src = featured.get("src") or featured.get("url")
                variant_copy["featured_image"] = {"src": src} if src else {}
            flat_variant = flatten_record({"variant": variant_copy})
            row = dict(base_row)
            row.update(flat_variant)
            attach_tag_groups(row)
            finalize_json_row(row, product, variant)
            rows.append(row)

    if not rows:
        return [], []

    columns = {key for row in rows for key in row.keys()}
    tag_group_columns = [col for col in columns if col.startswith("tags_group_")]
    tag_group_columns.sort(key=lambda col: (-tag_group_counts.get(col, 0), col))

    return rows, tag_group_columns


def extract_searchspring_results(payload: Any) -> List[Dict[str, Any]]:
    if isinstance(payload, dict):
        ss_data = payload.get("ssData")
        if isinstance(ss_data, dict):
            nested_results = extract_searchspring_results(ss_data)
            if nested_results:
                return nested_results
        primary = payload.get("results")
        if isinstance(primary, list):
            return [item for item in primary if isinstance(item, dict)]
        if isinstance(primary, dict):
            aggregated: List[Dict[str, Any]] = []
            for value in primary.values():
                if isinstance(value, list):
                    aggregated.extend([item for item in value if isinstance(item, dict)])
            if aggregated:
                return aggregated
        for key, value in payload.items():
            if isinstance(value, list):
                candidates = [item for item in value if isinstance(item, dict)]
                if candidates:
                    return candidates
    return []


def extract_searchspring_variants(product: Dict[str, Any]) -> List[Dict[str, Any]]:
    variants: List[Dict[str, Any]] = []
    seen_ids: Set[Any] = set()
    candidate_keys = [
        key
        for key in list(product.keys())
        if key.lower()
        in {
            "variants",
            "variant_list",
            "variantlist",
            "skus",
            "sku_list",
            "ss_variants",
        }
    ]
    for key in candidate_keys:
        value = product.pop(key, None)
        if isinstance(value, list):
            for item in value:
                if not isinstance(item, dict):
                    continue
                variant = dict(item)
                vid = variant.get("id")
                if vid is not None and vid in seen_ids:
                    continue
                if vid is not None:
                    seen_ids.add(vid)
                variants.append(variant)
        elif isinstance(value, dict):
            nodes = value.get("nodes") if isinstance(value.get("nodes"), list) else None
            if nodes is not None:
                for item in nodes:
                    if not isinstance(item, dict):
                        continue
                    variant = dict(item)
                    vid = variant.get("id")
                    if vid is not None and vid in seen_ids:
                        continue
                    if vid is not None:
                        seen_ids.add(vid)
                    variants.append(variant)
            else:
                variant = dict(value)
                vid = variant.get("id")
                if vid is not None and vid in seen_ids:
                    continue
                if vid is not None:
                    seen_ids.add(vid)
                variants.append(variant)

    for size_key in ("ss_size_json", "ss_sizes_json", "ss_sizes"):
        raw_value = product.pop(size_key, None)
        if raw_value in (None, ""):
            continue

        parsed: Any = None
        if isinstance(raw_value, list):
            parsed = raw_value
        elif isinstance(raw_value, dict):
            parsed = [raw_value]
        elif isinstance(raw_value, str):
            decoded = html.unescape(raw_value).strip()
            if not decoded:
                continue
            try:
                parsed = json.loads(decoded)
            except ValueError:
                continue

        if not isinstance(parsed, list):
            continue

        for item in parsed:
            if not isinstance(item, dict):
                continue
            lowered = {str(k).lower(): v for k, v in item.items()}
            variant = dict(item)

            label_value = lowered.get("label")
            if label_value not in (None, "") and "option1" not in variant:
                variant["option1"] = label_value

            variant_id = lowered.get("variant_id")
            if variant_id not in (None, "") and "id" not in variant:
                variant["id"] = variant_id

            available_qty = lowered.get("available")
            if available_qty not in (None, "") and "inventory_quantity" not in variant:
                variant["inventory_quantity"] = available_qty

            vid = variant.get("id")
            if vid is not None and vid in seen_ids:
                continue
            if vid is not None:
                seen_ids.add(vid)
            variants.append(variant)

    return variants


def fetch_searchspring_data(
    session: requests.Session, logger: logging.Logger
) -> Tuple[List[Dict[str, Any]], List[str]]:
    if not SEARCHSPRING_SITE_ID or not SEARCHSPRING_URL:
        return [], []

    page = 1
    rows: List[Dict[str, Any]] = []
    tag_group_counts: Counter[str] = Counter()
    base_url = SEARCHSPRING_URL.strip()

    parsed = urlsplit(base_url)
    base_query: Dict[str, Any] = {
        key: value for key, value in parse_qsl(parsed.query, keep_blank_values=True)
    }
    endpoint = urlunsplit((parsed.scheme, parsed.netloc, parsed.path, "", parsed.fragment))
    is_searchspring_host = "searchspring" in (parsed.netloc or "")

    while True:
        params: Dict[str, Any] = dict(base_query)
        params.update(SEARCHSPRING_EXTRA_PARAMS or {})

        if is_searchspring_host:
            if SEARCHSPRING_SITE_ID and not params.get("siteId"):
                params["siteId"] = SEARCHSPRING_SITE_ID
            params.setdefault("resultsFormat", "json")
            params.setdefault("resultsPerPage", 250)
            primary_url = primary_collection_url()
            if primary_url and not params.get("domain"):
                params["domain"] = primary_url
        elif SEARCHSPRING_SITE_ID and not params.get("siteId"):
            params["siteId"] = SEARCHSPRING_SITE_ID

        params["page"] = page

        logger.info("Fetching Searchspring page %s", page)
        try:
            response = session.get(endpoint, params=params, timeout=REQUEST_TIMEOUT, verify=False)
        except requests.RequestException as exc:
            logger.warning("Searchspring request failed on page %s: %s", page, exc)
            break

        if not response.ok:
            logger.warning(
                "Searchspring request returned status %s on page %s", response.status_code, page
            )
            break

        try:
            payload = response.json()
        except ValueError:
            logger.warning("Searchspring response on page %s was not valid JSON", page)
            break

        results = extract_searchspring_results(payload)
        if not results:
            logger.info("Searchspring page %s returned no results; stopping", page)
            break

        for product in results:
            if not isinstance(product, dict):
                continue
            product_copy = dict(product)
            variants = extract_searchspring_variants(product_copy)

            tags = collect_tag_values(product)
            tag_groups = group_tags_for_columns(tags)

            def attach_tag_groups(target_row: Dict[str, Any]) -> None:
                for column_name, tag_values in tag_groups.items():
                    joined = ", ".join(tag_values)
                    target_row[column_name] = joined
                    tag_group_counts[column_name] += 1

            image_candidates = [
                product_copy.get(key)
                for key in (
                    "image",
                    "image_url",
                    "imageUrl",
                    "image_link",
                    "thumbnail",
                    "thumbnail_url",
                    "thumbnailImageUrl",
                )
            ]
            image_src = next((candidate for candidate in image_candidates if candidate), None)
            if image_src:
                product_copy.setdefault("images", [{"src": image_src}])

            flat_product = flatten_record({"product": product_copy})
            base_row = dict(flat_product)
            if tags:
                base_row["product.tags_all"] = ", ".join(tags)

            for key in (
                "product.image",
                "product.image_url",
                "product.imageUrl",
                "product.image_link",
                "product.thumbnail",
                "product.thumbnail_url",
                "product.thumbnailImageUrl",
            ):
                if key in base_row and not base_row.get("product.images[0].src"):
                    base_row["product.images[0].src"] = base_row[key]

            if not variants:
                attach_tag_groups(base_row)
                finalize_json_row(base_row, product, None)
                rows.append(base_row)
                continue

            for variant in variants:
                if not isinstance(variant, dict):
                    continue
                variant_copy = dict(variant)
                if "inventory_quantity" not in variant_copy:
                    for candidate in (
                        "inventory_quantity",
                        "inventoryQuantity",
                        "inventory",
                        "qty",
                        "quantity",
                        "available_quantity",
                    ):
                        value = variant_copy.get(candidate)
                        if value not in (None, ""):
                            variant_copy["inventory_quantity"] = value
                            break
                if "availableForSale" not in variant_copy and isinstance(
                    variant_copy.get("available"), bool
                ):
                    variant_copy["availableForSale"] = variant_copy.get("available")

                flat_variant = flatten_record({"variant": variant_copy})
                row = dict(base_row)
                row.update(flat_variant)
                attach_tag_groups(row)
                finalize_json_row(row, product, variant_copy)
                rows.append(row)

        pagination = payload.get("pagination") if isinstance(payload, dict) else None
        next_page: Optional[int] = None
        if isinstance(pagination, dict):
            candidate = pagination.get("nextPage")
            if isinstance(candidate, int):
                next_page = candidate
            elif isinstance(candidate, str) and candidate.isdigit():
                next_page = int(candidate)
            elif pagination.get("page") and pagination.get("totalPages"):
                try:
                    current_page = int(pagination.get("page"))
                    total_pages = int(pagination.get("totalPages"))
                    if current_page < total_pages:
                        next_page = current_page + 1
                except (TypeError, ValueError):
                    next_page = None

        if next_page:
            page = next_page
            time.sleep(0.5)
            continue

        per_page_param = params.get("resultsPerPage")
        try:
            per_page_int = int(per_page_param)
        except (TypeError, ValueError):
            per_page_int = None
        if per_page_int and len(results) >= per_page_int:
            page += 1
            time.sleep(0.5)
            continue

        break

    if not rows:
        return [], []

    columns = {key for row in rows for key in row.keys()}
    tag_group_columns = [col for col in columns if col.startswith("tags_group_")]
    tag_group_columns.sort(key=lambda col: (-tag_group_counts.get(col, 0), col))

    return rows, tag_group_columns


def make_absolute(url: str, base: Any) -> str:
    if not url:
        return url
    if isinstance(base, (list, tuple)):
        base = base[0] if base else ""
    if not isinstance(base, str) or not base:
        primary = _primary_collection_url()
        base = primary or ""
    return urljoin(base, url)


def discover_tokens(
    session: requests.Session, html_blobs: List[Tuple[str, str]], logger: logging.Logger
) -> List[Tuple[str, str]]:
    tokens: Dict[str, str] = {}
    for base_url, html in html_blobs:
        if not html:
            continue
        for token in set(TOKEN_REGEX.findall(html)):
            tokens.setdefault(token, "collection_html")

        soup = BeautifulSoup(html, "html.parser")
        script_urls: List[str] = []
        for script in soup.find_all("script"):
            src = script.get("src")
            if src:
                absolute = make_absolute(src, base_url)
                script_urls.append(absolute)
                for token in set(TOKEN_REGEX.findall(absolute)):
                    tokens.setdefault(token, f"script_url:{absolute}")
            if script.string:
                for token in set(TOKEN_REGEX.findall(script.string)):
                    tokens.setdefault(token, "inline_script")

        for index, script_url in enumerate(script_urls[:MAX_SCRIPT_FETCHES]):
            logger.info(
                "Fetching script %s/%s for token discovery: %s",
                index + 1,
                min(len(script_urls), MAX_SCRIPT_FETCHES),
                script_url,
            )
            try:
                response = session.get(script_url, timeout=REQUEST_TIMEOUT, verify=False)
            except requests.RequestException as exc:
                logger.debug("Failed to fetch script %s: %s", script_url, exc)
                continue
            if not response.ok:
                logger.debug(
                    "Script %s returned status %s", script_url, response.status_code
                )
                continue
            for token in set(TOKEN_REGEX.findall(response.text)):
                tokens.setdefault(token, f"script_body:{script_url}")

    logger.info("Discovered %s potential tokens", len(tokens))
    return [(token, source) for token, source in tokens.items()]


def determine_graphql_endpoints() -> List[str]:
    endpoints: List[str] = []
    if GRAPHQL:
        endpoints.append(GRAPHQL.strip())
    if MYSHOPIFY:
        base = MYSHOPIFY.rstrip("/") + "/"
        for version in DEFAULT_GRAPHQL_VERSIONS:
            endpoints.append(urljoin(base, version))
    return list(dict.fromkeys(endpoint for endpoint in endpoints if endpoint))


def perform_graphql_request(
    session: requests.Session,
    endpoint: str,
    payload: Dict[str, Any],
    token: Optional[str],
) -> Tuple[Optional[requests.Response], Optional[Dict[str, Any]]]:
    headers = {"Content-Type": "application/json"}
    if token:
        headers["X-Shopify-Storefront-Access-Token"] = token
    try:
        response = session.post(
            endpoint,
            json=payload,
            headers=headers,
            timeout=REQUEST_TIMEOUT,
            verify=False,
        )
    except requests.RequestException:
        return None, None

    try:
        data = response.json()
    except ValueError:
        data = None
    return response, data


class GraphQLIntrospectionError(RuntimeError):
    pass


class GraphQLSchema:
    def __init__(
        self,
        session: requests.Session,
        endpoint: str,
        token: Optional[str],
        logger: logging.Logger,
    ) -> None:
        self.session = session
        self.endpoint = endpoint
        self.token = token
        self.logger = logger
        self._cache: Dict[str, Dict[str, Any]] = {}

    def get_type(self, type_name: Optional[str]) -> Optional[Dict[str, Any]]:
        if not type_name:
            return None
        if type_name in self._cache:
            return self._cache[type_name]

        payload = {"query": INTROSPECTION_QUERY, "variables": {"typeName": type_name}}
        response, data = perform_graphql_request(
            self.session, self.endpoint, payload, self.token
        )
        if response is None or not response.ok:
            raise GraphQLIntrospectionError(
                f"Introspection request failed for {type_name}: {getattr(response, 'status_code', 'error')}"
            )
        type_info = ((data or {}).get("data") or {}).get("__type") if data else None
        if not type_info:
            raise GraphQLIntrospectionError(f"Type {type_name} not found during introspection")
        self._cache[type_name] = type_info
        return type_info


class GraphQLQueryBuilder:
    DEFAULT_CONNECTION_LIMITS: Dict[str, int] = {
        "variants": GRAPHQL_PAGE_SIZE,
        "images": 50,
        "media": 50,
        "collections": 50,
        "components": 100,
        "groupedBy": 100,
        "quantityPriceBreaks": 100,
        "sellingPlanAllocations": 100,
        "sellingPlanGroups": 50,
        "storeAvailability": 100,
    }

    def __init__(
        self,
        session: requests.Session,
        endpoint: str,
        token: Optional[str],
        logger: logging.Logger,
        *,
        max_depth: int = 3,
        forbidden_fields: Optional[Dict[str, Sequence[str]]] = None,
        metafield_identifiers: Optional[Sequence[Tuple[str, str]]] = None,
    ) -> None:
        self.session = session
        self.endpoint = endpoint
        self.token = token
        self.logger = logger
        self.max_depth = max_depth
        self.metafield_identifiers = list(metafield_identifiers or METAFIELD_IDENTIFIERS)
        self.forbidden_fields: Dict[str, Set[str]] = defaultdict(set)
        for parent, names in DEFAULT_FORBIDDEN_FIELDS.items():
            self.forbidden_fields[parent].update(names)
        if forbidden_fields:
            for parent, names in forbidden_fields.items():
                self.forbidden_fields[parent].update(names)
        self.schema = GraphQLSchema(session, endpoint, token, logger)
        self.variant_selection = self._build_type_selection(
            "ProductVariant", max(1, max_depth - 1)
        )
        if not self.variant_selection:
            self.variant_selection = self._build_type_selection("ProductVariant", max_depth)
        if not self.variant_selection:
            raise GraphQLIntrospectionError("Unable to build variant selection set")
        self.product_selection = self._build_type_selection("Product", max_depth)
        if not self.product_selection:
            raise GraphQLIntrospectionError("Unable to build product selection set")
        metafields_selection = self._build_metafields_selection()
        collections_selection = self._build_collections_selection()
        if metafields_selection:
            self.product_selection = "\n".join(
                part for part in [self.product_selection, metafields_selection] if part
            )
        if collections_selection:
            self.product_selection = "\n".join(
                part for part in [self.product_selection, collections_selection] if part
            )
        self.collection_query = self._build_collection_query()
        self.products_query = self._build_products_query()

    def _indent(self, text: str, spaces: int = 2) -> str:
        pad = " " * spaces
        return "\n".join(f"{pad}{line}" if line else pad for line in text.splitlines())

    def _should_include_field(self, parent_type: str, field: Dict[str, Any]) -> bool:
        name = field.get("name")
        if not name or name.startswith("__"):
            return False
        if field_has_required_args(field):
            return False
        if name in self.forbidden_fields.get(parent_type, set()):
            return False
        if parent_type == "ProductVariant" and name == "product":
            return False
        if name in {"sellingPlanGroups", "sellingPlanAllocations"}:
            return False
        return True

    def _build_field_args(self, field: Dict[str, Any]) -> str:
        args = []
        arg_index = {arg.get("name"): arg for arg in field.get("args", [])}
        if "first" in arg_index:
            limit = self.DEFAULT_CONNECTION_LIMITS.get(field.get("name", ""), GRAPHQL_PAGE_SIZE)
            args.append(f"first: {limit}")
        return f"({', '.join(args)})" if args else ""

    def _build_scalar_snapshot(self, type_name: str) -> str:
        type_info = self.schema.get_type(type_name)
        if not type_info:
            return ""
        scalars: List[str] = []
        for field in type_info.get("fields", []):
            if not self._should_include_field(type_name, field):
                continue
            kind, _name, _wrappers = unwrap_type(field.get("type"))
            if kind in {"SCALAR", "ENUM"}:
                scalars.append(field.get("name"))
        return "\n".join(scalars)

    def _build_connection_body(
        self,
        connection_name: str,
        depth: int,
        visited: Sequence[str],
        parent_type: Optional[str],
    ) -> str:
        type_info = self.schema.get_type(connection_name)
        if not type_info or depth <= 0:
            return ""

        lines: List[str] = []
        for field in type_info.get("fields", []):
            fname = field.get("name")
            if fname == "pageInfo":
                lines.append("pageInfo {\n  hasNextPage\n  endCursor\n}")
            elif fname == "edges":
                base_kind, edge_type_name, _ = unwrap_type(field.get("type"))
                if base_kind != "OBJECT" or not edge_type_name:
                    continue
                edge_info = self.schema.get_type(edge_type_name)
                if not edge_info:
                    continue
                edge_lines: List[str] = []
                for edge_field in edge_info.get("fields", []):
                    ename = edge_field.get("name")
                    if ename == "cursor":
                        edge_lines.append("cursor")
                    elif ename == "node":
                        node_kind, node_type_name, _ = unwrap_type(edge_field.get("type"))
                        if node_kind != "OBJECT" or not node_type_name:
                            continue
                        if node_type_name in visited:
                            node_body = self._build_scalar_snapshot(node_type_name)
                        else:
                            node_body = self._build_type_selection(
                                node_type_name,
                                depth - 1,
                                visited=tuple(visited) + (node_type_name,),
                            )
                        if not node_body and node_type_name == parent_type:
                            node_body = self._build_scalar_snapshot(node_type_name)
                        if node_body:
                            edge_lines.append(
                                f"node {{\n{self._indent(node_body)}\n}}"
                            )
                if edge_lines:
                    lines.append(
                        f"edges {{\n{self._indent('\n'.join(edge_lines))}\n}}"
                    )
        return "\n".join(lines)

    def _build_field_selection(
        self,
        parent_type: str,
        field: Dict[str, Any],
        depth: int,
        visited: Sequence[str],
    ) -> Optional[str]:
        name = field.get("name")
        if not self._should_include_field(parent_type, field):
            return None

        if parent_type == "Product" and name == "variants":
            return self._build_variants_field(field)

        base_kind, base_name, wrappers = unwrap_type(field.get("type"))
        if base_kind in {"SCALAR", "ENUM"}:
            return name
        if base_kind == "LIST" and not base_name and wrappers:
            # List ultimately resolves to another type stored deeper in ofType.
            inner = field.get("type", {})
            while inner and inner.get("kind") == "LIST":
                inner = inner.get("ofType")
            base_kind, base_name, _ = unwrap_type(inner)
        if base_kind == "OBJECT" and base_name:
            if base_name in visited or depth <= 0:
                return None
            new_visited = tuple(visited) + (base_name,)
            if base_name.endswith("Connection"):
                body = self._build_connection_body(
                    base_name, depth, new_visited, parent_type=parent_type
                )
            else:
                body = self._build_type_selection(
                    base_name, depth, visited=tuple(visited)
                )
            if not body:
                return None
            args = self._build_field_args(field)
            return f"{name}{args} {{\n{self._indent(body)}\n}}"
        return None

    def _build_metafields_selection(self) -> str:
        if not self.metafield_identifiers:
            return ""

        product_type = self.schema.get_type("Product") or {}
        fields = product_type.get("fields", [])
        metafields_field = next(
            (field for field in fields if field.get("name") == "metafields"), None
        )
        metafield_field = next(
            (field for field in fields if field.get("name") == "metafield"), None
        )

        selection_body = "namespace\nkey\ntype\nvalue"
        identifiers_literal = ", ".join(
            f'{{namespace: "{ns}", key: "{key}"}}'
            for ns, key in self.metafield_identifiers
        )

        if metafields_field and any(arg.get("name") == "identifiers" for arg in metafields_field.get("args", [])):
            return (
                "metafields(identifiers: ["
                + identifiers_literal
                + f"]) {{\n  {selection_body}\n}}"
            )

        if metafield_field and all(
            any(arg.get("name") == name for arg in metafield_field.get("args", []))
            for name in ("namespace", "key")
        ):
            lines: List[str] = []
            for idx, (ns, key) in enumerate(self.metafield_identifiers):
                alias = f"mf_{idx}"
                lines.append(
                    f"{alias}: metafield(namespace: \"{ns}\", key: \"{key}\") {{\n  {selection_body}\n}}"
                )
            return "\n".join(lines)

        self.logger.debug(
            "Metafields selection not added; schema lacks identifiers/namespace+key support"
        )
        return ""

    def _build_collections_selection(self) -> str:
        product_type = self.schema.get_type("Product") or {}
        fields = product_type.get("fields", [])
        collections_field = next(
            (field for field in fields if field.get("name") == "collections"), None
        )
        if not collections_field:
            return ""
        args = self._build_field_args(collections_field)
        body = (
            "edges {\n"
            "  node {\n"
            "    id\n"
            "    handle\n"
            "    title\n"
            "  }\n"
            "}"
        )
        return f"collections{args} {{\n{self._indent(body)}\n}}"

    def _build_variants_field(self, field: Dict[str, Any]) -> Optional[str]:
        args = self._build_field_args(field)
        body = (
            "pageInfo {\n  hasNextPage\n  endCursor\n}\n"
            "edges {\n"
            "  cursor\n"
            "  node {\n"
            f"{self._indent(self.variant_selection, 4)}\n"
            "  }\n"
            "}"
        )
        return f"variants{args} {{\n{self._indent(body)}\n}}"

    def _build_type_selection(
        self,
        type_name: str,
        depth: int,
        *,
        visited: Sequence[str] = (),
    ) -> str:
        if depth <= 0 or type_name in visited:
            return ""
        type_info = self.schema.get_type(type_name)
        if not type_info:
            return ""

        new_visited = tuple(visited) + (type_name,)
        selections: List[str] = []
        for field in type_info.get("fields", []):
            selection = self._build_field_selection(type_name, field, depth - 1, new_visited)
            if selection:
                selections.append(selection)
        return "\n".join(selections)

    def _build_collection_query(self) -> str:
        product_block = self._indent(self.product_selection)
        return (
            "query CollectionProducts($handle: String!, $cursor: String, $pageSize: Int!) {\n"
            "  collection(handle: $handle) {\n"
            "    id\n"
            "    handle\n"
            "    title\n"
            "    products(first: $pageSize, after: $cursor) {\n"
            "      pageInfo {\n"
            "        hasNextPage\n"
            "        endCursor\n"
            "      }\n"
            "      edges {\n"
            "        cursor\n"
            "        node {\n"
            f"{product_block}\n"
            "        }\n"
            "      }\n"
            "    }\n"
            "  }\n"
            "}"
        )

    def _build_products_query(self) -> str:
        product_block = self._indent(self.product_selection)
        return (
            "query ProductsProbe($cursor: String, $pageSize: Int!, $query: String) {\n"
            "  products(first: $pageSize, after: $cursor, query: $query) {\n"
            "    pageInfo {\n"
            "      hasNextPage\n"
            "      endCursor\n"
            "    }\n"
            "    edges {\n"
            "      cursor\n"
            "      node {\n"
            f"{product_block}\n"
            "      }\n"
            "    }\n"
            "  }\n"
            "}"
        )

def probe_graphql_endpoints(
    session: requests.Session,
    endpoints: Sequence[str],
    tokens_with_source: Sequence[Tuple[Optional[str], str]],
    logger: logging.Logger,
    *,
    include_unauthenticated: bool = True,
) -> Tuple[List[Dict[str, Any]], List[str], Dict[str, Set[Optional[str]]]]:
    access_rows: List[Dict[str, Any]] = []
    operational: List[str] = []
    success_map: Dict[str, Set[Optional[str]]] = {endpoint: set() for endpoint in endpoints}

    deduped_tokens: List[Tuple[Optional[str], str]] = []
    seen_keys: Set[Tuple[Optional[str], str]] = set()
    for token, source in tokens_with_source:
        normalized = token.strip() if isinstance(token, str) else token
        normalized = normalized or None
        key = (normalized, source)
        if key in seen_keys:
            continue
        seen_keys.add(key)
        deduped_tokens.append((normalized, source))

    for endpoint in endpoints:
        attempts: List[Tuple[Optional[str], str]] = list(deduped_tokens)
        if include_unauthenticated:
            attempts.append((None, "unauthenticated"))

        for token, token_source in attempts:
            payload = {"query": SHOP_PROBE_QUERY}
            response, data = perform_graphql_request(session, endpoint, payload, token)
            entry: Dict[str, Any] = {
                "endpoint": endpoint,
                "token": token or "",
                "token_source": token_source,
                "status_code": getattr(response, "status_code", ""),
                "ok": bool(response and response.ok),
            }
            if response is None:
                entry["note"] = "request_exception"
            elif not response.ok:
                entry["note"] = f"HTTP_{response.status_code}"
            else:
                shop = ((data or {}).get("data") or {}).get("shop") if data else None
                if shop:
                    entry["shop_name"] = shop.get("name")
                    entry["primary_domain"] = (shop.get("primaryDomain") or {}).get("url")
                    entry["note"] = "success"
                    success_map.setdefault(endpoint, set()).add(token)
                    if token is None and endpoint not in operational:
                        operational.append(endpoint)
                else:
                    errors = (data or {}).get("errors") if data else None
                    entry["note"] = format_error_note(errors) if errors else "no_shop_data"
            access_rows.append(entry)
    return access_rows, operational, success_map


def apply_tag_filter(product: Dict[str, Any]) -> bool:
    if not GRAPHQL_FILTER_TAG:
        return True
    tags = product.get("tags") or []
    lowered = {str(tag).lower() for tag in tags}
    return GRAPHQL_FILTER_TAG.lower() in lowered


def probe_collection_filters(
    session: requests.Session,
    endpoint: str,
    token: Optional[str],
    handle: str,
    logger: logging.Logger,
) -> Dict[str, List[str]]:
    for query in FILTER_PROBE_QUERIES:
        payload = {"query": query, "variables": {"handle": handle}}
        response, data = perform_graphql_request(session, endpoint, payload, token)
        if response is None or not response.ok:
            continue

        filters_block = None
        collection = ((data or {}).get("data") or {}).get("collection") if data else None
        if isinstance(collection, dict):
            products = collection.get("products")
            if isinstance(products, dict):
                filters_block = products.get("filters") or products.get("productFilters")

        if not filters_block:
            continue

        filters: Dict[str, List[str]] = {}
        for fil in filters_block:
            if not isinstance(fil, dict):
                continue
            label = fil.get("label") or fil.get("id")
            values = fil.get("values") or []
            if not label:
                continue
            val_labels: List[str] = []
            for val in values:
                if isinstance(val, dict):
                    if val.get("label"):
                        val_labels.append(str(val.get("label")))
                    elif val.get("id"):
                        val_labels.append(str(val.get("id")))
            if val_labels:
                filters[str(label)] = val_labels
        if filters:
            logger.info("Discovered %s filter groups for collection %s", len(filters), handle)
            return filters

    logger.debug("No filters discovered for collection %s", handle)
    return {}


class ViewJSONEnrichmentState:
    def __init__(self, enabled: bool, fields: Sequence[str], probe_limit: int) -> None:
        self.enabled = bool(enabled)
        self.fields = [field.strip() for field in fields if str(field).strip()]
        self.probe_limit = max(int(probe_limit), 0)
        self.probe_attempts = 0
        self.probe_hits = 0
        self.disabled_after_probe = False
        self.cache: Dict[str, Dict[str, Any]] = {}
        self.warned_urls: Set[str] = set()


def _normalize_view_json_url(online_store_url: str) -> str:
    parsed = urlsplit(online_store_url)
    params = [(k, v) for k, v in parse_qsl(parsed.query, keep_blank_values=True) if k != "view"]
    params.append(("view", "json"))
    return urlunsplit((parsed.scheme, parsed.netloc, parsed.path, urlencode(params), parsed.fragment))


def _lookup_view_json_field(payload: Any, field: str) -> Any:
    current: Any = payload
    for part in field.split("."):
        key = part.strip()
        if not key:
            return None
        if isinstance(current, dict):
            current = current.get(key)
            continue
        if isinstance(current, list) and key.isdigit():
            idx = int(key)
            if 0 <= idx < len(current):
                current = current[idx]
                continue
        return None
    return current


def _extract_view_json_values(payload: Dict[str, Any], fields: Sequence[str]) -> Dict[str, Any]:
    extracted: Dict[str, Any] = {}
    for field in fields:
        value = _lookup_view_json_field(payload, field)
        if value in (None, "", [], {}):
            continue
        if isinstance(value, (dict, list)):
            extracted[f"viewjson.{field}"] = json.dumps(value, ensure_ascii=False)
        else:
            extracted[f"viewjson.{field}"] = str(value)
    return extracted


def _get_view_json_enrichment(
    session: requests.Session,
    product: Dict[str, Any],
    logger: logging.Logger,
    state: Optional[ViewJSONEnrichmentState],
) -> Dict[str, Any]:
    if state is None or not state.enabled or not state.fields:
        return {}
    if state.probe_limit and state.probe_attempts >= state.probe_limit and state.probe_hits == 0:
        state.enabled = False
        state.disabled_after_probe = True
        logger.info(
            "View JSON enrichment disabled after %s probe attempts with no useful fields.",
            state.probe_attempts,
        )
        return {}

    cache_key = str(product.get("id") or product.get("handle") or "")
    if cache_key and cache_key in state.cache:
        return dict(state.cache[cache_key])

    online_store_url = str(product.get("onlineStoreUrl") or "").strip()
    if not online_store_url:
        if cache_key:
            state.cache[cache_key] = {}
        return {}

    view_url = _normalize_view_json_url(online_store_url)
    response: Optional[requests.Response] = None
    try:
        response = session.get(view_url, timeout=REQUEST_TIMEOUT)
        response.raise_for_status()
        payload = response.json()
    except (requests.RequestException, ValueError) as exc:
        if view_url not in state.warned_urls:
            state.warned_urls.add(view_url)
            logger.warning("View JSON enrichment failed for %s -> %s", view_url, exc)
        if cache_key:
            state.cache[cache_key] = {}
        if state.probe_attempts < state.probe_limit:
            state.probe_attempts += 1
        return {}

    if not isinstance(payload, dict):
        if view_url not in state.warned_urls:
            state.warned_urls.add(view_url)
            logger.warning("View JSON enrichment returned non-object JSON for %s", view_url)
        if cache_key:
            state.cache[cache_key] = {}
        if state.probe_attempts < state.probe_limit:
            state.probe_attempts += 1
        return {}

    extracted = _extract_view_json_values(payload, state.fields)
    if state.probe_attempts < state.probe_limit:
        state.probe_attempts += 1
        if extracted:
            state.probe_hits += 1
    if cache_key:
        state.cache[cache_key] = extracted
    return dict(extracted)


def flatten_graphql_product(
    collection_info: Dict[str, Any],
    edge_cursor: str,
    product: Dict[str, Any],
    variant_edge: Optional[Dict[str, Any]],
    *,
    session: Optional[requests.Session] = None,
    logger: Optional[logging.Logger] = None,
    view_json_state: Optional[ViewJSONEnrichmentState] = None,
) -> Dict[str, Any]:
    row: Dict[str, Any] = dict(collection_info)
    collections_handles, collections_titles = extract_collections(product, collection_info)
    if collections_handles:
        row["collections.handle"] = ",".join(collections_handles)
    if collections_titles:
        row["collections.title"] = ",".join(collections_titles)

    metafields = collect_metafields(product)
    if metafields:
        row["metafields"] = json.dumps(metafields)

    product_copy = dict(product)
    for drop_key in (
        "encodedVariantAvailability",
        "encodedVariantExistence",
        "featuredImage",
        "images",
        "media",
        "isGiftCard",
    ):
        product_copy.pop(drop_key, None)
    product_copy.pop("collections", None)
    variants = product_copy.pop("variants", None)
    flat_product = flatten_record({"product": product_copy})
    row.update(flat_product)

    if session is not None and logger is not None:
        row.update(_get_view_json_enrichment(session, product, logger, view_json_state))

    option_columns = build_option_columns(product.get("options") or [])
    for key, value in option_columns.items():
        row[key] = value

    collection_filters = collection_info.get("collection_filters") if isinstance(collection_info, dict) else {}
    derived_filters = derive_filter_values(product, metafields)
    filter_values = select_filters_for_product(collection_filters, product, derived_filters)
    apply_filter_columns(row, filter_values)

    if variant_edge is None:
        finalize_storefront_row(row, product, None)
        return row

    variant = dict(variant_edge.get("node") or {})
    for drop_key in ("quantityRule", "image"):
        variant.pop(drop_key, None)
    flat_variant = flatten_record({"variant": variant})
    row.update(flat_variant)
    finalize_storefront_row(row, product, variant)
    return row


def collect_storefront_from_collections(
    session: requests.Session,
    endpoint: str,
    token: Optional[str],
    logger: logging.Logger,
) -> Tuple[List[Dict[str, Any]], Optional[int], str]:
    forbidden: Dict[str, Set[str]] = defaultdict(set)
    for parent, names in DEFAULT_FORBIDDEN_FIELDS.items():
        forbidden[parent].update(names)

    first_status: Optional[int] = None
    view_json_state = ViewJSONEnrichmentState(
        VIEW_JSON_ENRICHMENT_ENABLED,
        VIEW_JSON_FIELDS,
        VIEW_JSON_PROBE_LIMIT,
    )

    while True:
        try:
            builder = GraphQLQueryBuilder(
                session,
                endpoint,
                token,
                logger,
                forbidden_fields=forbidden,
                metafield_identifiers=METAFIELD_IDENTIFIERS,
            )
        except GraphQLIntrospectionError as exc:
            logger.debug("Unable to build collection query for %s: %s", endpoint, exc)
            return [], None, "builder_error"

        query_text = builder.collection_query
        rows: List[Dict[str, Any]] = []
        need_retry = False
        newly_blocked: Dict[str, Set[str]] = defaultdict(set)
        collection_filters_cache: Dict[str, Dict[str, List[str]]] = {}

        for handle in STOREFRONT_COLLECTION_HANDLES:
            if handle not in collection_filters_cache:
                collection_filters_cache[handle] = probe_collection_filters(
                    session, endpoint, token, handle, logger
                )
            handle_filters = collection_filters_cache.get(handle) or {}
            cursor: Optional[str] = None
            while True:
                payload = {
                    "query": query_text,
                    "variables": {
                        "handle": handle,
                        "cursor": cursor,
                        "pageSize": GRAPHQL_PAGE_SIZE,
                    },
                }
                response, data = perform_graphql_request(
                    session, endpoint, payload, token
                )
                if first_status is None and response is not None:
                    first_status = response.status_code
                if response is None:
                    return [], first_status, "request_exception"
                if not response.ok:
                    return [], first_status, f"HTTP_{response.status_code}"

                payload_data = (data or {}).get("data") if data else None
                collection = (payload_data or {}).get("collection") if payload_data else None
                errors = (data or {}).get("errors") if data else None

                if not collection:
                    if errors:
                        unrecoverable = True
                        for error in errors:
                            path = error.get("path") or []
                            field_name = extract_field_from_error_path(path)
                            if not field_name:
                                continue
                            unrecoverable = False
                            target_type = infer_error_target_type(path)
                            if field_name not in forbidden[target_type]:
                                forbidden[target_type].add(field_name)
                                newly_blocked[target_type].add(field_name)
                                need_retry = True
                        if unrecoverable:
                            return [], first_status, format_error_note(errors)
                        break
                    return [], first_status, "no_collection_data"

                if errors:
                    logger.debug(
                        "Collection query returned %s errors for handle %s on %s",
                        len(errors),
                        handle,
                        endpoint,
                    )
                    new_field_added = False
                    for error in errors:
                        path = error.get("path") or []
                        field_name = extract_field_from_error_path(path)
                        if not field_name:
                            continue
                        target_type = infer_error_target_type(path)
                        if field_name not in forbidden[target_type]:
                            forbidden[target_type].add(field_name)
                            newly_blocked[target_type].add(field_name)
                            need_retry = True
                            new_field_added = True
                    if need_retry:
                        break
                    if not new_field_added:
                        return [], first_status, format_error_note(errors)

                collection_info = {
                    "collection_id": collection.get("id"),
                    "collection_handle": collection.get("handle"),
                    "collection_title": collection.get("title"),
                    "collection_filters": handle_filters,
                }
                products_connection = collection.get("products") or {}
                edges: Iterable[Dict[str, Any]] = products_connection.get("edges") or []
                for edge in edges:
                    product = edge.get("node") or {}
                    if not apply_tag_filter(product):
                        continue
                    variants_connection = product.get("variants") or {}
                    variant_entries = extract_graphql_variant_entries(
                        variants_connection
                    )
                    if not variant_entries:
                        rows.append(
                            flatten_graphql_product(
                                collection_info,
                                edge.get("cursor", ""),
                                product,
                                None,
                                session=session,
                                logger=logger,
                                view_json_state=view_json_state,
                            )
                        )
                    else:
                        for variant_edge in variant_entries:
                            rows.append(
                                flatten_graphql_product(
                                    collection_info,
                                    edge.get("cursor", ""),
                                    product,
                                    variant_edge,
                                    session=session,
                                    logger=logger,
                                    view_json_state=view_json_state,
                                )
                            )
                page_info = products_connection.get("pageInfo") or {}
                if page_info.get("hasNextPage"):
                    cursor = page_info.get("endCursor")
                    logger.info(
                        "Collection %s has additional Storefront pages; continuing",
                        handle,
                    )
                    time.sleep(0.5)
                else:
                    break

            if need_retry:
                break

        if need_retry:
            blocked_summary = {
                parent: sorted(fields)
                for parent, fields in newly_blocked.items()
                if fields
            }
            if blocked_summary:
                logger.info(
                    "Retrying collection query without restricted fields: %s",
                    blocked_summary,
                )
            else:
                logger.debug(
                    "Encountered errors but no removable fields; aborting with failure"
                )
                return [], first_status, "errors"
            continue

        note = "success" if rows else "no_rows"
        return rows, first_status, note


def build_product_query_string() -> Optional[str]:
    query_parts: List[str] = []
    if GRAPHQL_FILTER_TAG:
        if " " in GRAPHQL_FILTER_TAG:
            query_parts.append(f'tag:"{GRAPHQL_FILTER_TAG}"')
        else:
            query_parts.append(f"tag:{GRAPHQL_FILTER_TAG}")
    for handle in STOREFRONT_COLLECTION_HANDLES:
        query_parts.append(f"collection:{handle}")
    return " ".join(query_parts) if query_parts else None


def collect_storefront_from_products(
    session: requests.Session,
    endpoint: str,
    token: Optional[str],
    logger: logging.Logger,
) -> Tuple[List[Dict[str, Any]], Optional[int], str]:
    rows: List[Dict[str, Any]] = []
    cursor: Optional[str] = None
    query_string = build_product_query_string()
    first_status: Optional[int] = None
    view_json_state = ViewJSONEnrichmentState(
        VIEW_JSON_ENRICHMENT_ENABLED,
        VIEW_JSON_FIELDS,
        VIEW_JSON_PROBE_LIMIT,
    )
    forbidden: Dict[str, Set[str]] = defaultdict(set)
    for parent, names in DEFAULT_FORBIDDEN_FIELDS.items():
        forbidden[parent].update(names)

    try:
        builder = GraphQLQueryBuilder(
            session,
            endpoint,
            token,
            logger,
            forbidden_fields=forbidden,
            metafield_identifiers=METAFIELD_IDENTIFIERS,
        )
    except GraphQLIntrospectionError as exc:
        logger.debug("Unable to build products query for %s: %s", endpoint, exc)
        return [], None, "builder_error"

    query_text = builder.products_query

    while True:
        payload = {
            "query": query_text,
            "variables": {
                "cursor": cursor,
                "pageSize": GRAPHQL_PAGE_SIZE,
                "query": query_string,
            },
        }
        response, data = perform_graphql_request(session, endpoint, payload, token)
        if first_status is None and response is not None:
            first_status = response.status_code
        if response is None:
            return [], first_status, "request_exception"
        if not response.ok:
            return [], first_status, f"HTTP_{response.status_code}"

        products_connection = ((data or {}).get("data") or {}).get("products") if data else None
        if not products_connection:
            errors = (data or {}).get("errors") if data else None
            if errors:
                return [], first_status, format_error_note(errors)
            return [], first_status, "no_products_data"

        errors = (data or {}).get("errors") if data else None
        if errors:
            logger.debug(
                "Products query returned %s errors on %s",
                len(errors),
                endpoint,
            )

        edges: Iterable[Dict[str, Any]] = products_connection.get("edges") or []
        for edge in edges:
            product = edge.get("node") or {}
            if not apply_tag_filter(product):
                continue
            variants_connection = product.get("variants") or {}
            variant_entries = extract_graphql_variant_entries(variants_connection)
            if not variant_entries:
                rows.append(
                    flatten_graphql_product(
                        {"collection_handle": ""},
                        edge.get("cursor", ""),
                        product,
                        None,
                        session=session,
                        logger=logger,
                        view_json_state=view_json_state,
                    )
                )
            else:
                for variant_edge in variant_entries:
                    rows.append(
                        flatten_graphql_product(
                            {"collection_handle": ""},
                            edge.get("cursor", ""),
                            product,
                            variant_edge,
                            session=session,
                            logger=logger,
                            view_json_state=view_json_state,
                        )
                    )
        page_info = products_connection.get("pageInfo") or {}
        if page_info.get("hasNextPage"):
            cursor = page_info.get("endCursor")
            logger.info("Products query returned more pages; continuing")
            time.sleep(0.5)
        else:
            break
    note = "success" if rows else "no_rows"
    return rows, first_status, note


def fallback_collect_storefront(
    session: requests.Session,
    endpoints: Sequence[str],
    logger: logging.Logger,
) -> Tuple[List[Dict[str, Any]], Optional[Dict[str, Any]]]:
    if not endpoints:
        return [], None

    if STOREFRONT_COLLECTION_HANDLES:
        return fallback_collect_from_collections(session, endpoints, logger)
    return fallback_collect_from_products(session, endpoints, logger)


def fallback_collect_from_collections(
    session: requests.Session,
    endpoints: Sequence[str],
    logger: logging.Logger,
) -> Tuple[List[Dict[str, Any]], Optional[Dict[str, Any]]]:
    for endpoint in endpoints:
        logger.info(
            "Attempting unauthenticated Storefront fallback via %s", endpoint
        )
        rows: List[Dict[str, Any]] = []
        first_status: Optional[int] = None
        success = True
        filters_cache: Dict[str, Dict[str, List[str]]] = {}
        view_json_state = ViewJSONEnrichmentState(
            VIEW_JSON_ENRICHMENT_ENABLED,
            VIEW_JSON_FIELDS,
            VIEW_JSON_PROBE_LIMIT,
        )

        for handle in STOREFRONT_COLLECTION_HANDLES:
            if handle not in filters_cache:
                filters_cache[handle] = probe_collection_filters(
                    session, endpoint, None, handle, logger
                )
            handle_filters = filters_cache.get(handle) or {}
            cursor: Optional[str] = None
            while True:
                payload = {
                    "query": FALLBACK_COLLECTION_QUERY,
                    "variables": {
                        "handle": handle,
                        "cursor": cursor,
                        "pageSize": GRAPHQL_PAGE_SIZE,
                    },
                }
                response, data = perform_graphql_request(
                    session, endpoint, payload, token=None
                )
                if first_status is None and response is not None:
                    first_status = response.status_code
                if response is None or not response.ok:
                    logger.debug(
                        "Fallback Storefront request failed for %s (handle=%s): %s",
                        endpoint,
                        handle,
                        getattr(response, "status_code", "error"),
                    )
                    success = False
                    break

                collection = (
                    ((data or {}).get("data") or {}).get("collection") if data else None
                )
                if not collection:
                    logger.debug(
                        "Fallback Storefront returned no collection data for handle '%s'",
                        handle,
                    )
                    success = False
                    break

                collection_info = {
                    "collection_id": collection.get("id"),
                    "collection_handle": collection.get("handle"),
                    "collection_title": collection.get("title"),
                    "collection_filters": handle_filters,
                }

                products_connection = collection.get("products") or {}
                edges: Iterable[Dict[str, Any]] = products_connection.get("edges") or []
                for edge in edges:
                    product = edge.get("node") or {}
                    if not apply_tag_filter(product):
                        continue
                    variants_connection = product.get("variants") or {}
                    variant_entries = extract_graphql_variant_entries(
                        variants_connection
                    )
                    if not variant_entries:
                        rows.append(
                            flatten_graphql_product(
                                collection_info,
                                edge.get("cursor", ""),
                                product,
                                None,
                                session=session,
                                logger=logger,
                                view_json_state=view_json_state,
                            )
                        )
                    else:
                        for variant_edge in variant_entries:
                            rows.append(
                                flatten_graphql_product(
                                    collection_info,
                                    edge.get("cursor", ""),
                                    product,
                                    variant_edge,
                                    session=session,
                                    logger=logger,
                                    view_json_state=view_json_state,
                                )
                            )

                page_info = products_connection.get("pageInfo") or {}
                if page_info.get("hasNextPage"):
                    cursor = page_info.get("endCursor")
                    time.sleep(0.5)
                else:
                    break

            if not success:
                break

        if rows and success:
            access_entry = {
                "endpoint": endpoint,
                "token": "",
                "token_source": "fallback_unauthenticated",
                "status_code": first_status or "",
                "ok": True,
                "note": "fallback_success",
            }
            return rows, access_entry

    return [], None


def fallback_collect_from_products(
    session: requests.Session,
    endpoints: Sequence[str],
    logger: logging.Logger,
) -> Tuple[List[Dict[str, Any]], Optional[Dict[str, Any]]]:
    query_string = build_product_query_string()
    for endpoint in endpoints:
        logger.info(
            "Attempting unauthenticated Storefront products fallback via %s",
            endpoint,
        )
        rows: List[Dict[str, Any]] = []
        cursor: Optional[str] = None
        first_status: Optional[int] = None
        fallback_query = build_fallback_products_query()
        view_json_state = ViewJSONEnrichmentState(
            VIEW_JSON_ENRICHMENT_ENABLED,
            VIEW_JSON_FIELDS,
            VIEW_JSON_PROBE_LIMIT,
        )

        while True:
            payload = {
                "query": fallback_query,
                "variables": {
                    "cursor": cursor,
                    "pageSize": GRAPHQL_PAGE_SIZE,
                    "query": query_string,
                },
            }
            response, data = perform_graphql_request(
                session, endpoint, payload, token=None
            )
            if first_status is None and response is not None:
                first_status = response.status_code
            if response is None or not response.ok:
                logger.debug(
                    "Fallback products request failed for %s: %s",
                    endpoint,
                    getattr(response, "status_code", "error"),
                )
                rows = []
                break

            products_connection = (
                ((data or {}).get("data") or {}).get("products") if data else None
            )
            if not products_connection:
                logger.debug(
                    "Fallback products query returned no data for endpoint %s",
                    endpoint,
                )
                rows = []
                break

            edges: Iterable[Dict[str, Any]] = products_connection.get("edges") or []
            for edge in edges:
                product = edge.get("node") or {}
                if not apply_tag_filter(product):
                    continue
                variants_connection = product.get("variants") or {}
                variant_entries = extract_graphql_variant_entries(variants_connection)
                if not variant_entries:
                    rows.append(
                        flatten_graphql_product(
                            {"collection_handle": ""},
                            edge.get("cursor", ""),
                            product,
                            None,
                            session=session,
                            logger=logger,
                            view_json_state=view_json_state,
                        )
                    )
                else:
                    for variant_edge in variant_entries:
                        rows.append(
                            flatten_graphql_product(
                                {"collection_handle": ""},
                                edge.get("cursor", ""),
                                product,
                                variant_edge,
                                session=session,
                                logger=logger,
                                view_json_state=view_json_state,
                            )
                        )

            page_info = products_connection.get("pageInfo") or {}
            if page_info.get("hasNextPage"):
                cursor = page_info.get("endCursor")
                time.sleep(0.5)
            else:
                break

        if rows:
            access_entry = {
                "endpoint": endpoint,
                "token": "",
                "token_source": "fallback_unauthenticated",
                "status_code": first_status or "",
                "ok": True,
                "note": "fallback_success",
            }
            return rows, access_entry

    return [], None


def gather_storefront_data(
    session: requests.Session,
    html_blobs: List[Tuple[str, str]],
    logger: logging.Logger,
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    endpoints = determine_graphql_endpoints()
    if not endpoints:
        logger.info("No GraphQL endpoints configured; skipping Storefront extraction")
        return [], []

    provided_tokens = [
        (token, "provided_token") for token in normalize_tokens(X_SHOPIFY_STOREFRONT_ACCESS_TOKEN)
    ]
    access_rows, _operational, success_map = probe_graphql_endpoints(
        session, endpoints, provided_tokens, logger
    )
    endpoints_to_use = list(dict.fromkeys(endpoints))
    token_success_map: Dict[str, Set[Optional[str]]] = {
        endpoint: set(tokens) for endpoint, tokens in success_map.items()
    }

    def attempt_with_token(
        token: Optional[str], source: str
    ) -> Optional[List[Dict[str, Any]]]:
        endpoints_iterable = endpoints_to_use
        if token is not None:
            eligible = [
                endpoint
                for endpoint in endpoints_to_use
                if token in token_success_map.get(endpoint, set())
            ]
            if not eligible:
                logger.debug(
                    "Skipping token %s entirely; no endpoints reported a successful probe",
                    token,
                )
                return None
            endpoints_iterable = eligible

        for endpoint in endpoints_iterable:
            if STOREFRONT_COLLECTION_HANDLES:
                rows, status, note = collect_storefront_from_collections(
                    session, endpoint, token, logger
                )
            else:
                rows, status, note = collect_storefront_from_products(
                    session, endpoint, token, logger
                )

            access_rows.append(
                {
                    "endpoint": endpoint,
                    "token": token or "",
                    "token_source": source,
                    "status_code": status or "",
                    "ok": note == "success",
                    "note": note,
                }
            )

            if rows:
                logger.info(
                    "Storefront extraction succeeded with endpoint %s using token source %s",
                    endpoint,
                    source,
                )
                return rows
        return None

    attempted_sources: set = set()

    if provided_tokens:
        for provided_token, source in provided_tokens:
            if (provided_token, source) in attempted_sources:
                continue
            result = attempt_with_token(provided_token, source)
            attempted_sources.add((provided_token, source))
            if result:
                return result, access_rows

    discovered_tokens: List[Tuple[Optional[str], str]] = []
    if html_blobs:
        new_tokens = discover_tokens(session, html_blobs, logger)
        if new_tokens:
            discovery_rows, _ops, discovery_success = probe_graphql_endpoints(
                session,
                endpoints_to_use,
                new_tokens,
                logger,
                include_unauthenticated=False,
            )
            access_rows.extend(discovery_rows)
            for endpoint, tokens in discovery_success.items():
                if tokens:
                    token_success_map.setdefault(endpoint, set()).update(tokens)
        discovered_tokens.extend(new_tokens)

    for token, source in discovered_tokens:
        if (token, source) in attempted_sources:
            continue
        result = attempt_with_token(token, source)
        attempted_sources.add((token, source))
        if result:
            return result, access_rows

    if (None, "no_token") not in attempted_sources:
        result = attempt_with_token(None, "no_token")
        attempted_sources.add((None, "no_token"))
        if result:
            return result, access_rows

    fallback_rows, fallback_entry = fallback_collect_storefront(
        session, endpoints_to_use, logger
    )
    if fallback_rows:
        logger.info("Storefront fallback succeeded without a token")
        if fallback_entry:
            access_rows.append(fallback_entry)
        return fallback_rows, access_rows

    logger.warning("Storefront extraction did not return any rows")
    return [], access_rows


def export_workbook(
    json_rows: List[Dict[str, Any]],
    storefront_rows: List[Dict[str, Any]],
    access_rows: List[Dict[str, Any]],
    searchspring_rows: List[Dict[str, Any]],
    *,
    json_priority_columns: Optional[Sequence[str]] = None,
    searchspring_priority_columns: Optional[Sequence[str]] = None,
) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "JSON"
    json_columns = (
        build_column_order(json_rows, extra_priority=json_priority_columns)
        if json_rows
        else list(COLUMN_ORDER_BASE)
    )
    write_sheet(sheet, json_rows, column_order=json_columns)

    searchspring_sheet = workbook.create_sheet("SearchSpring")
    searchspring_columns = (
        build_column_order(
            searchspring_rows, extra_priority=searchspring_priority_columns
        )
        if searchspring_rows
        else list(COLUMN_ORDER_BASE)
    )
    write_sheet(searchspring_sheet, searchspring_rows, column_order=searchspring_columns)

    storefront_sheet = workbook.create_sheet("Storefront")
    storefront_columns = (
        build_column_order(storefront_rows) if storefront_rows else list(COLUMN_ORDER_BASE)
    )
    write_sheet(storefront_sheet, storefront_rows, column_order=storefront_columns)

    access_sheet = workbook.create_sheet("Storefront_access")
    write_sheet(access_sheet, access_rows)

    timestamp = datetime.now(timezone.utc).astimezone().strftime("%Y-%m-%d_%H-%M-%S")
    output_path = OUTPUT_DIR / f"{BRAND_SLUG}_probe_{timestamp}.xlsx"
    workbook.save(output_path)
    return output_path


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------
def main() -> None:
    parser = argparse.ArgumentParser(description="Retail data probe")
    parser.add_argument(
        "--metafield",
        dest="metafields",
        help="Comma-separated list of namespace:key metafields to request via Storefront API",
        default="",
    )
    args = parser.parse_args()

    global METAFIELD_IDENTIFIERS
    METAFIELD_IDENTIFIERS = parse_metafield_identifiers(args.metafields)

    logger = configure_logging()
    session = build_session()
    global COLLECTION_TITLE_MAP
    COLLECTION_TITLE_MAP = fetch_collection_titles(session, logger)
    html_blobs = fetch_collection_html(session, logger)
    json_rows, tag_group_columns = fetch_collection_json(session, logger)
    if SEARCHSPRING_SITE_ID and SEARCHSPRING_URL:
        searchspring_rows, searchspring_tag_columns = fetch_searchspring_data(session, logger)
    else:
        logger.info("Searchspring configuration missing; skipping Searchspring extraction")
        searchspring_rows, searchspring_tag_columns = [], []
    storefront_rows, access_rows = gather_storefront_data(session, html_blobs, logger)
    output_path = export_workbook(
        json_rows,
        storefront_rows,
        access_rows,
        searchspring_rows,
        json_priority_columns=tag_group_columns,
        searchspring_priority_columns=searchspring_tag_columns,
    )
    logger.info("Workbook written to %s", output_path.as_posix())


if __name__ == "__main__":
    main()
